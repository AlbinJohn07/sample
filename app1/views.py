
from functools import cache
from typing import OrderedDict
from django.shortcuts import get_list_or_404, render
from .serializers import RegisterSerializer
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
#(2) Views for the User Login Views for the User Login
from django.shortcuts import get_list_or_404, render
from .serializers import RegisterSerializer
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
# #Create your views here.
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth import login
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework.exceptions import ValidationError
from .serializers import LoginSerializer

from django.shortcuts import render

def dojo_app(request):
    return render (request,'index.html')

class LoginAPIView(APIView):

    """
    User Login API View
    """
    def post(self, request):
        print(request.data)
        serializer = LoginSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            user = serializer.validated_data['user']
            
            # Log the user in
            login(request, user)
            
            # Generate JWT tokens
            refresh = RefreshToken.for_user(user)
            access_token = str(refresh.access_token)
            print(user.role)
            
            # Return user info & tokens
            return Response({
                'message': 'Login successful',
                'access_token': access_token,
                'refresh_token': str(refresh),
                'user': {
                    'email': user.email,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'employeeid': user.employeeid,
                    'role': user.role,
                    'hq': user.hq,
                    'factory': user.factory,
                    'department': user.department,
                    'status': user.status
                }
            }, status=status.HTTP_200_OK)
        
        # Extract error message
        error_message = "Authentication failed"
        if serializer.errors:
            for field, errors in serializer.errors.items():
                if errors:
                    if isinstance(errors, list) and errors:
                        error_message = errors[0]
                    else:
                        error_message = str(errors)
                    break
        
        return Response({
            'error': True,
            'message': error_message
        }, status=status.HTTP_400_BAD_REQUEST)
    





#(1) Views for the User Register

from django.db import IntegrityError
from django.shortcuts import render
from .serializers import RegisterSerializer
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from .models import User

class RegisterView(generics.GenericAPIView):
    serializer_class = RegisterSerializer

    def post(self, request):
        user_data = request.data
        serializer = self.serializer_class(data=user_data)

        try:
            serializer.is_valid(raise_exception=True)

            # Check if user already exists by email
            if User.objects.filter(email=user_data.get("email")).exists():
                return Response({
                    "message": "Registration failed",
                    "errors": {"email": "This email is already registered."}
                }, status=status.HTTP_400_BAD_REQUEST)

            # Check if employee ID already exists
            if User.objects.filter(employeeid=user_data.get("employeeid")).exists():
                return Response({
                    "message": "Registration failed",
                    "errors": {"employeeid": "This employee ID is already in use."}
                }, status=status.HTTP_400_BAD_REQUEST)

            # Save the user
            serializer.save()

            return Response({
                "message": "User registered successfully!"
            }, status=status.HTTP_201_CREATED)

        except ValidationError as e:
            # Handle specific validation errors
            return Response({
                "message": "Validation failed",
                "errors": e.detail
            }, status=status.HTTP_400_BAD_REQUEST)

        except IntegrityError:
            # Handle database integrity errors (like duplicate entries)
            return Response({
                "message": "Database error",
                "errors": {"detail": "Duplicate entry or constraint violation."}
            }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            # Handle unexpected errors
            return Response({
                "message": "Unexpected error occurred",
                "errors": {"detail": str(e)}
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


#(3) Views for the User Logout

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from .serializers import LogoutSerializer

class LogoutAPIView(APIView):
    """
    User Logout API View
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = LogoutSerializer(data=request.data)
        if serializer.is_valid():
            refresh_token = serializer.validated_data["refresh_token"]

            try:
                token = RefreshToken(refresh_token)
                token.blacklist()  # Blacklist the refresh token
                return Response({"message": "Logout successful"}, status=status.HTTP_200_OK)
            except Exception as e:
                return Response({"error": "Invalid token"}, status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    


















from django.shortcuts import render

# Create your views here.
# from rest_framework import viewsets, status
# from rest_framework.response import Response
# from .models import EmployeeMaster
# from .serializers import EmployeeSerializer

# class EmployeeViewSet(viewsets.ModelViewSet):
#     queryset = EmployeeMaster.objects.all()
#     serializer_class = EmployeeSerializer

#     def create(self, request, *args, **kwargs):
#         serializer = self.get_serializer(data=request.data)
#         if not serializer.is_valid():
#             # Print validation errors to console for debugging
#             print("Validation errors:", serializer.errors)
#             # Return detailed error response
#             return Response({
#                 'status': 'error',
#                 'errors': serializer.errors,
#                 'message': 'Invalid data provided'
#             }, status=status.HTTP_400_BAD_REQUEST)
        
#         try:
#             self.perform_create(serializer)
#             headers = self.get_success_headers(serializer.data)
#             return Response({
#                 'status': 'success',
#                 'data': serializer.data,
#                 'message': 'Employee created successfully'
#             }, status=status.HTTP_201_CREATED, headers=headers)
#         except Exception as e:
#             # Print exception to console
#             print("Exception occurred:", str(e))
#             return Response({
#                 'status': 'error',
#                 'message': str(e),
#                 'details': 'An error occurred while creating employee'
#             }, status=status.HTTP_400_BAD_REQUEST)

#@A0005
# this view is for the employee master and updated 26/08/2025
# Create your views here.
from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.http import Http404
from .models import EmployeeMaster
from .serializers import EmployeeSerializer

class EmployeeViewSet(viewsets.ModelViewSet):
  
    queryset = EmployeeMaster.objects.filter(is_active=True)  # Default queryset for router
    serializer_class = EmployeeSerializer

    def get_queryset(self):
     
        if self.action in ['destroy', 'mark_read', 'restore']:
            return EmployeeMaster.objects.all()  # Include inactive for these actions
        return EmployeeMaster.objects.filter(is_active=True)  # Only active for normal operations
#@A0006
    # ----------------- CREATE -----------------
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            print("Validation errors:", serializer.errors)
            return Response({
                'status': 'error',
                'errors': serializer.errors,
                'message': 'Invalid data provided'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            return Response({
                'status': 'success',
                'data': serializer.data,
                'message': 'Employee created successfully'
            }, status=status.HTTP_201_CREATED, headers=headers)
        except Exception as e:
            print("Exception occurred:", str(e))
            return Response({
                'status': 'error',
                'message': str(e),
                'details': 'An error occurred while creating employee'
            }, status=status.HTTP_400_BAD_REQUEST)

    # # ----------------- DESTROY (soft delete) -----------------
    # def destroy(self, request, *args, **kwargs):
       
    #     try:
    #         instance = self.get_object()
    #     except Http404:
    #         return Response({
    #             "status": "error",
    #             "message": "Employee not found"
    #         }, status=status.HTTP_404_NOT_FOUND)

    #     # Check if already inactive
    #     if not instance.is_active:
    #         return Response({
    #             "status": "success",
    #             "message": f"Employee {instance.name} is already removed"
    #         }, status=status.HTTP_200_OK)

    #     # Soft delete
    #     instance.is_active = False
    #     instance.save()

    #     return Response({
    #         "status": "success",
    #         "message": f"Employee {instance.name} has been removed from the active list"
    #     }, status=status.HTTP_200_OK)

    @action(detail=True, methods=["patch"], url_path="mark_read")
    def mark_read(self, request, pk=None):
     
        try:
            employee = self.get_object()
        except Http404:
            return Response({
                "status": "error",
                "message": "Employee not found"
            }, status=status.HTTP_404_NOT_FOUND)

        employee.read = True
        employee.save()

        return Response({
            "status": "success",
            "message": f"Employee {employee.id} marked as read"
        }, status=status.HTTP_200_OK)

    # @action(detail=True, methods=["patch"], url_path="restore")
    # def restore(self, request, pk=None):
       
    #     try:
    #         employee = self.get_object()
    #     except Http404:
    #         return Response({
    #             "status": "error",
    #             "message": "Employee not found"
    #         }, status=status.HTTP_404_NOT_FOUND)

    #     if employee.is_active:
    #         return Response({
    #             "status": "success",
    #             "message": f"Employee {employee.name} is already active"
    #         }, status=status.HTTP_200_OK)

    #     employee.is_active = True
    #     employee.save()

    #     return Response({
    #         "status": "success",
    #         "message": f"Employee {employee.name} has been restored to the active list"
    #     }, status=status.HTTP_200_OK)

from rest_framework import viewsets
from .models import (
     Station, OperatorSkill,
    TrainingTopic, OperatorTraining, MonthlyAssignment
)
from .serializers import (
     StationSerializer, OperatorSkillSerializer,
    TrainingTopicSerializer, OperatorTrainingSerializer, MonthlyAssignmentSerializer
)


# class OperatorViewSet(viewsets.ModelViewSet):
#     queryset = Operator.objects.all()
#     serializer_class = OperatorSerializer


class StationViewSet(viewsets.ModelViewSet):
    queryset = Station.objects.all()
    serializer_class = StationSerializer


class OperatorSkillViewSet(viewsets.ModelViewSet):
    queryset = OperatorSkill.objects.all()
    serializer_class = OperatorSkillSerializer


class TrainingTopicViewSet(viewsets.ModelViewSet):
    queryset = TrainingTopic.objects.all()
    serializer_class = TrainingTopicSerializer


class OperatorTrainingViewSet(viewsets.ModelViewSet):
    queryset = OperatorTraining.objects.all()
    serializer_class = OperatorTrainingSerializer


class MonthlyAssignmentViewSet(viewsets.ModelViewSet):
    queryset = MonthlyAssignment.objects.all()
    serializer_class = MonthlyAssignmentSerializer









from rest_framework import viewsets
from .models import HQ, Factory, Department, Line, Level
from .serializers import HQSerializer, FactorySerializer, DepartmentSerializer, LineSerializer, LevelSerializer

class HQViewSet(viewsets.ModelViewSet):
    queryset = HQ.objects.all()
    serializer_class = HQSerializer

class FactoryViewSet(viewsets.ModelViewSet):
    queryset = Factory.objects.all()
    serializer_class = FactorySerializer

class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer

class LineViewSet(viewsets.ModelViewSet):
    queryset = Line.objects.all()
    serializer_class = LineSerializer

class LevelViewSet(viewsets.ModelViewSet):
    queryset = Level.objects.all()
    serializer_class = LevelSerializer





from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import OperatorLevelTracking, EmployeeMaster
from .serializers import OperatorLevelTrackingSerializer
from datetime import date, timedelta

@api_view(['GET'])
def get_today_milestones(request):
    today = date.today()
    milestone_rules = OperatorLevelTracking.objects.all()
    milestone_data = []

    for rule in milestone_rules:
        expected_join_date = today - timedelta(days=rule.day)
        matched_employees = EmployeeMaster.objects.filter(joining_date=expected_join_date)

        for employee in matched_employees:
            milestone_data.append({
                "operator_name": employee.name,
                "level_name": rule.level.name,
                "day": rule.day,
                "milestone_date": today,
                "message": f"{employee.name} has completed milestone: {rule.level.name} on Day {rule.day}"
            })

    return Response({
        "date": str(today),
        "milestones": milestone_data
    })






from datetime import date, timedelta
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import OperatorLevelEmailTracking, TrackingEmail, EmployeeMaster
from .utils import send_milestone_email

@api_view(['GET'])
def get_today_email_milestones(request):
    today = date.today()
    rules = OperatorLevelEmailTracking.objects.all()
    email_logs = []

    for rule in rules:
        milestone_day = rule.day
        expected_date = today - timedelta(days=milestone_day)

        # Get all employees who joined exactly 'day' days ago
        matching_employees = EmployeeMaster.objects.filter(joining_date=expected_date)

        recipient_list = [e.email for e in rule.emails.all() if e.email]

        for employee in matching_employees:
            if recipient_list:
                subject = "Milestone Alert"
                message = f"{employee.name} has reached milestone: {rule.level.name} on Day {rule.day}."
                send_milestone_email(subject, message, recipient_list)
                email_logs.append({
                    "employee": employee.name,
                    "joined_on": str(employee.joining_date),
                    "level": rule.level.name,
                    "day": rule.day,
                    "recipients": recipient_list,
                    "status": "Email sent"
                })

    return Response({
        "message": "Milestone emails sent to matching employees.",
        "email_logs": email_logs
    })









from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework import status
from .models import Machine, MachineAllocation
from .serializers import MachineSerializer, MachineAllocationSerializer

class MachineViewSet(viewsets.ModelViewSet):
    queryset = Machine.objects.all()
    serializer_class = MachineSerializer


from rest_framework import viewsets, status
from rest_framework.response import Response
from django.core.mail import send_mail
from .models import MachineAllocation, MachineAllocationTrackingEmail
from .serializers import MachineAllocationSerializer

class MachineAllocationViewSet(viewsets.ModelViewSet):
    queryset = MachineAllocation.objects.all()
    serializer_class = MachineAllocationSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            self.perform_create(serializer)

            # Send email to all tracking emails
            allocation = serializer.instance
            subject = "Machine Allocation Approval Request"
            message = (
                f"Machine '{allocation.machine.name}' has been allocated to "
                f"'{allocation.employee.name}'.\n\n"
                "Please review and approve this allocation request."
            )
            from_email = None  # Will use DEFAULT_FROM_EMAIL
            recipient_list = list(MachineAllocationTrackingEmail.objects.values_list('email', flat=True))

            if recipient_list:
                send_mail(subject, message, from_email, recipient_list)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)

        # Ensure only 'approval_status' is being updated
        if set(serializer.validated_data.keys()) != {'approval_status'}:
            return Response(
                {'error': 'Only approval_status can be updated.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        self.perform_update(serializer)
        return Response(serializer.data)







from rest_framework import viewsets
from .models import SkillTraining
from .serializers import SkillTrainingSerializer

class SkillTrainingViewSet(viewsets.ModelViewSet):
    queryset = SkillTraining.objects.all()
    serializer_class = SkillTrainingSerializer






from rest_framework import viewsets
from .models import SubTopic
from .serializers import SubTopicSerializer

class SubTopicViewSet(viewsets.ModelViewSet):
    queryset = SubTopic.objects.all()
    serializer_class = SubTopicSerializer



from rest_framework import viewsets
from .models import SubTopic
from .serializers import SubTopicDaySerializer

class SubTopicDayViewSet(viewsets.ModelViewSet):
    queryset = SubTopic.objects.all()
    serializer_class = SubTopicDaySerializer




from rest_framework import viewsets
from .models import SubTopicContent
from .serializers import SubTopicContentSerializer

class SubTopicContentViewSet(viewsets.ModelViewSet):
    queryset = SubTopicContent.objects.all()
    serializer_class = SubTopicContentSerializer





from rest_framework import viewsets
from .models import Days
from .serializers import DaysSerializer

class DaysViewSet(viewsets.ModelViewSet):
    queryset = Days.objects.all()
    serializer_class = DaysSerializer







from rest_framework import viewsets
from .models import TrainingContent
from .serializers import TrainingContentSerializer

class TrainingContentViewSet(viewsets.ModelViewSet):
    queryset = TrainingContent.objects.all()
    serializer_class = TrainingContentSerializer

from rest_framework.views import APIView
class TrainingContentCreateView(APIView):
    def post(self, request):
        serializer = TrainingContentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    




from rest_framework import viewsets
from .models import LevelTwoProduction, LevelTwoLine,LevelTwoSubStation
from .serializers import LevelTwoProductionSerializer, LevelTwoLineSerializer,LevelTwoSubStationSerializer

class LevelTwoProductionViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoProduction.objects.all()
    serializer_class = LevelTwoProductionSerializer


class LevelTwoLineViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoLine.objects.all()
    serializer_class = LevelTwoLineSerializer



class LevelTwoSubStationViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoSubStation.objects.all()
    serializer_class = LevelTwoSubStationSerializer

    def get_queryset(self):
        line_id = self.request.query_params.get('line_id')
        if line_id:
            return LevelTwoSubStation.objects.filter(line_id=line_id)
        return LevelTwoSubStation.objects.all()


from rest_framework import viewsets
from .models import (
    LevelTwoTraineeInfo,
    LevelTwoTrainingTopic,
    LevelTwoOJTDay,
    LevelTwoOJTScore,
)
from .serializers import (
    LevelTwoTraineeInfoSerializer,
    LevelTwoTrainingTopicSerializer,
    LevelTwoOJTDaySerializer,
    LevelTwoOJTScoreSerializer,
)

class LevelTwoTraineeInfoViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoTraineeInfo.objects.all()
    serializer_class = LevelTwoTraineeInfoSerializer


class LevelTwoTrainingTopicViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoTrainingTopic.objects.all()
    serializer_class = LevelTwoTrainingTopicSerializer


class LevelTwoOJTDayViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoOJTDay.objects.all()
    serializer_class = LevelTwoOJTDaySerializer


from rest_framework import viewsets
from .models import LevelTwoOJTScore
from .serializers import LevelTwoOJTScoreSerializer
from .utils import check_and_update_operator_skill  # Make sure to import this

class LevelTwoOJTScoreViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoOJTScore.objects.all()
    serializer_class = LevelTwoOJTScoreSerializer

    def perform_create(self, serializer):
        # Save the new OJT score
        instance = serializer.save()

        # Update training status
        instance.trainee.calculate_and_save_training_status()

        # ✅ Call the check_and_update_operator_skill function
        if instance.trainee.traineeId:
            check_and_update_operator_skill(instance.trainee.traineeId)





from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import EmployeeMaster

class EmployeeNameByCodeAPIView(APIView):
    def get(self, request, pay_code):
        try:
            employee = EmployeeMaster.objects.get(pay_code=pay_code)
            return Response({'name': employee.name, 'pay_code':pay_code}, status=status.HTTP_200_OK)
        except EmployeeMaster.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)







from rest_framework import viewsets
from .models import EmployeeLevelAssignment
from .serializers import EmployeeLevelAssignmentSerializer

class EmployeeLevelAssignmentViewSet(viewsets.ModelViewSet):
    queryset = EmployeeLevelAssignment.objects.all()
    serializer_class = EmployeeLevelAssignmentSerializer







from rest_framework import viewsets
from .models import LevelTwoTraineeInfo
from .serializers import NestedLevelTwoTraineeInfoSerializer

class NestedLevelTwoTraineeInfoViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoTraineeInfo.objects.all()
    serializer_class = NestedLevelTwoTraineeInfoSerializer






from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import LevelTwoTraineeInfo
from .serializers import NestedLevelTwoTraineeInfoSerializer

class GetTraineeByCodeView(APIView):
    def get(self, request, trainee_id, station_id):
        try:
            trainee = LevelTwoTraineeInfo.objects.get(traineeId=trainee_id, station_id=station_id)
        except LevelTwoTraineeInfo.DoesNotExist:
            return Response({"error": "Trainee not found with this ID and station"}, status=status.HTTP_404_NOT_FOUND)

        # ✅ Call function to calculate and save status
        trainee.calculate_and_save_training_status()

        serializer = NestedLevelTwoTraineeInfoSerializer(trainee)
        return Response(serializer.data, status=status.HTTP_200_OK)






from rest_framework import viewsets
from .models import LevelTwoQuality, LevelTwoQualityLine,LevelTwoQualitySubStation
from .serializers import LevelTwoQualitySerializer, LevelTwoQualityLineSerializer,LevelTwoQualitySubStationSerializer

class LevelTwoQualityViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoQuality.objects.all()
    serializer_class = LevelTwoQualitySerializer


class LevelTwoQualityLineViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoQualityLine.objects.all()
    serializer_class = LevelTwoQualityLineSerializer

class LevelTwoQualitySubStationViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoQualitySubStation.objects.all()
    serializer_class = LevelTwoQualitySubStationSerializer




from rest_framework import viewsets
from .models import (
    LevelTwoQATraineeInfo,
    LevelTwoQATrainingTopic,
    LevelTwoQAOJTDay,
    LevelTwoQAOJTScore,
)
from .serializers import (
    LevelTwoQATraineeInfoSerializer,
    LevelTwoQATrainingTopicSerializer,
    LevelTwoQAOJTDaySerializer,
    LevelTwoQAOJTScoreSerializer,
)


class LevelTwoQATraineeInfoViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoQATraineeInfo.objects.all()
    serializer_class = LevelTwoQATraineeInfoSerializer


class LevelTwoQATrainingTopicViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoQATrainingTopic.objects.all()
    serializer_class = LevelTwoQATrainingTopicSerializer


class LevelTwoQAOJTDayViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoQAOJTDay.objects.all()
    serializer_class = LevelTwoQAOJTDaySerializer


class LevelTwoQAOJTScoreViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoQAOJTScore.objects.all()
    serializer_class = LevelTwoQAOJTScoreSerializer






from rest_framework import viewsets
from .models import LevelTwoQATraineeInfo
from .serializers import NestedLevelTwoQATraineeInfoSerializer

class NestedLevelTwoQATraineeInfoViewSet(viewsets.ModelViewSet):
    queryset = LevelTwoQATraineeInfo.objects.all()
    serializer_class = NestedLevelTwoQATraineeInfoSerializer





from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import LevelTwoQATraineeInfo
from .serializers import NestedLevelTwoQATraineeInfoSerializer

# URL: /api/qa-trainee/<trainee_id>/<line_id>/

class GetQATraineeByCodeView(APIView):
    def get(self, request, trainee_id, station_id):
        try:
            trainee = LevelTwoQATraineeInfo.objects.get(traineeId=trainee_id, station_id=station_id)
        except LevelTwoQATraineeInfo.DoesNotExist:
            return Response({"error": "Trainee not found with this ID and station"}, status=status.HTTP_404_NOT_FOUND)

        # ✅ Call function to calculate and save status
        trainee.calculate_and_save_training_status()

        serializer = NestedLevelTwoQATraineeInfoSerializer(trainee)
        return Response(serializer.data, status=status.HTTP_200_OK)




from rest_framework import viewsets
from .models import LevelThreeProduction, LevelThreeLine, LevelThreeSubStation
from .serializers import (
    LevelThreeProductionSerializer,
    LevelThreeLineSerializer,
    LevelThreeSubStationSerializer
)


class LevelThreeProductionViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeProduction.objects.all()
    serializer_class = LevelThreeProductionSerializer


class LevelThreeLineViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeLine.objects.all()
    serializer_class = LevelThreeLineSerializer


class LevelThreeSubStationViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeSubStation.objects.all()
    serializer_class = LevelThreeSubStationSerializer













from django.shortcuts import render

# Create your views here.

from rest_framework import viewsets
from .models import LevelThreeTraineeInfo, LevelThreeTrainingTopic, LevelThreeOJTDay, LevelThreeOJTScore
from .serializers import (
    LevelThreeTraineeInfoSerializer,
    LevelThreeTrainingTopicSerializer,
    LevelThreeOJTDaySerializer,
    LevelThreeOJTScoreSerializer,
)

class LevelThreeTraineeInfoViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeTraineeInfo.objects.all()
    serializer_class = LevelThreeTraineeInfoSerializer

class LevelThreeTrainingTopicViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeTrainingTopic.objects.all()
    serializer_class = LevelThreeTrainingTopicSerializer

class LevelThreeOJTDayViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeOJTDay.objects.all()
    serializer_class = LevelThreeOJTDaySerializer




from .utils import check_and_update_operator_skill_level_three

class LevelThreeOJTScoreViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeOJTScore.objects.all()
    serializer_class = LevelThreeOJTScoreSerializer

    def perform_create(self, serializer):
    # Save the new OJT score
     instance = serializer.save()

    # Update training status
     instance.trainee.calculate_and_save_training_status()

    # ✅ Call the Level 3 operator skill update
     if instance.trainee.trainee_Id:
         check_and_update_operator_skill_level_three(instance.trainee.trainee_Id)

    def get_queryset(self):
        trainee_name = self.request.query_params.get('trainee')
        if trainee_name:
            return LevelThreeOJTScore.objects.filter(trainee__trainee_name=trainee_name)
        return super().get_queryset()
    




from rest_framework import viewsets
from .models import LevelThreeTraineeInfo
from .serializers import NestedLevelThreeTraineeInfoSerializer

class NestedLevelThreeTraineeInfoViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeTraineeInfo.objects.all()
    serializer_class = NestedLevelThreeTraineeInfoSerializer





from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import LevelThreeTraineeInfo
from .serializers import NestedLevelThreeTraineeInfoSerializer

# URL: /api/levelthree-trainee/<trainee_id>/<station_id>/

class GetLevelThreeTraineeByCodeView(APIView):
    def get(self, request, trainee_id, station_id):
        try:
            trainee = LevelThreeTraineeInfo.objects.get(trainee_id=trainee_id, station_id=station_id)

        except LevelThreeTraineeInfo.DoesNotExist:
            return Response({"error": "Trainee not found with this ID and station"}, status=status.HTTP_404_NOT_FOUND)

        # ✅ Call function to calculate and save status
        trainee.calculate_and_save_training_status()

        serializer = NestedLevelThreeTraineeInfoSerializer(trainee)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

    # def get(self, request, trainee_id, station_id):
    #     try:
    #         trainee = LevelThreeTraineeInfo.objects.get(trainee_Id=trainee_id, station_id=station_id)
    #     except LevelThreeTraineeInfo.DoesNotExist:
    #         return Response(
    #             {"error": "Trainee not found with this ID and station"},
    #             status=status.HTTP_404_NOT_FOUND
    #         )

    #     serializer = NestedLevelThreeTraineeInfoSerializer(trainee)
    #     return Response(serializer.data, status=status.HTTP_200_OK)






from rest_framework import viewsets
from .models import LevelThreeQuality, LevelThreeQualityLine,LevelThreeQualitySubStation
from .serializers import LevelThreeQualitySerializer, LevelThreeQualityLineSerializer,LevelThreeQualitySubStationSerializer

class LevelThreeQualityViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeQuality.objects.all()
    serializer_class = LevelThreeQualitySerializer


class LevelThreeQualityLineViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeQualityLine.objects.all()
    serializer_class = LevelThreeQualityLineSerializer




class LevelThreeQualitySubStationViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeQualitySubStation.objects.all()
    serializer_class = LevelThreeQualitySubStationSerializer







from rest_framework import viewsets
from .models import (
    LevelThreeQATraineeInfo,
    LevelThreeQATrainingTopic,
    LevelThreeQAOJTDay,
    LevelThreeQAOJTScore,
)
from .serializers import (
    LevelThreeQATraineeInfoSerializer,
    LevelThreeQATrainingTopicSerializer,
    LevelThreeQAOJTDaySerializer,
    LevelThreeQAOJTScoreSerializer,
)

class LevelThreeQATraineeInfoViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeQATraineeInfo.objects.all()
    serializer_class = LevelThreeQATraineeInfoSerializer


class LevelThreeQATrainingTopicViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeQATrainingTopic.objects.all()
    serializer_class = LevelThreeQATrainingTopicSerializer


class LevelThreeQAOJTDayViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeQAOJTDay.objects.all()
    serializer_class = LevelThreeQAOJTDaySerializer


class LevelThreeQAOJTScoreViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeQAOJTScore.objects.all()
    serializer_class = LevelThreeQAOJTScoreSerializer








from rest_framework import viewsets
from .models import LevelThreeQATraineeInfo
from .serializers import NestedLevelThreeQATraineeInfoSerializer

class NestedLevelThreeQATraineeInfoViewSet(viewsets.ModelViewSet):
    queryset = LevelThreeQATraineeInfo.objects.all()
    serializer_class = NestedLevelThreeQATraineeInfoSerializer






from rest_framework import status
from .models import LevelTwoQATraineeInfo
from .serializers import NestedLevelThreeQATraineeInfoSerializer

# URL: /api/qa-trainee/<trainee_id>/<line_id>/

class GetThreeQATraineeByCodeView(APIView):
    def get(self, request, trainee_id, station_id):
        try:
            trainee = LevelThreeQATraineeInfo.objects.get(traineeId=trainee_id, station_id=station_id)
        except LevelThreeQATraineeInfo.DoesNotExist:
            return Response({"error": "Trainee not found with this ID and station"}, status=status.HTTP_404_NOT_FOUND)

        # ✅ Call function to calculate and save status
        trainee.calculate_and_save_training_status()

        serializer = NestedLevelThreeQATraineeInfoSerializer(trainee)
        return Response(serializer.data, status=status.HTTP_200_OK)






from rest_framework import viewsets
from .models import ARVRTrainingContent
from .serializers import ARVRTrainingContentSerializer

class ARVRTrainingContentViewSet(viewsets.ModelViewSet):
    queryset = ARVRTrainingContent.objects.all()
    serializer_class = ARVRTrainingContentSerializer






from rest_framework import viewsets
from .models import MCQQuestion
from .serializers import MCQQuestionSerializer

class MCQQuestionViewSet(viewsets.ModelViewSet):
    queryset = MCQQuestion.objects.all()
    serializer_class = MCQQuestionSerializer







from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import EmployeeMaster

class EmployeeNameByCodeAPIView(APIView):
    def get(self, request, pay_code):
        try:
            employee = EmployeeMaster.objects.get(pay_code=pay_code)
            return Response({'name': employee.name}, status=status.HTTP_200_OK)
        except EmployeeMaster.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)







from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status
from tablib import Dataset
from .resources import BiometricAttendanceResource

class ExcelUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, format=None):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        file_format = 'xls' if file_obj.name.endswith('.xls') else 'xlsx'
        dataset = Dataset()
        try:
            imported_data = dataset.load(file_obj.read(), format=file_format)
        except Exception as e:
            return Response({'error': f'Failed to read Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        resource = BiometricAttendanceResource()
        result = resource.import_data(dataset, dry_run=True)

        if result.has_errors():
            errors = []
            for row_number, row_errors in result.row_errors():
                for error in row_errors:
                    errors.append(f"Row {row_number}: {str(error.error)}")
            return Response({'error': 'Import failed', 'details': errors}, status=status.HTTP_400_BAD_REQUEST)

        # Perform actual import
        resource.import_data(dataset, dry_run=False)
        return Response({'success': 'Data imported successfully'}, status=status.HTTP_201_CREATED)






import os
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from tablib import Dataset
from .resources import BiometricAttendanceResource

class ExcelUploadFromPathView(APIView):
    # Set the path to your Excel file here
    EXCEL_FILE_PATH = r"E:\attendance.xlsx"  # Change this to your actual path

    def post(self, request, format=None):
        if not os.path.exists(self.EXCEL_FILE_PATH):
            return Response({'error': 'File not found on server path'}, status=status.HTTP_400_BAD_REQUEST)

        file_format = 'xls' if self.EXCEL_FILE_PATH.endswith('.xls') else 'xlsx'

        dataset = Dataset()
        try:
            with open(self.EXCEL_FILE_PATH, 'rb') as file_obj:
                imported_data = dataset.load(file_obj.read(), format=file_format)
        except Exception as e:
            return Response({'error': f'Failed to read Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        resource = BiometricAttendanceResource()
        result = resource.import_data(imported_data, dry_run=True)

        if result.has_errors():
            errors = []
            for row_number, row_errors in result.row_errors():
                for error in row_errors:
                    errors.append(f"Row {row_number}: {str(error.error)}")
            return Response({'error': 'Import failed', 'details': errors}, status=status.HTTP_400_BAD_REQUEST)

        # Perform actual import
        resource.import_data(imported_data, dry_run=False)
        return Response({'success': 'Data imported successfully'}, status=status.HTTP_201_CREATED)




from rest_framework import viewsets
from .models import BiometricAttendance
from .serializers import BiometricAttendanceSerializer

class BiometricAttendanceViewSet(viewsets.ModelViewSet):
    queryset = BiometricAttendance.objects.all()
    serializer_class = BiometricAttendanceSerializer







from rest_framework import viewsets
from .models import MultiSkilling
from .serializers import NewMultiSkillingSerializer

class NewMultiSkillingViewSet(viewsets.ModelViewSet):
    queryset = MultiSkilling.objects.all()
    serializer_class = NewMultiSkillingSerializer



from rest_framework.views import APIView
from rest_framework.response import Response
from .models import EmployeeMaster, MultiSkilling
from rest_framework import status

class AllEmployeesWithActiveSkillsView(APIView):
    def get(self, request):
        name_query = request.GET.get('name', '')

        # Filter employees by name (case-insensitive)
        employees = EmployeeMaster.objects.filter(name__icontains=name_query)

        result = []
        for emp in employees:
            # Fetch only active skills for this employee
            active_skills = MultiSkilling.objects.filter(
                employee=emp, status='active'
            ).select_related('skill_level', 'station')

            skills = [
                {
                    "skill": skill.skill,
                    "skill_level": skill.skill_level.skill_level,
                    "start_date": skill.start_date,
                    "end_date": skill.end_date,
                    "notes": skill.notes,
                    "status": skill.status,
                }
                for skill in active_skills
            ]

            result.append({
                "employee_id": emp.id,
                "pay_code": emp.pay_code,
                "card_no": emp.card_no,
                "name": emp.name,
                "department": emp.department,
                "section": emp.section,
                "designation_category": emp.desig_category,
                "joining_date": emp.joining_date,
                "skills": skills
            })

        return Response(result, status=status.HTTP_200_OK)












from rest_framework.views import APIView
from rest_framework.response import Response
from .models import MultiSkilling, EmployeeMaster
from django.db.models import Prefetch, Q

class GroupedEmployeeSkillsView(APIView):
    def get(self, request):
        valid_statuses = ['scheduled', 'inprogress', 'completed']
        skills = MultiSkilling.objects.filter(status__in=valid_statuses) \
            .select_related('employee', 'skill_level', 'station') \
            .order_by('employee_id', 'start_date')

        grouped_data = {}

        for skill in skills:
            emp = skill.employee
            emp_id = emp.id

            if emp_id not in grouped_data:
                grouped_data[emp_id] = {
                    "employee_id": emp.id,
                    "pay_code": emp.pay_code,
                    "card_no": emp.card_no,
                    "name": emp.name,
                    "department": emp.department,
                    "section": emp.section,
                    "joining_date": emp.joining_date,
                    "skills": []
                }

            grouped_data[emp_id]["skills"].append({
                "station": skill.station.skill if skill.station else None,
                "station_number": skill.station.station_number if skill.station else None,
                "start_date": skill.start_date,
                "end_date": skill.end_date,
                "notes": skill.notes,
                "status": skill.status,
                "skill_level": skill.skill_level.name if skill.skill_level else None
            })

        return Response(list(grouped_data.values()))

from rest_framework.views import APIView
from rest_framework.response import Response
from .models import EmployeeMaster, MultiSkilling
from rest_framework import status

class AllEmployeesWithCompletedSkillsView(APIView):
    def get(self, request):
        name_query = request.GET.get('name', '')

        # Filter employees by name (case-insensitive)
        employees = EmployeeMaster.objects.filter(name__icontains=name_query)

        result = []
        for emp in employees:
            # Fetch only active skills for this employee
            active_skills = MultiSkilling.objects.filter(
                employee=emp, status='completed'
            ).select_related('skill_level', 'station')

            skills = [
                {
                    "skill": skill.skill,
                    "skill_level": skill.skill_level.skill_level,
                    "start_date": skill.start_date,
                    "end_date": skill.end_date,
                    "notes": skill.notes,
                    "status": skill.status,
                }
                for skill in active_skills
            ]

            result.append({
                "employee_id": emp.id,
                "pay_code": emp.pay_code,
                "card_no": emp.card_no,
                "name": emp.name,
                "department": emp.department,
                "section": emp.section,
                "designation_category": emp.desig_category,
                "joining_date": emp.joining_date,
                "skills": skills
            })

        return Response(result, status=status.HTTP_200_OK)














from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status as http_status
from .serializers import RefreshMultiSkillingSerializer
from .models import MultiSkilling

@api_view(['POST'])
def create_rescheduled_multiskilling(request):
    data = request.data.copy()  # Copy the request data
    data['status'] = 'rescheduled'  # Force status to 'rescheduled'

    serializer = RefreshMultiSkillingSerializer(data=data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=http_status.HTTP_201_CREATED)
    return Response(serializer.errors, status=http_status.HTTP_400_BAD_REQUEST)






from rest_framework.views import APIView
from rest_framework.response import Response
from .models import EmployeeMaster, MultiSkilling
from rest_framework import status

class AllEmployeesWithRescheduledSkillsView(APIView):
    def get(self, request):
        name_query = request.GET.get('name', '')

        # Filter employees by name (case-insensitive)
        employees = EmployeeMaster.objects.filter(name__icontains=name_query)

        result = []
        for emp in employees:
            # Fetch only 'rescheduled' skills for this employee
            rescheduled_skills = MultiSkilling.objects.filter(
                employee=emp, status='rescheduled'
            ).select_related('skill_level', 'station')

            # If the employee has no rescheduled skills, skip
            if not rescheduled_skills.exists():
                continue

            skills = [
                {
                    "id": skill.id,
                    "skill": skill.skill,
                    "notes": skill.notes,
                    "status": skill.status,
                    "reason": skill.reason,
                    "refreshment_date": skill.refreshment_date,
                }
                for skill in rescheduled_skills
            ]

            result.append({
                "employee_id": emp.id,
                "pay_code": emp.pay_code,
                "card_no": emp.card_no,
                "name": emp.name,
                "department": emp.department,
                "section": emp.section,
                "designation_category": emp.desig_category,
                "joining_date": emp.joining_date,
                "skills": skills
            })

        return Response(result, status=status.HTTP_200_OK)









from rest_framework import viewsets
from .models import TrainingReport
from .serializers import TrainingReportSerializer

class TrainingReportViewSet(viewsets.ModelViewSet):
    queryset = TrainingReport.objects.all().order_by('-month')
    serializer_class = TrainingReportSerializer



from rest_framework import viewsets
from .models import UnifiedDefectReport
from .serializers import UnifiedDefectReportSerializer

class UnifiedDefectReportViewSet(viewsets.ModelViewSet):
    queryset = UnifiedDefectReport.objects.all().order_by('-month')
    serializer_class = UnifiedDefectReportSerializer





from rest_framework.views import APIView
from rest_framework.response import Response
from .models import TrainingReport
from .serializers import TrainingReportSerializer
from django.db.models import Sum

class TrainingSummaryView(APIView):
    def get(self, request):
        summary = TrainingReport.objects.aggregate(
            new_operators_joined=Sum("new_operators_joined"),
            new_operators_trained=Sum("new_operators_trained"),
            total_trainings_planned=Sum("total_trainings_planned"),
            total_trainings_actual=Sum("total_trainings_actual")
        )
        return Response(summary)





from rest_framework.generics import ListAPIView
from .models import TrainingReport
from .serializers import TrainingReportSerializer

class OperatorsJoinedVsTrainedView(ListAPIView):
    queryset = TrainingReport.objects.all().order_by("month")
    serializer_class = TrainingReportSerializer




from rest_framework.views import APIView
from .models import UnifiedDefectReport
from rest_framework.response import Response

class MSILDefectsView(APIView):
    def get(self, request):
        data = UnifiedDefectReport.objects.filter(category='MSIL').order_by('month')
        return Response([
            {
                'month': d.month,
                'total_defects': d.total_defects,
                'ctq_defects': d.ctq_defects,
            }
            for d in data
        ])





class CTQDefectsAllPlantsView(APIView):
    def get(self, request):
        data = UnifiedDefectReport.objects.filter(category='All Plants').order_by('month')
        return Response([
            {
                'month': d.month,
                'total_defects': d.total_defects,
                'ctq_defects': d.ctq_defects,
            }
            for d in data
        ])




class InternalRejectionView(APIView):
    def get(self, request):
        internal = UnifiedDefectReport.objects.all()
        rejection = internal.aggregate(
            total_internal_rejection=Sum('total_internal_rejection'),
            ctq_internal_rejection=Sum('ctq_internal_rejection')
        )
        return Response(rejection)



from rest_framework import viewsets
from .models import TrainingReport
from .serializers import TrainingReportSerializer

class TrainingReportViewSet(viewsets.ModelViewSet):
    queryset = TrainingReport.objects.all().order_by('-month')
    serializer_class = TrainingReportSerializer



from rest_framework import viewsets
from .models import UnifiedDefectReport
from .serializers import UnifiedDefectReportSerializer

class UnifiedDefectReportViewSet(viewsets.ModelViewSet):
    queryset = UnifiedDefectReport.objects.all().order_by('-month')
    serializer_class = UnifiedDefectReportSerializer

from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import viewsets
from .models import TrainingContent
from .serializers import TrainingContentSerializer


class SubtopicWiseTrainingContentViewSet(viewsets.ModelViewSet):
    queryset = TrainingContent.objects.all()
    serializer_class = TrainingContentSerializer

    @action(detail=False, methods=['get'], url_path=r'(?P<id>\d+)')
    def subtopicwise(self, request, id=None):
        contents = TrainingContent.objects.filter(subtopic_content_id=id)
        serializer = self.get_serializer(contents, many=True)
        return Response(serializer.data)
    


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import MCQQuestion
from .serializers import MCQQuestionSerializer

class MCQBySubtopicView(APIView):
    def get(self, request, subtopic_id):
        mcqs = MCQQuestion.objects.filter(subtopic_content_id=subtopic_id)
        serializer = MCQQuestionSerializer(mcqs, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

# import xlrd
from datetime import datetime
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import EmployeeMaster
from .serializers import ExcelUploadSerializer
import xlrd


def parse_excel_date(value, datemode):
    if isinstance(value, float):  # Excel serial date
        return xlrd.xldate.xldate_as_datetime(value, datemode).date()
    elif isinstance(value, str):
        try:
            return datetime.strptime(value.strip(), "%d/%m/%Y").date()
        except:
            return None
    return None

class EmployeeExcelUploadView(APIView):
    def post(self, request, *args, **kwargs):
        serializer = ExcelUploadSerializer(data=request.data)
        if serializer.is_valid():
            excel_file = serializer.validated_data['file']
            try:
                workbook = xlrd.open_workbook(file_contents=excel_file.read())
                sheet = workbook.sheet_by_index(0)

                # Normalize headers (lowercase, no leading/trailing whitespace), skip first column
                raw_headers = [str(cell.value).strip().lower() for cell in sheet.row(5)][1:]

                for row_idx in range(6, sheet.nrows):  # Start after header row
                    row_cells = sheet.row(row_idx)[1:]  # Skip "Sr. No." column
                    row_data = {raw_headers[i]: row_cells[i].value for i in range(len(raw_headers))}

                    # Safely fetch and convert dates
                    birth_date = parse_excel_date(row_data.get('birth date'), workbook.datemode)
                    joining_date = parse_excel_date(row_data.get('joining date'), workbook.datemode)

                    if not birth_date or not joining_date:
                        continue  # Skip rows with invalid dates

                    # Create or update the employee record
                    EmployeeMaster.objects.update_or_create(
                        pay_code=row_data.get('pay code'),
                        defaults={
                            'card_no': row_data.get('card no.'),
                            'sex': row_data.get('sex'),
                            'birth_date': birth_date,
                            'name': row_data.get('name'),
                            'guardian_name': row_data.get("guardian's name", ''),
                            'department': row_data.get('department'),
                            'section': row_data.get('section'),
                            'desig_category': row_data.get('desig/categor') or None,
                            'joining_date': joining_date,
                            'auth_shift': row_data.get('auth shift'),
                            'shift_type': row_data.get('shift type'),
                            'shift_pattern': row_data.get('shift pattern'),
                            'first_weekly_off': row_data.get('1st weekly') or '',  # corrected key
                            'second_weekly_off': None,
                            'second_weekly_off_fh': None,
                            'ot_allowed_rate': False,
                            'round_the_clock': False,
                        }
                    )

                return Response({"message": "Employees uploaded successfully"}, status=status.HTTP_201_CREATED)

            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)






# from rest_framework import viewsets
# from rest_framework.decorators import action
# from rest_framework.response import Response
# from .models import MachineAllocation, OperatorSkill, Machine
# from .serializers import EmployeeNameSerializer

# class EmployeeMachineAllocationViewSet(viewsets.ModelViewSet):
#     queryset = MachineAllocation.objects.all()
#     serializer_class = ...  # Your main serializer

#     @action(detail=False, methods=['get'], url_path='eligible-employees')
#     def eligible_employees(self, request):
#         machine_id = request.query_params.get('machine_id')
#         if not machine_id:
#             return Response({'error': 'machine_id is required'}, status=400)

#         try:
#             machine = Machine.objects.get(id=machine_id)
#         except Machine.DoesNotExist:
#             return Response({'error': 'Machine not found'}, status=404)

#         matching_skills = OperatorSkill.objects.filter(station__skill=machine.process)
#         employee_ids = matching_skills.values_list('operator_id', flat=True).distinct()
#         employees = EmployeeMaster.objects.filter(id__in=employee_ids)

#         serializer = EmployeeNameSerializer(employees, many=True)
#         return Response(serializer.data)


# easytest


from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import *
from .serializers import *
from rest_framework import viewsets


class KeyEventCreateView(APIView):
    def post(self, request):
        serializer = KeyEventSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'Key event saved'}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class LatestKeyEventView(APIView):
    def get(self, request):
        try:
            latest_event = KeyEvent.objects.latest('timestamp')
            serializer = KeyEventSerializer(latest_event)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except KeyEvent.DoesNotExist:
            return Response({"message": "No key events yet."}, status=status.HTTP_404_NOT_FOUND)

        
# api/views.py
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .serializers import ConnectEventSerializer

@api_view(['POST'])
def connect_event_create(request):
    serializer = ConnectEventSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)




@api_view(['POST'])
def vote_event_create(request):
    serializer = VoteEventSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)




from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from .models import Question, EmployeeMaster, Level, Station, TestSession, Score
from .serializers import QuestionSerializer, EmployeeSerializer, ScoreSerializer, SimpleScoreSerializer

from rest_framework.views import APIView
from rest_framework.response import Response
from .models import Question
from .serializers import QuestionSerializer


from rest_framework import generics

class QuestionListCreateView(generics.ListCreateAPIView):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer

    def get_queryset(self):
        paper_id = self.request.query_params.get('paper_id')
        if paper_id:
            return self.queryset.filter(question_paper__id=paper_id)
        return self.queryset
    
from rest_framework import generics

class QuestionDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer




class QuestionPaperListCreateView(generics.ListCreateAPIView):
    queryset = QuestionPaper.objects.all()
    serializer_class = QuestionPaperSerializer

class QuestionsByPaperView(generics.ListAPIView):
    serializer_class = QuestionSerializer

    def get_queryset(self):
        paper_id = self.kwargs.get('paper_id')
        return Question.objects.filter(question_paper_id=paper_id)




class EmployeeListCreateView(APIView):
    def get(self, request):
        employees = EmployeeMaster.objects.all()
        serializer = EmployeeSerializer(employees, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = EmployeeSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ScoreListView(APIView):
    def get(self, request):
        # Assuming you use caching for latest test session
        session_key = cache.get("latest_test_session")
        if not session_key:
            return Response([])

        scores = Score.objects.filter(session_key=session_key).select_related('employee', 'level', 'skill')
        serializer = ScoreSerializer(scores, many=True)
        return Response(serializer.data)


class KeyIdToEmployeeNameMap(APIView):
    def get(self, request):
        mapping = TestSession.objects.select_related('employee').all()
        return Response({s.key_id: s.employee.name for s in mapping})

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from .models import EmployeeMaster, Station, TestSession, QuestionPaper

# class StartTestSessionView(APIView):
#     def post(self, request):
#         try:
#             test_name = request.data.get("test_name")
#             assignments = request.data.get("assignments", [])
#             question_paper_id = request.data.get("question_paper_id")

#             if not test_name or not assignments:
#                 return Response(
#                     {"error": "Test name and assignments are required."},
#                     status=status.HTTP_400_BAD_REQUEST,
#                 )

#             question_paper = None
#             if question_paper_id:
#                 question_paper = get_object_or_404(QuestionPaper, id=question_paper_id)

#             for item in assignments:
#                 key_id = item.get("key_id")
#                 employee_id = item.get("employee_id")

#                 if not key_id or not employee_id:
#                     return Response(
#                         {"error": "key_id and employee_id are required in each assignment."},
#                         status=status.HTTP_400_BAD_REQUEST,
#                     )

#                 employee = get_object_or_404(EmployeeMaster, id=employee_id)

#                 # You can skip station logic if not using skill anymore
#                 TestSession.objects.create(
#                     test_name=test_name,
#                     key_id=key_id,
#                     employee=employee,
#                     level=None,
#                     skill=None,  # Assuming nullable=True on model
#                     question_paper=question_paper,
#                 )

#             return Response({"status": "ok"}, status=status.HTTP_200_OK)

#         except Exception as e:
#             return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class StartTestSessionView(APIView):
    def post(self, request):
        try:
            print("Incoming Request Data:", request.data)

            test_name = request.data.get("test_name")
            assignments = request.data.get("assignments", [])
            question_paper_id = request.data.get("question_paper_id")
            level = request.data.get("level")  # string
            skill_id = request.data.get("skill")  # foreign key

            if not test_name or not assignments:
                response_data = {"error": "Test name and assignments are required."}
                print("Response:", response_data)
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

            question_paper = None
            if question_paper_id:
                question_paper = get_object_or_404(QuestionPaper, id=question_paper_id)

            skill = None
            if skill_id:
                skill = get_object_or_404(Station, id=skill_id)

            for item in assignments:
                key_id = item.get("key_id")
                employee_id = item.get("employee_id")

                if not key_id or not employee_id:
                    response_data = {"error": "key_id and employee_id are required in each assignment."}
                    print("Response:", response_data)
                    return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

                employee = get_object_or_404(EmployeeMaster, id=employee_id)

                TestSession.objects.create(
                    test_name=test_name,
                    key_id=key_id,
                    employee=employee,
                    level=level,  # now just saving the string directly
                    skill=skill,
                    question_paper=question_paper,
                )

            response_data = {"status": "ok"}
            print("Response:", response_data)
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as e:
            response_data = {"error": str(e)}
            print("Response:", response_data)
            return Response(response_data, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



##################################################################################################################################################


from rest_framework.views import APIView
from rest_framework.response import Response
from .models import TestSession, Question, Score, LevelTwoTraineeInfo
import traceback
from .utils import check_and_update_operator_skill, check_and_update_operator_skill_level_three


from .models import OperatorSkill, Station


from rest_framework.views import APIView
from rest_framework.response import Response
from .models import TestSession, Question, Score, LevelTwoTraineeInfo
import traceback
from .utils import check_and_update_operator_skill, check_and_update_operator_skill_level_three


from .models import OperatorSkill, Station


from rest_framework.views import APIView
from rest_framework.response import Response
from .models import TestSession, Question, Score, LevelTwoTraineeInfo, OperatorSkill, Station
import traceback
from .utils import check_and_update_operator_skill, check_and_update_operator_skill_level_three


from rest_framework.views import APIView
from rest_framework.response import Response
from .models import TestSession, Question, Score, EmployeeMaster, OperatorSkill, Station
import traceback
from .utils import check_and_update_operator_skill, check_and_update_operator_skill_level_three


from rest_framework.views import APIView
from rest_framework.response import Response
from .models import TestSession, Question, Score, EmployeeMaster, OperatorSkill, Station
import traceback
from .utils import check_and_update_operator_skill, check_and_update_operator_skill_level_three


class EndTestSessionView(APIView):
    def post(self, request):
        try:
            key_id_to_answers = request.data  # { key_id: [answers] }
            results = []

            sessions = TestSession.objects.select_related(
                'employee', 'skill', 'question_paper'
            ).all()

            for session in sessions:
                key_id = str(session.key_id)
                if key_id not in key_id_to_answers:
                    continue

                employee = session.employee  # should be EmployeeMaster instance
                test_name = session.test_name
                question_paper = session.question_paper
                level = session.level   # FK or string
                skill = session.skill

                questions = list(
                    Question.objects.filter(question_paper=question_paper).order_by('id')
                )
                answers = key_id_to_answers.get(key_id, [])

                correct_count = 0
                for i, ans in enumerate(answers):
                    if i < len(questions) and ans == questions[i].correct_index:
                        correct_count += 1

                percentage = round((correct_count / len(questions)) * 100) if questions else 0
                passed = percentage >= 80

                # ✅ Create the Score
                Score.objects.create(
                    employee=employee,
                    marks=correct_count,
                    percentage=percentage,
                    passed=passed,
                    test_name=test_name,
                    test=session,
                    level=level,
                    skill=skill
                )

                # ✅ If passed, update skills
                if passed:
                    try:
                        # Check employee directly in EmployeeMaster
                        emp = EmployeeMaster.objects.get(pay_code=employee.pay_code)

                        # Custom skill updates
                        check_and_update_operator_skill(emp.pay_code)
                        check_and_update_operator_skill_level_three(emp.pay_code)

                        # 🚀 Save employee to "General" station if Level 1
                        level_name = str(level) if not hasattr(level, "name") else level.name
                        print("DEBUG: Level name ->", level_name)

                        if level_name == "Level 1":
                            general_station, _ = Station.objects.get_or_create(skill="General")
                            operator_skill, created = OperatorSkill.objects.get_or_create(
                                operator=emp,
                                station=general_station,
                                defaults={"skill_level": "Level 1"}
                            )
                            if not created:
                                operator_skill.skill_level = "Level 1"
                                operator_skill.save()
                            print(f"DEBUG: Employee {emp.name} assigned to General station.")

                    except EmployeeMaster.DoesNotExist:
                        print(f"DEBUG: Employee with pay_code {employee.pay_code} not found")

                results.append({
                    'name': employee.name,
                    'marks': correct_count,
                    'percentage': percentage,
                    'passed': passed
                })

            # ✅ Delete sessions
            TestSession.objects.all().delete()

            return Response({
                'test_name': test_name,
                'results': results
            }, status=200)

        except Exception as e:
            traceback.print_exc()
            return Response({'error': str(e)}, status=500)
        



class PastTestSessionsView(APIView):
    def get(self, request):
        qs = Score.objects.values('test_name').distinct()
        return Response([s['test_name'] for s in qs])


class ScoresByTestView(APIView):
    def get(self, request, name):
        scores = (
            Score.objects
            .filter(test_name=name)
            .select_related('employee', 'skill')  # ✅ FIX: Removed 'level'
        )

        questions_count = Question.objects.count() or 1

        data = []
        for s in scores:
            data.append({
                'employee_id': s.employee.id,
                'name': s.employee.name,
                'marks': s.marks,
                'percentage': s.percentage,
                'level_name': s.level if s.level else '',  # ✅ FIX: removed .name
                'skill': s.skill.skill if s.skill else '',  # assuming Station.skill is a string
                'section': s.employee.section if s.employee.section else '',
            })
        print("GET /api/scores-by-session/ response:", data) 
        return Response(data)

    
class ResultSummaryAPIView(APIView):
    def get(self, request):
        scores = Score.objects.select_related('employee', 'level', 'skill')
        data = []
        for score in scores:
            percentage = round((score.marks / 10) * 100, 2)  # Adjust total marks accordingly
            result = 'Pass' if score.marks >= 8 else 'Retraining' if score.marks >= 5 else 'Fail'

            data.append({
                "employee_id": score.employee.id,
                "name": score.employee.name,
                "marks": score.marks,
                "percentage": percentage,
                "section": score.employee.section,  # assuming CharField
                "level_name": score.level.name if score.level else '',
                "skill": score.skill.skill if score.skill else '',  # Station.skill string
                "result": result,
            })

        serializer = SimpleScoreSerializer(data, many=True)
        return Response(serializer.data)


class SkillListView(APIView):
    def get(self, request):
        skills = Station.objects.values_list('skill', flat=True).distinct()
        return Response(skills)


class ScoresBySessionView(APIView):
    def get(self, request, session_key):
        scores = Score.objects.filter(session_key=session_key).select_related('employee', 'level', 'skill')
        serializer = ScoreSerializer(scores, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)





1

#Employee Card 



from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import OperatorSkill, EmployeeMaster
from .serializers import OperatorCardSkillSerializer

class OperatorSkillByNameView(APIView):
    def get(self, request):
        name = request.query_params.get('name')
        if name:
            try:
                employee = EmployeeMaster.objects.get(name=name)
                operator_skills = OperatorSkill.objects.filter(operator=employee)
                serializer = OperatorCardSkillSerializer(operator_skills, many=True)
                return Response(serializer.data)
            except EmployeeMaster.DoesNotExist:
                return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)
        return Response({'error': 'Name parameter is required'}, status=status.HTTP_400_BAD_REQUEST)







from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import EmployeeMaster, Score
from .serializers import CardScoreSerializer

class ScoreByEmployeeNameView(APIView):
    def get(self, request):
        name = request.query_params.get('name')
        if name:
            try:
                employee = EmployeeMaster.objects.get(name=name)
                scores = Score.objects.filter(test__employee=employee)
                serializer = CardScoreSerializer(scores, many=True)
                return Response(serializer.data)
            except EmployeeMaster.DoesNotExist:
                return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)
        return Response({'error': 'Name parameter is required'}, status=status.HTTP_400_BAD_REQUEST)







from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import EmployeeMaster, MultiSkilling
from .serializers import CardMultiSkillingSerializer

class MultiSkillingByEmployeeView(APIView):
    def get(self, request):
        name = request.query_params.get('name')
        if not name:
            return Response({'error': 'Name parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            employee = EmployeeMaster.objects.get(name=name)
        except EmployeeMaster.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)

        records = MultiSkilling.objects.filter(employee=employee)
        serializer = CardMultiSkillingSerializer(records, many=True)
        return Response(serializer.data)






from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import EmployeeMaster, RefreshmentTraining
from .serializers import CardRefreshmentTrainingSerializer

class RefreshmentTrainingByNameView(APIView):
    def get(self, request):
        name = request.query_params.get('name')
        if not name:
            return Response({'error': 'Name parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            employee = EmployeeMaster.objects.get(name=name)
        except EmployeeMaster.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)

        records = RefreshmentTraining.objects.filter(employee=employee)
        serializer = CardRefreshmentTrainingSerializer(records, many=True)
        return Response(serializer.data)






from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import EmployeeMaster
from .serializers import CardEmployeeMasterSerializer

class CardEmployeeDetailByNameView(APIView):
    def get(self, request):
        name = request.query_params.get('name')
        if not name:
            return Response({'error': 'Name parameter is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = EmployeeMaster.objects.get(name=name)
            serializer = CardEmployeeMasterSerializer(employee)
            return Response(serializer.data)
        except EmployeeMaster.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)







# from rest_framework.response import Response
# from rest_framework import status
# from rest_framework.views import APIView

# from .models import (
#     EmployeeMaster,
#     OperatorSkill,
#     Score,
#     MultiSkilling,
#     RefreshmentTraining
# )

# from .serializers import (
#     CardEmployeeMasterSerializer,
#     OperatorCardSkillSerializer,
#     CardScoreSerializer,
#     CardMultiSkillingSerializer,
#     CardRefreshmentTrainingSerializer
# )

# class EmployeeCardDetailsView(APIView):
#     def get(self, request):
#         card_no = request.query_params.get('card_no')
#         if not card_no:
#             return Response({'error': 'card_no parameter is required'}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             employee = EmployeeMaster.objects.get(card_no=card_no)
#         except EmployeeMaster.DoesNotExist:
#             return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)

#         # Fetch and serialize all related data
#         employee_data = CardEmployeeMasterSerializer(employee).data
#         operator_skills = OperatorCardSkillSerializer(OperatorSkill.objects.filter(operator=employee), many=True).data
#         scores = CardScoreSerializer(Score.objects.filter(employee=employee), many=True).data
#         multi_skilling = CardMultiSkillingSerializer(MultiSkilling.objects.filter(employee=employee), many=True).data
#         refreshment_training = CardRefreshmentTrainingSerializer(RefreshmentTraining.objects.filter(employee=employee), many=True).data

#         # Construct full response
#         response_data = {
#             'employee': employee_data,
#             'operator_skills': operator_skills,
#             'scores': scores,
#             'multi_skilling': multi_skilling,
#             'refreshment_training': refreshment_training,
#         }

#         # Print to console
#         print("==== Employee Card Details ====")
#         import pprint
#         pprint.pprint(response_data)  # pretty-print for readability
#         print("================================")

#         return Response(response_data)




from rest_framework.response import Response
from rest_framework import status
from rest_framework.views import APIView

from .models import (
    EmployeeMaster,
    OperatorSkill,
    Score,
    MultiSkilling,
    RefreshmentTraining
)

from .serializers import (
    CardEmployeeMasterSerializer,
    OperatorCardSkillSerializer,
    CardScoreSerializer,
    CardMultiSkillingSerializer,
    CardRefreshmentTrainingSerializer,
    CardHanchouExamResultSerializer,  # <-- 2. IMPORT the new serializer
    CardTrainingAttendanceSerializer,
    CardScheduleSerializer
)

class EmployeeCardDetailsView(APIView):
    def get(self, request):
        card_no = request.query_params.get('card_no')
        if not card_no:
            return Response({'error': 'card_no parameter is required'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            employee = EmployeeMaster.objects.get(card_no=card_no)
        except EmployeeMaster.DoesNotExist:
            return Response({'error': 'Employee not found'}, status=status.HTTP_404_NOT_FOUND)

        # Fetch and serialize all related data
        employee_data = CardEmployeeMasterSerializer(employee).data
        operator_skills = OperatorCardSkillSerializer(OperatorSkill.objects.filter(operator=employee), many=True).data
        scores = CardScoreSerializer(Score.objects.filter(employee=employee), many=True).data
        # multi_skilling = CardMultiSkillingSerializer(MultiSkilling.objects.filter(employee=employee), many=True).data
        # refreshment_training = CardRefreshmentTrainingSerializer(RefreshmentTraining.objects.filter(employee=employee), many=True).data
        hanchou_results = CardHanchouExamResultSerializer(
            HanchouExamResult.objects.filter(employee=employee),
            many=True
        ).data
        shokuchou_results = CardShokuchouExamResultSerializer(
            ShokuchouExamResult.objects.filter(employee=employee),
            many=True
        ).data

        # --- ★ UPDATED SCHEDULE LOGIC ★ ---
        
        # 1. Get ALL schedules for this employee, REMOVE the category filter
        all_schedules = Schedule.objects.filter(employees=employee)

        # 2. Use the renamed serializer
        serializer_context = {'employee': employee}
        scheduled_trainings_data = CardScheduleSerializer(
            all_schedules, 
            many=True, 
            context=serializer_context
        ).data

        # --- ★ END: REVISED LOGIC ★ ---




        attendances_data = [] # Default to an empty list
        employee_name = employee.name

        try:
            # Step 1: Find all UserInfo records that match the employee's name.
            # We use __iexact for a case-insensitive match, which is safer.
            # This might return multiple users if names are not unique.
            user_info_records = UserInfo.objects.filter(first_name__iexact=employee_name)

            if user_info_records.exists():
                # Step 2: Use the found UserInfo records to query for their attendance.
                # The `user__in` lookup efficiently finds attendance for all matched users.
                attendances = TrainingAttendance.objects.filter(user__in=user_info_records)
                
                # Step 3: Serialize the data
                attendances_data = CardTrainingAttendanceSerializer(attendances, many=True).data

        except Exception as e:
            # If anything goes wrong, we'll log it and continue with empty attendance
            print(f"Error fetching attendance for {employee_name}: {e}")
            attendances_data = []

        # Construct full response
        response_data = {
            'employee': employee_data,
            'operator_skills': operator_skills,
            'scores': scores,
            # 'multi_skilling': multi_skilling,
            'scheduled_trainings': scheduled_trainings_data,
            'hanchou_results': hanchou_results, 
            'shokuchou_results': shokuchou_results,
            'attendance': attendances_data,
        }

        # Print to console
        print("==== Employee Card Details ====")
        import pprint
        pprint.pprint(response_data)  # pretty-print for readability
        print("================================")

        return Response(response_data)






 
# employeereportpdfview
from django.http import Http404, HttpResponse, JsonResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from io import BytesIO
import json
import traceback
from .models import EmployeeMaster, OperatorSkill, Score, MultiSkilling, RefreshmentTraining

@method_decorator(csrf_exempt, name='dispatch')
class EmployeeReportPDFView(View):
    def post(self, request, *args, **kwargs):
        """
        Handle PDF generation requests
        Accepts both form data and JSON input
        """
        try:
            print("\n=== Received PDF generation request ===")
            
            # 1. Parse input data
            card_no = self._get_card_number(request)
            if not card_no:
                return JsonResponse({'error': 'card_no is required'}, status=400)
            print(f"Processing card_no: {card_no}")

            # 2. Get employee record
            try:
                employee = EmployeeMaster.objects.get(card_no=card_no)
                print(f"Found employee: {employee.name}")
            except EmployeeMaster.DoesNotExist:
                print(f"Employee not found for card_no: {card_no}")
                return JsonResponse({'error': 'Employee not found'}, status=404)

            # 3. Generate PDF content
            print("Generating PDF content...")
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            story = self.create_pdf_content(employee)
            
            # 4. Build PDF document
            print("Building PDF document...")
            doc.build(story)
            buffer.seek(0)
            print("PDF generation completed successfully")

            # 5. Return PDF response
            response = HttpResponse(
                buffer.getvalue(), 
                content_type='application/pdf'
            )
            response['Content-Disposition'] = (
                f'attachment; filename="employee_report_{card_no}.pdf"'
            )
            return response
            
        except Exception as e:
            print("\n!!! PDF generation failed !!!")
            traceback.print_exc()
            return JsonResponse(
                {
                    'error': 'Internal server error',
                    'detail': str(e),
                    'traceback': traceback.format_exc()
                }, 
                status=500
            )

    def _get_card_number(self, request):
        """Helper method to extract card_no from request"""
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                return data.get('card_no')
            except json.JSONDecodeError:
                return None
        return request.POST.get('card_no')

    def create_pdf_content(self, employee):
        """Generate the PDF content structure"""
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        story.append(Paragraph(f"Employee Comprehensive Report", styles['Title']))
        story.append(Spacer(1, 12))
        
        # Add all sections
        self._add_basic_info(story, styles, employee)
        self._add_operator_skills(story, styles, employee)
        self._add_scores(story, styles, employee)
        self._add_multi_skills(story, styles, employee)
        self._add_refreshment_training(story, styles, employee)
        
        return story


    def _get_table_style(self):
        """Returns a consistent, professional style for all tables"""
        return TableStyle([
            # Header styling
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4682B4')),  # Steel blue
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('TOPPADDING', (0,0), (-1,0), 4),
            
            # Data row styling
            ('ALIGN', (0,1), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E0E0E0')),
            
            # Zebra striping
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F8F8')]),
            
            # Cell padding
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ])

    def _add_basic_info(self, story, styles, employee):
        """Employee basic information with improved layout"""
        story.append(Paragraph("Basic Information", styles['Heading2']))
        story.append(Spacer(1, 8))
        
        basic_data = [
            ["Field", "Value"],
            ["Name:", employee.name],
            ["Card No:", employee.card_no],
            ["Department:", employee.department],
            ["Section:", employee.section],
            ["Designation:", employee.desig_category],
            ["Joining Date:", employee.joining_date.strftime('%Y-%m-%d') if employee.joining_date else "N/A"],
            ["Gender:", employee.sex],
            ["Birth Date:", employee.birth_date.strftime('%Y-%m-%d') if employee.birth_date else "N/A"],
            ["Guardian:", employee.guardian_name or "N/A"]
        ]
        
        basic_table = Table(basic_data, colWidths=[150, 300])
        style = TableStyle([
            ('BACKGROUND', (0,0), (1,0), colors.HexColor('#4682B4')),
            ('TEXTCOLOR', (0,0), (1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (1,0), 'CENTER'),
            ('FONTNAME', (0,0), (1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (1,0), 10),
            ('ALIGN', (0,1), (0,-1), 'LEFT'),
            ('FONTNAME', (0,1), (0,-1), 'Helvetica-Bold'),
            ('ALIGN', (1,1), (1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E0E0E0')),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ])
        basic_table.setStyle(style)
        story.append(basic_table)
        story.append(Spacer(1, 24))

    def _add_operator_skills(self, story, styles, employee):
        """Operator skills with professional table design"""
        skills = OperatorSkill.objects.filter(operator=employee).select_related('station')
        if not skills.exists():
            return
            
        story.append(Paragraph("Operator Skills", styles['Heading2']))
        story.append(Spacer(1, 8))
        
        skill_data = [["Station", "Skill Level"]]
        for skill in skills:
            skill_data.append([
                str(skill.station),
                skill.skill_level or "N/A"
            ])
        
        skills_table = Table(skill_data, colWidths=[250, 100])
        skills_table.setStyle(self._get_table_style())
        story.append(skills_table)
        story.append(Spacer(1, 24))


    def _add_scores(self, story, styles, employee):
        """Scores table with proper result formatting"""
        scores = Score.objects.filter(employee=employee)
        if not scores.exists():
            return
            
        story.append(Paragraph("Scores and Assessments", styles['Heading2']))
        story.append(Spacer(1, 8))
        
        score_data = [["Test", "Marks", "%", "Result", "Date"]]
        for score in scores:
            score_data.append([
                score.test_name or "N/A",
                str(score.marks) if score.marks is not None else "N/A",
                f"{score.percentage}%" if score.percentage is not None else "N/A",
                "Pass" if score.passed else "Fail",
                score.created_at.strftime("%d-%b-%Y") if score.created_at else "N/A"
            ])
        
        scores_table = Table(score_data, colWidths=[150, 60, 60, 80, 80])
        style = self._get_table_style()
        style.add('ALIGN', (1,1), (2,-1), 'RIGHT')
        
        # Add result coloring
        for row in range(1, len(score_data)):
            result = scores[row-1].passed  
            for style_command in self._get_result_badge_style(result, row):
                style.add(*style_command)
        
        scores_table.setStyle(style)
        story.append(scores_table)
        story.append(Spacer(1, 24))

    def _add_multi_skills(self, story, styles, employee):
        """Multi-skilling with proper status formatting"""
        multi_skills = MultiSkilling.objects.filter(employee=employee).select_related('station', 'skill_level')
        if not multi_skills.exists():
            return
            
        story.append(Paragraph("Multi-Skilling", styles['Heading2']))
        story.append(Spacer(1, 8))
        
        multi_data = [["Skill", "Status", "Station", "Level", "Start", "End"]]
        for skill in multi_skills:
            multi_data.append([
                skill.skill or "N/A",
                skill.status.capitalize() if skill.status else "N/A",
                str(skill.station) if skill.station else "N/A",
                skill.skill_level.skill_level if skill.skill_level else "N/A",
                skill.start_date.strftime('%d-%b-%Y') if skill.start_date else "N/A",
                skill.end_date.strftime('%d-%b-%Y') if skill.end_date else "N/A"
            ])
        
        multi_table = Table(multi_data, colWidths=[120, 80, 100, 60, 80, 80])
        style = self._get_table_style()
        
        # Add status coloring
        for row in range(1, len(multi_data)):
            status = multi_skills[row-1].status
            for style_command in self._get_status_badge_style(status, row):
                style.add(*style_command)
        
        multi_table.setStyle(style)
        story.append(multi_table)
        story.append(Spacer(1, 24))

    def _get_result_badge_style(self, result, row):
        """Returns properly structured style commands for result badges"""
        if isinstance(result, bool):
            color = colors.green if result else colors.red
        else:
            color = colors.green if str(result).lower() == "pass" else colors.red
        
        # Return a complete style command tuple
        return [
            ('TEXTCOLOR', (3, row), (3, row), color),
            ('FONTNAME', (3, row), (3, row), 'Helvetica-Bold')
        ]

    def _get_status_badge_style(self, status, row):
        """Returns properly structured style commands for status badges"""
        status_colors = {
            'active': colors.green,
            'completed': colors.blue,
            'inprogress': colors.orange,
            'scheduled': colors.purple,
            'inactive': colors.gray,
            'rescheduled': colors.darkblue, 
        }
        color = status_colors.get(status.lower(), colors.black)
        return [
            ('TEXTCOLOR', (1, row), (1, row), color),
            ('FONTNAME', (1, row), (1, row), 'Helvetica-Bold')
        ]

    def _add_refreshment_training(self, story, styles, employee):
        """Refreshment training with professional layout"""
        trainings = RefreshmentTraining.objects.filter(employee=employee).select_related(
            'station', 'skill', 'skill_level'
        )
        if not trainings.exists():
            return
            
        story.append(Paragraph("Refreshment Training", styles['Heading2']))
        story.append(Spacer(1, 8))
        
        training_data = [["Skill", "Station", "Level", "Start Date", "End Date", "Reason"]]
        for training in trainings:
            training_data.append([
                training.skill.skill if training.skill else "N/A",
                str(training.station) if training.station else "N/A",
                training.skill_level.skill_level if training.skill_level else "N/A",
                training.start_date.strftime('%d-%b-%Y') if training.start_date else "N/A",
                training.end_date.strftime('%d-%b-%Y') if training.end_date else "N/A",
                training.reason_for_refreshment or "N/A"
            ])
        
        training_table = Table(training_data, colWidths=[120, 100, 60, 80, 80, 150])
        style = self._get_table_style()
        style.add('ALIGN', (5,1), (5,-1), 'LEFT') 
        training_table.setStyle(style)
        story.append(training_table)



from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import MainDepartment
from .serializers import MainDepartmentSerializer

class MainDepartmentListView(APIView):
    def get(self, request):
        departments = MainDepartment.objects.all()
        serializer = MainDepartmentSerializer(departments, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)






from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import MainLine
from .serializers import MainLineByDepartmentSerializer

class MainLinesByDepartmentView(APIView):
    def get(self, request, department_id):
        main_lines = MainLine.objects.filter(department_id=department_id)
        serializer = MainLineByDepartmentSerializer(main_lines, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)




from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import SubLine
from .serializers import SubLineByMainLineSerializer,StationByLineSerializer

class SubLinesByMainLineView(APIView):
    def get(self, request, main_line_id):
        sub_lines = SubLine.objects.filter(main_line_id=main_line_id)
        serializer = SubLineByMainLineSerializer(sub_lines, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)







class StationsBySubLineView(APIView):
    def get(self, request, sub_line_id):
        stations = Station.objects.filter(sub_line_id=sub_line_id)
        serializer = StationByLineSerializer(stations, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)



from django.http import HttpResponse, JsonResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from io import BytesIO
import json
import traceback
from .models import EmployeeMaster, OperatorSkill, Score, MultiSkilling, RefreshmentTraining

@method_decorator(csrf_exempt, name='dispatch')
class EmployeeReportPDFView(View):
    def post(self, request, *args, **kwargs):
        """
        Handle PDF generation requests
        Accepts both form data and JSON input
        """
        try:
            print("\n=== Received PDF generation request ===")
            
            # 1. Parse input data
            card_no = self._get_card_number(request)
            if not card_no:
                return JsonResponse({'error': 'card_no is required'}, status=400)
            print(f"Processing card_no: {card_no}")

            # 2. Get employee record
            try:
                employee = EmployeeMaster.objects.get(card_no=card_no)
                print(f"Found employee: {employee.name}")
            except EmployeeMaster.DoesNotExist:
                print(f"Employee not found for card_no: {card_no}")
                return JsonResponse({'error': 'Employee not found'}, status=404)

            # 3. Generate PDF content
            print("Generating PDF content...")
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=50, bottomMargin=50, leftMargin=50, rightMargin=50)
            story = self.create_pdf_content(employee)
            
            # 4. Build PDF document
            print("Building PDF document...")
            doc.build(story)
            buffer.seek(0)
            print("PDF generation completed successfully")

            # 5. Return PDF response
            response = HttpResponse(
                buffer.getvalue(), 
                content_type='application/pdf'
            )
            response['Content-Disposition'] = (
                f'attachment; filename="employee_report_{card_no}.pdf"'
            )
            return response
            
        except Exception as e:
            print("\n!!! PDF generation failed !!!")
            traceback.print_exc()
            return JsonResponse(
                {
                    'error': 'Internal server error',
                    'detail': str(e),
                    'traceback': traceback.format_exc()
                }, 
                status=500
            )

    def _get_card_number(self, request):
        """Helper method to extract card_no from request"""
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                return data.get('card_no')
            except json.JSONDecodeError:
                return None
        return request.POST.get('card_no')

    def create_pdf_content(self, employee):
        """Generate the PDF content structure"""
        styles = getSampleStyleSheet()
        story = []
        
        # Title with proper alignment
        title = Paragraph(f"Employee History Report", styles['Title'])
        title.alignment = 1  # Center alignment
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Add all sections
        self._add_basic_info(story, styles, employee)
        self._add_operator_skills(story, styles, employee)
        self._add_scores(story, styles, employee)
        self._add_multi_skills(story, styles, employee)
        self._add_refreshment_training(story, styles, employee)
        
        return story

    def _get_base_table_style(self):
        """Returns the base table style that all tables will use"""
        return TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4682B4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            
            # Data row styling
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),  # Default all data to left alignment
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            
            # Grid and borders
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
            
            # Zebra striping
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F8F8')]),
            
            # Consistent padding for all cells
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ])

    def _add_section_heading(self, story, styles, heading_text):
        """Add a consistent section heading"""
        heading = Paragraph(heading_text, styles['Heading2'])
        heading.alignment = 0  # Left alignment
        story.append(heading)
        story.append(Spacer(1, 10))

    def _add_basic_info(self, story, styles, employee):
        """Employee basic information with consistent layout"""
        self._add_section_heading(story, styles, "Basic Information")
        
        # Prepare data
        basic_data = [
            ["Field", "Value"],
            ["Name", employee.name or "N/A"],
            ["Card No", employee.card_no or "N/A"],
            ["Department", employee.department or "N/A"],
            ["Section", employee.section or "N/A"],
            ["Designation", employee.desig_category or "N/A"],
            ["Joining Date", employee.joining_date.strftime('%d-%b-%Y') if employee.joining_date else "N/A"],
            ["Gender", employee.sex or "N/A"],
            ["Birth Date", employee.birth_date.strftime('%d-%b-%Y') if employee.birth_date else "N/A"],
            ["Guardian", employee.guardian_name or "N/A"]
        ]
        
        # Create table with consistent width (500 points total)
        basic_table = Table(basic_data, colWidths=[150, 350])
        
        # Apply base style
        style = self._get_base_table_style()
        # Make field column bold
        style.add('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold')
        
        basic_table.setStyle(style)
        story.append(basic_table)
        story.append(Spacer(1, 20))

    def _add_operator_skills(self, story, styles, employee):
        """Operator skills with consistent layout"""
        skills = OperatorSkill.objects.filter(operator=employee).select_related('station')
        if not skills.exists():
            return
            
        self._add_section_heading(story, styles, "Operator Skills")
        
        # Prepare data
        skill_data = [["Station", "Skill Level"]]
        for skill in skills:
            skill_data.append([
                str(skill.station) if skill.station else "N/A",
                skill.skill_level or "N/A"
            ])
        
        # Create table with consistent width (500 points total)
        skills_table = Table(skill_data, colWidths=[300, 200])
        
        # Apply base style
        style = self._get_base_table_style()
        # Center align skill level column
        style.add('ALIGN', (1, 1), (1, -1), 'CENTER')
        
        skills_table.setStyle(style)
        story.append(skills_table)
        story.append(Spacer(1, 20))

    def _add_scores(self, story, styles, employee):
        """Scores table with consistent layout"""
        scores = Score.objects.filter(employee=employee)
        if not scores.exists():
            return
            
        self._add_section_heading(story, styles, "Scores and Assessments")
        
        # Prepare data
        score_data = [["Test Name", "Marks", "Percentage", "Result", "Date"]]
        for score in scores:
            score_data.append([
                score.test_name or "N/A",
                str(score.marks) if score.marks is not None else "N/A",
                f"{score.percentage}%" if score.percentage is not None else "N/A",
                "Pass" if score.passed else "Fail",
                score.created_at.strftime("%d-%b-%Y") if score.created_at else "N/A"
            ])
        
        # Create table with consistent width (500 points total)
        scores_table = Table(score_data, colWidths=[150, 70, 90, 70, 120])
        
        # Apply base style
        style = self._get_base_table_style()
        # Right align numerical columns
        style.add('ALIGN', (1, 1), (1, -1), 'RIGHT')   # Marks
        style.add('ALIGN', (2, 1), (2, -1), 'RIGHT')   # Percentage
        # Center align result and date
        style.add('ALIGN', (3, 1), (3, -1), 'CENTER')  # Result
        style.add('ALIGN', (4, 1), (4, -1), 'CENTER')  # Date
        
        # Add result coloring
        for row in range(1, len(score_data)):
            if row - 1 < len(scores):
                result = scores[row - 1].passed
                color = colors.green if result else colors.red
                style.add('TEXTCOLOR', (3, row), (3, row), color)
                style.add('FONTNAME', (3, row), (3, row), 'Helvetica-Bold')
        
        scores_table.setStyle(style)
        story.append(scores_table)
        story.append(Spacer(1, 20))


    def _add_multi_skills(self, story, styles, employee):
        """Multi-skilling with consistent layout"""
        multi_skills = MultiSkilling.objects.filter(employee=employee).select_related('station', 'skill_level')
        if not multi_skills.exists():
            return
            
        self._add_section_heading(story, styles, "Multi-Skilling")
        
        # Prepare data with proper column distribution
        multi_data = [["Skill", "Status", "Station", "Level", "Start Date", "End Date"]]
        for skill in multi_skills:
            # Ensure skill name is clean and status is separate
            skill_name = (skill.skill or "N/A").strip()
            status = (skill.status or "N/A").strip().capitalize()
            
            multi_data.append([
                skill_name,
                status,
                str(skill.station) if skill.station else "N/A",
                skill.skill_level.skill_level if skill.skill_level else "N/A",
                skill.start_date.strftime('%d-%b-%Y') if skill.start_date else "N/A",
                skill.end_date.strftime('%d-%b-%Y') if skill.end_date else "N/A"
            ])
        
        # Create table with adjusted column widths
        # Total width should be around 500 (letter width minus margins)
        col_widths = [120, 80, 80, 50, 90, 80]  # Total = 500
        
        multi_table = Table(multi_data, colWidths=col_widths)
        
        # Apply base style
        style = self._get_base_table_style()
        # Center align status, level, and dates
        style.add('ALIGN', (1, 1), (1, -1), 'CENTER')  # Status
        style.add('ALIGN', (3, 1), (3, -1), 'CENTER')  # Level
        style.add('ALIGN', (4, 1), (4, -1), 'CENTER')  # Start Date
        style.add('ALIGN', (5, 1), (5, -1), 'CENTER')  # End Date
        
        # Add word wrapping for all columns
        style.add('WORDWRAP', (0, 0), (-1, -1), True)
        
        # Add status coloring
        for row in range(1, len(multi_data)):
            if row - 1 < len(multi_skills):
                status = multi_skills[row - 1].status
                if status:
                    color = self._get_status_color(status.lower())
                    style.add('TEXTCOLOR', (1, row), (1, row), color)
                    style.add('FONTNAME', (1, row), (1, row), 'Helvetica-Bold')
        
        multi_table.setStyle(style)
        story.append(multi_table)
        story.append(Spacer(1, 20))

    def _add_refreshment_training(self, story, styles, employee):
        """Refreshment training with consistent layout"""
        trainings = RefreshmentTraining.objects.filter(employee=employee).select_related(
            'station', 'skill', 'skill_level'
        )
        if not trainings.exists():
            return
            
        self._add_section_heading(story, styles, "Refreshment Training")
        
        # Prepare data
        training_data = [["Skill", "Station", "Level", "Start Date", "End Date", "Reason"]]
        for training in trainings:
            training_data.append([
                training.skill.skill if training.skill else "N/A",
                str(training.station) if training.station else "N/A",
                training.skill_level.skill_level if training.skill_level else "N/A",
                training.start_date.strftime('%d-%b-%Y') if training.start_date else "N/A",
                training.end_date.strftime('%d-%b-%Y') if training.end_date else "N/A",
                training.reason_for_refreshment or "N/A"
            ])
        
        # Create table with consistent width (500 points total)
        training_table = Table(training_data, colWidths=[90, 90, 60, 80, 80, 100])
        
        # Apply base style
        style = self._get_base_table_style()
        # Center align level and dates
        style.add('ALIGN', (2, 1), (2, -1), 'CENTER')  # Level
        style.add('ALIGN', (3, 1), (3, -1), 'CENTER')  # Start Date
        style.add('ALIGN', (4, 1), (4, -1), 'CENTER')  # End Date
        
        training_table.setStyle(style)
        story.append(training_table)

    def _get_status_color(self, status):
        """Get appropriate color for status"""
        status_colors = {
            'active': colors.green,
            'completed': colors.blue,
            'inprogress': colors.orange,
            'in progress': colors.orange,
            'scheduled': colors.purple,
            'inactive': colors.red,
            'rescheduled': colors.navy,
        }
        return status_colors.get(status.lower(), colors.black)
    

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models import EmployeeMaster
from .serializers import EmployeeSerializer
from datetime import datetime

class EmployeeExcelViewSet(viewsets.ModelViewSet):
    queryset = EmployeeMaster.objects.all()
    serializer_class = EmployeeSerializer
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        
        queryset = self.get_queryset()

        wb = Workbook()
        ws = wb.active
        ws.title = "EMPLOYEE MASTER"

        ws.merge_cells('A1:S1')
        company_cell = ws['A1']
        company_cell.value = "Company Name: KRISHNA MARUTI SEAT -JOSHI SAI ENTERPRISES, NEHA ENTERPRISES, MADHU ENTERPRISES, Amar Infosoft Private Limited,"
        company_cell.font = Font(bold=True, size=10)
        company_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells('A2:S2')
        run_date_cell = ws['A2']
        run_date_cell.value = f"Run Date & Time: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        run_date_cell.font = Font(size=10)
        run_date_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells('A4:S4')
        title_cell = ws['A4']
        title_cell.value = "EMPLOYEE MASTER"
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        headers = [
            'Srl. No.',
            'Pay Code',
            'Card No.',
            'Sex',
            'Birth Date',
            'Name',
            'Guardian\'s Name',
            'Department',
            'Section',
            'Desig/Category',
            'Joining Date',
            'Auth Shift',
            'Shift Type',
            'Shift Pattern',
            '1st Weekly Off',
            '2nd Weekly Off',
            '2nd Weekly Off',
            'OT Allowed/Rate',
            'Round the Clock'
        ]
        header_font = Font(bold=True, size=10)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.row_dimensions[6].height = 30
        
        for row_num, employee in enumerate(queryset, 7):
            birth_date = employee.birth_date.strftime('%d/%m/%Y') if employee.birth_date else ''
            joining_date = employee.joining_date.strftime('%d/%m/%Y') if employee.joining_date else ''
            
            row_data = [
                row_num - 6,  # Serial number starting from 1
                employee.pay_code,
                employee.card_no,
                employee.sex,
                birth_date,
                employee.name.upper() if employee.name else '',
                employee.guardian_name.upper() if employee.guardian_name else '',
                employee.department,
                employee.section,
                employee.desig_category or '',
                joining_date,
                employee.auth_shift,
                employee.shift_type,
                employee.shift_pattern,
                employee.first_weekly_off,
                employee.second_weekly_off or '',
                employee.second_weekly_off_fh or '',
                'Y' if employee.ot_allowed_rate else 'N',
                'Y' if employee.round_the_clock else 'N'
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="center" if col_num in [1, 3, 4, 18, 19] else "left", 
                    vertical="center"
                )
                cell.font = Font(size=9)
        
        # Set column widths
        column_widths = {
            'A': 6, 'B': 10, 'C': 10, 'D': 4, 'E': 12, 'F': 15, 'G': 15, 
            'H': 20, 'I': 20, 'J': 15, 'K': 12, 'L': 10, 'M': 10, 'N': 12, 
            'O': 12, 'P': 12, 'Q': 12, 'R': 10, 'S': 12
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'EMPLOYEE_MASTER_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        
        return response









 # management review dashboard
# management review dashboard
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.utils import timezone
from datetime import datetime
from .models import ManagementReview
from .serializers import (TrainingDataSerializer, DefectsDataSerializer,OperatorsChartSerializer, TrainingPlansChartSerializer,DefectsChartSerializer)

class CurrentMonthTrainingDataView(APIView):
    def get(self, request):
        current_month = timezone.now().replace(day=1)
        try:
            data = ManagementReview.objects.get(
                month_year__year=current_month.year,
                month_year__month=current_month.month
            )
            serializer = TrainingDataSerializer(data)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except ManagementReview.DoesNotExist:
            return Response(
                {"message": "No data found for current month"},
                status=status.HTTP_404_NOT_FOUND
            )

class CurrentMonthDefectsDataView(APIView):
    def get(self, request):
        current_month = timezone.now().replace(day=1)
        data = (
            ManagementReview.objects
            .filter(
                month_year__year=current_month.year,
                month_year__month=current_month.month
            )
            .order_by('-id')
            .first()
        )
        if not data:
            return Response(
                {"message": "No data found for current month"},
                status=status.HTTP_404_NOT_FOUND
            )
        serializer = DefectsDataSerializer(data)
        return Response(serializer.data, status=status.HTTP_200_OK)

class PreviousMonthDefectsDataView(APIView):
    def get(self, request):
        current_month = timezone.now().date().replace(day=1)
        # Compute previous month with year wrap-around
        if current_month.month == 1:
            prev_year = current_month.year - 1
            prev_month = 12
        else:
            prev_year = current_month.year
            prev_month = current_month.month - 1
        data = (
            ManagementReview.objects
            .filter(
                month_year__year=prev_year,
                month_year__month=prev_month
            )
            .order_by('-id')
            .first()
        )
        if not data:
            return Response(
                {"message": "No data found for previous month"},
                status=status.HTTP_404_NOT_FOUND
            )
        serializer = DefectsDataSerializer(data)
        return Response(serializer.data, status=status.HTTP_200_OK)

class NextMonthDefectsDataView(APIView):
    def get(self, request):
        current_month = timezone.now().date().replace(day=1)
        # Compute next month with year wrap-around
        if current_month.month == 12:
            next_year = current_month.year + 1
            next_month = 1
        else:
            next_year = current_month.year
            next_month = current_month.month + 1
        data = (
            ManagementReview.objects
            .filter(
                month_year__year=next_year,
                month_year__month=next_month
            )
            .order_by('-id')
            .first()
        )
        if not data:
            return Response(
                {"message": "No data found for next month"},
                status=status.HTTP_404_NOT_FOUND
            )
        serializer = DefectsDataSerializer(data)
        return Response(serializer.data, status=status.HTTP_200_OK)

class OperatorsChartView(APIView):
    def get(self, request):
        current_year = timezone.now().year
        data = ManagementReview.objects.filter(
            month_year__year=current_year
        ).order_by('month_year')
        serializer = OperatorsChartSerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class TrainingPlansChartView(APIView):
    def get(self, request):
        current_year = timezone.now().year
        data = ManagementReview.objects.filter(
            month_year__year=current_year
        ).order_by('month_year')
        serializer = TrainingPlansChartSerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class DefectsChartView(APIView):
    def get(self, request):
        current_year = timezone.now().year
        data = ManagementReview.objects.filter(
            month_year__year=current_year
        ).order_by('month_year')
        serializer = DefectsChartSerializer(data, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)





# views.py

# import pandas as pd
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# from .models import AdvancedManpowerCTQ
# from .serializers import AdvancedManpowerCTQUploadSerializer

# class AdvancedManpowerCTQUploadView(APIView):
#     def post(self, request):
#         serializer = AdvancedManpowerCTQUploadSerializer(data=request.data)
#         if serializer.is_valid():
#             excel_file = serializer.validated_data['file']
#             try:
#                 # Read Excel with header at row 1
#                 df = pd.read_excel(excel_file, header=1)
#                 df.columns = df.columns.str.strip()  # Remove trailing/leading spaces

#                 for _, row in df.iterrows():
#                     AdvancedManpowerCTQ.objects.update_or_create(
#                         month_year_ctq=pd.to_datetime(row['Month & Year']).date(),
#                         defaults={
#                             'total_stations_ctq': int(row['Total Stations']),
#                             'operator_required_ctq': int(row['Operator Required']),
#                             'operator_availability_ctq': int(row['Operator Availability']),
#                             'buffer_manpower_required_ctq': int(row['Buffer Man Power Required']),
#                             'buffer_manpower_availability_ctq': int(row['Buffer Man Power Availability']),
#                             'attrition_trend_ctq': int(row['Attrition Trend']),
#                             'absentee_trend_ctq': int(row['Absentee Trend']),
#                             'planned_units_ctq': int(row['Planned Units']),
#                             'actual_production_ctq': int(row['Actual Production']),
#                         }
#                     )
#                 return Response({"message": "Data uploaded successfully."}, status=status.HTTP_201_CREATED)

#             except Exception as e:
#                 return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)




import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import ManagementReview
from .serializers import ManagementReviewUploadSerializer

class ManagementReviewUploadView(APIView):
    def post(self, request):
        serializer = ManagementReviewUploadSerializer(data=request.data)
        if serializer.is_valid():
            excel_file = serializer.validated_data['file']
            try:
                df = pd.read_excel(excel_file, header=1)
                df.columns = df.columns.str.strip()  # 🧽 Clean trailing/leading spaces from column names

                for _, row in df.iterrows():
                    ManagementReview.objects.update_or_create(
                        month_year=pd.to_datetime(row['Month & Year']).date(),
                        defaults={
                            'new_operators_joined': int(row['New Operators Joined']),
                            'new_operators_trained': int(row['New Operators Trained']),
                            'total_training_plans': int(row['Total Training Plans']),
                            'total_trainings_actual': int(row['Total Trainings Actual']),
                            'total_defects_msil': int(row['Total Defects at MSIL']),
                            'ctq_defects_msil': int(row['CTQ Defects at MSIL']),
                            'total_defects_tier1': int(row['Total Defects at Tier-1']),
                            'ctq_defects_tier1': int(row['CTQ Defects at Tier-1']),  # fixed trailing space issue
                            'total_internal_rejection': int(row['Total Internal Rejection']),
                            'ctq_internal_rejection': int(row['CTQ Internal Rejection']),
                        }
                    )
                return Response({"message": "Data uploaded successfully."}, status=status.HTTP_201_CREATED)

            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



from rest_framework import viewsets
from .models import ManagementReview
from .serializers import ManagementReviewSerializer

class ManagementReviewViewSet(viewsets.ModelViewSet):
    queryset = ManagementReview.objects.all().order_by('-month_year')
    serializer_class = ManagementReviewSerializer

from rest_framework import viewsets
from .models import CompanyLogo
from .serializers import CompanyLogoSerializer

class CompanyLogoViewSet(viewsets.ModelViewSet):
    queryset = CompanyLogo.objects.all()
    serializer_class = CompanyLogoSerializer



from rest_framework import generics
from .models import SubLine
from .serializers import SubLineSerializer

class SubLineListAPIView(generics.ListAPIView):
    queryset = SubLine.objects.all()
    serializer_class = SubLineSerializer




import pandas as pd
from django.db import transaction
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import (
    OperatorSkill, EmployeeMaster, Station,
    SubLine, MainLine, MainDepartment
)

@csrf_exempt
def upload_operator_skills(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            df = pd.read_excel(file)

            failed_rows = []

            with transaction.atomic():
                for index, row in df.iterrows():
                    try:
                        pay_code = str(row['pay_code']).strip()
                        department_name = str(row['department']).strip()
                        main_line_name = str(row['main_line']).strip()
                        sub_line_name = str(row['sub_line']).strip()
                        skill_name = str(row['skill']).strip()
                        skill_level = str(row['skill_level']).strip()

                        # Get Employee
                        employee = EmployeeMaster.objects.get(pay_code=pay_code)

                        # Get Station via nested lookups
                        department = MainDepartment.objects.get(name=department_name)
                        main_line = MainLine.objects.get(name=main_line_name, department=department)
                        sub_line = SubLine.objects.get(name=sub_line_name, main_line=main_line)
                        station = Station.objects.get(skill=skill_name, sub_line=sub_line)

                        # Create or Update OperatorSkill
                        OperatorSkill.objects.update_or_create(
                            operator=employee,
                            station=station,
                            defaults={
                                'skill_level': skill_level
                            }
                        )

                    except Exception as e:
                        failed_rows.append({
                            'row_index': index + 2,
                            'error': str(e)
                        })
                        # Trigger rollback if any row fails
                        raise

            return JsonResponse({
                'status': 'success',
                'message': f"{df.shape[0]} records uploaded successfully.",
                'failed': []
            })

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': 'Upload failed. No records were saved.',
                'error': str(e),
                'failed': failed_rows
            })

    return JsonResponse({'status': 'error', 'message': 'Invalid request. Upload a file using POST method.'})











class StationDeleteView(APIView):
    def delete(self, request, pk):
        station = get_object_or_404(Station, pk=pk)
        station.delete()
        return Response({"message": "Station deleted successfully."}, status=status.HTTP_204_NO_CONTENT)

class StationUpdateView(APIView):
    def put(self, request, pk):
        station = get_object_or_404(Station, pk=pk)
        serializer = StationSerializer(station, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status

class MachineAllocationApprovalViewSet(viewsets.ModelViewSet):
    queryset = MachineAllocation.objects.all()
    serializer_class = MachineAllocationApprovalSerializer

    @action(detail=True, methods=['put'], url_path='set-status')
    def set_status(self, request, pk=None):
        allocation = self.get_object()
        status_value = request.data.get('approval_status')

        if status_value not in dict(MachineAllocation.APPROVAL_STATUS_CHOICES):
            return Response({'error': 'Invalid status'}, status=status.HTTP_400_BAD_REQUEST)

        allocation.approval_status = status_value
        allocation.save()
        return Response({
            'status': 'success',
            'id': allocation.id,
            'approval_status': allocation.approval_status
        })

    @action(detail=True, methods=['put'], url_path='reject')
    def reject(self, request, pk=None):
        allocation = self.get_object()
        allocation.approval_status = 'rejected'
        allocation.save()
        return Response({
            'status': 'rejected',
            'id': allocation.id,
            'approval_status': allocation.approval_status
        }, status=status.HTTP_200_OK)





from .serializers import EmployeeWithStatusSerializer

class EmployeeMachineAllocationViewSet(viewsets.ModelViewSet):
    queryset = MachineAllocation.objects.all()
    serializer_class = ...  # your main MachineAllocation serializer

    @action(detail=False, methods=['get'], url_path='eligible-employees')
    def eligible_employees(self, request):
        machine_id = request.query_params.get('machine_id')
        if not machine_id:
            return Response({'error': 'machine_id is required'}, status=400)

        try:
            machine = Machine.objects.get(id=machine_id)
        except Machine.DoesNotExist:
            return Response({'error': 'Machine not found'}, status=404)

        matching_skills = OperatorSkill.objects.filter(station__skill=machine.process)
        employee_ids = matching_skills.values_list('operator_id', flat=True).distinct()
        employees = EmployeeMaster.objects.filter(id__in=employee_ids)

        serializer = EmployeeWithStatusSerializer(employees, many=True, context={'machine_id': machine_id})
        return Response(serializer.data)
    
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Department
from .serializers import DepartmentSerializer

class DepartmentByFactoryView(APIView):
    def get(self, request):
        factory_id = request.query_params.get('factory')
        if not factory_id:
            return Response({"error": "Factory ID is required"}, status=status.HTTP_400_BAD_REQUEST)

        departments = Department.objects.filter(factory_id=factory_id)
        serializer = DepartmentSerializer(departments, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

from rest_framework import viewsets
from .models import AdvancedManpowerCTQ
from .serializers import NewAdvancedManpowerCTQSerializer

class NewAdvancedManpowerCTQViewSet(viewsets.ModelViewSet):
    queryset = AdvancedManpowerCTQ.objects.all().order_by('-month_year_ctq')
    serializer_class = NewAdvancedManpowerCTQSerializer





from rest_framework.views import APIView
from rest_framework.response import Response
from datetime import datetime
from .models import AdvancedManpowerCTQ, OperatorRequirement
from .serializers import (
    AdvancedManpowerCTQSerializer,
    OperatorTrendSerializer,
    BufferManpowerTrendSerializer,
    AttritionTrendSerializer,
    AbsenteeTrendSerializer,
    OperatorRequirementSerializer
)

class ManpowerCTQTrendsView(APIView):
    def get(self, request):
        plant = request.query_params.get('plant')
        factory_id = request.query_params.get('factory_id')
        department_id = request.query_params.get('department_id')
        today = datetime.today()

        # ===== CTQ Queryset =====
        ctq_queryset = AdvancedManpowerCTQ.objects.all().order_by('month_year_ctq')
        if plant:
            ctq_queryset = ctq_queryset.filter(plant=plant)
        if factory_id:
            ctq_queryset = ctq_queryset.filter(factory_id=factory_id)
        if department_id:
            ctq_queryset = ctq_queryset.filter(department_id=department_id)

        # Current month data
        current_month_ctq = ctq_queryset.filter(
            month_year_ctq__year=today.year,
            month_year_ctq__month=today.month
        )

        # ===== OperatorRequirement Queryset =====
        operator_queryset = OperatorRequirement.objects.all().order_by('-month')
        if factory_id:
            operator_queryset = operator_queryset.filter(factory_id=factory_id)
        if department_id:
            operator_queryset = operator_queryset.filter(department_id=department_id)

        # ===== Build Flat Response =====
        return Response({
            "current_month": AdvancedManpowerCTQSerializer(current_month_ctq, many=True).data,
            "operator_trend": OperatorTrendSerializer(ctq_queryset, many=True).data,
            "buffer_trend": BufferManpowerTrendSerializer(ctq_queryset, many=True).data,
            "attrition_trend": AttritionTrendSerializer(ctq_queryset, many=True).data,
            "absentee_trend": AbsenteeTrendSerializer(ctq_queryset, many=True).data,
            "operator_requirements": OperatorRequirementSerializer(operator_queryset, many=True).data,
        })
    
from rest_framework import viewsets
from .models import OperatorRequirement
from .serializers import OperatorRequirementSerializer

class OperatorRequirementViewSet(viewsets.ModelViewSet):
    queryset = OperatorRequirement.objects.all().order_by('-month')
    serializer_class = OperatorRequirementSerializer


import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import AdvancedManpowerCTQ, Factory, Department, HQ
from .serializers import AdvancedManpowerCTQSerializer

class UploadAdvancedManpowerCTQView(APIView):
    def post(self, request):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(excel_file)

            for _, row in df.iterrows():
                factory_name = row['factory_name']
                department_name = row['department_name']

                # Create or get dummy HQ (if not already assigned)
                hq, _ = HQ.objects.get_or_create(name='Default HQ')

                factory, _ = Factory.objects.get_or_create(name=factory_name, defaults={'hq': hq})
                department, _ = Department.objects.get_or_create(name=department_name, factory=factory)

                # Direct creation using model (instead of serializer if name usage)
                AdvancedManpowerCTQ.objects.create(
                    month_year_ctq=row['month_year_ctq'],
                    total_stations_ctq=row['total_stations_ctq'],
                    operator_required_ctq=row['operator_required_ctq'],
                    operator_availability_ctq=row['operator_availability_ctq'],
                    buffer_manpower_required_ctq=row['buffer_manpower_required_ctq'],
                    buffer_manpower_availability_ctq=row['buffer_manpower_availability_ctq'],
                    attrition_trend_ctq=row['attrition_trend_ctq'],
                    absentee_trend_ctq=row['absentee_trend_ctq'],
                    planned_units_ctq=row['planned_units_ctq'],
                    actual_production_ctq=row['actual_production_ctq'],
                    factory=factory,
                    department=department
                )

            return Response({"message": "Excel data uploaded successfully."}, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)








import os
from rest_framework.views import APIView
from rest_framework.response import Response
from django.conf import settings

class LogoView(APIView):
    def get(self, request):
        image_path = 'logos/nl_logo.jpg'
        full_url = request.build_absolute_uri(settings.MEDIA_URL + image_path)
        return Response({'logo_url': full_url})

# views.py
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import FileResponse
from .models import UploadedFile
from .serializers import UploadedFileSerializer
import os

class UploadedFileListView(APIView):
    def get(self, request):
        files = UploadedFile.objects.all().order_by('-uploaded_at')
        serializer = UploadedFileSerializer(files, many=True)
        return Response(serializer.data)

class FileDownloadView(APIView):
    def get(self, request, file_id):
        try:
            uploaded_file = UploadedFile.objects.get(id=file_id)
            file_path = uploaded_file.file.path
            response = FileResponse(open(file_path, 'rb'))
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
            return response
        except UploadedFile.DoesNotExist:
            return Response({'error': 'File not found'}, status=404)









from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Score
from .serializers import LevelOneScoreSerializer

class LevelOneEmployeesView(APIView):
    def get(self, request):
        level_one_scores = Score.objects.filter(level="Level 1").select_related('employee')
        serializer = LevelOneScoreSerializer(level_one_scores, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)





from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import Score
from .serializers import LevelTwoGroupedEmployeeScoreSerializer
from collections import defaultdict

class GroupedScoreByEmployeeView(APIView):
    def get(self, request):
        scores = Score.objects.filter(level="Level 2").select_related('employee', 'skill')

        grouped_data = defaultdict(list)
        for score in scores:
            key = (score.employee.id, score.employee.name)
            grouped_data[key].append(score)

        result = []
        for (employee_id, employee_name), score_list in grouped_data.items():
            result.append({
                "employee_id": employee_id,
                "employee_name": employee_name,
                "scores": score_list
            })

        serializer = LevelTwoGroupedEmployeeScoreSerializer(result, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)





from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from collections import defaultdict
from .models import Score
from .serializers import LevelThreeGroupedEmployeeScoreSerializer

class GroupedLevelThreeScoreByEmployeeView(APIView):
    def get(self, request):
        scores = Score.objects.filter(level="Level 3").select_related('employee', 'skill')

        grouped_data = defaultdict(list)
        for score in scores:
            key = (score.employee.id, score.employee.name)
            grouped_data[key].append(score)

        result = []
        for (employee_id, employee_name), score_list in grouped_data.items():
            result.append({
                "employee_id": employee_id,
                "employee_name": employee_name,
                "scores": score_list
            })

        serializer = LevelThreeGroupedEmployeeScoreSerializer(result, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)











# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# from django_celery_beat.models import PeriodicTask, CrontabSchedule
# import json
# from datetime import datetime
# from django.utils import timezone
# import pytz

# class SetTaskTimeView(APIView):
#     def post(self, request):
#         time_str = request.data.get("time")  # Expecting format like "02:30 PM"

#         if not time_str:
#             return Response({"error": "Time is required (e.g., 02:30 PM)."}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             # Convert 12-hour format with AM/PM to 24-hour format
#             time_obj = datetime.strptime(time_str.strip(), "%I:%M %p")
#             hour = time_obj.hour
#             minute = time_obj.minute
#         except ValueError:
#             return Response({"error": "Invalid time format. Use HH:MM AM/PM."}, status=status.HTTP_400_BAD_REQUEST)

#         # Delete old schedules for this task to avoid conflicts
#         old_task = PeriodicTask.objects.filter(name="daily_import_excel").first()
#         if old_task and old_task.crontab:
#             old_crontab = old_task.crontab
#             old_task.delete()
#             # Check if any other tasks use this schedule
#             if not PeriodicTask.objects.filter(crontab=old_crontab).exists():
#                 old_crontab.delete()

#         # Create new schedule
#         schedule, created = CrontabSchedule.objects.get_or_create(
#             minute=str(minute),
#             hour=str(hour),
#             day_of_week='*',
#             day_of_month='*',
#             month_of_year='*',
#         )

#         # Create new task
#         task = PeriodicTask.objects.create(
#             name="daily_import_excel",
#             crontab=schedule,
#             task='app1.tasks.import_attendance_from_excel',
#             enabled=True,
#             args=json.dumps([]),
#         )

#         # Get current time in Asia/Kolkata timezone
#         kolkata_tz = pytz.timezone('Asia/Kolkata')
#         current_kolkata_time = timezone.now().astimezone(kolkata_tz)

#         return Response({
#             "success": True,
#             "message": f"Task scheduled daily at {time_str} (24-hour: {hour:02}:{minute:02})",
#             "scheduled_time": f"{hour:02}:{minute:02}",
#             "current_time_utc": timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
#             "current_time_kolkata": current_kolkata_time.strftime('%Y-%m-%d %H:%M:%S'),
#             "timezone": "Asia/Kolkata",
#             "task_enabled": task.enabled
#         })
    



def schedule_task(task_name, task_path, time_str):
    from datetime import datetime
    from django.utils import timezone
    import pytz
    from django_celery_beat.models import PeriodicTask, CrontabSchedule
    import json

    try:
        dt = datetime.strptime(time_str.strip(), "%I:%M %p")
        hour, minute = dt.hour, dt.minute
    except ValueError:
        return None, "Invalid time format. Use HH:MM AM/PM."

    existing_task = PeriodicTask.objects.filter(name=task_name).first()
    if existing_task:
        old_cron = existing_task.crontab
        existing_task.delete()
        if not PeriodicTask.objects.filter(crontab=old_cron).exists():
            old_cron.delete()

    cron, _ = CrontabSchedule.objects.get_or_create(
        minute=str(minute), hour=str(hour),
        day_of_week='*', day_of_month='*', month_of_year='*',
    )

    task = PeriodicTask.objects.create(
        name=task_name,
        crontab=cron,
        task=task_path,
        args=json.dumps([]),
        enabled=True
    )

    kolkata = timezone.now().astimezone(pytz.timezone('Asia/Kolkata'))
    return {
        "message": f"{task_name} scheduled at {time_str} (24-hour: {hour:02}:{minute:02})",
        "current_kolkata_time": kolkata.strftime('%Y-%m-%d %H:%M:%S'),
        "task_enabled": task.enabled
    }, None





from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

class SetAttendanceTaskTimeView(APIView):
    def post(self, request):
        time_str = request.data.get("time")
        data, error = schedule_task("daily_import_excel", "app1.tasks.import_attendance_from_excel", time_str)
        if error:
            return Response({"error": error}, status=status.HTTP_400_BAD_REQUEST)
        
        response_data = {"success": True, **data}
        print(response_data)  # <-- Print the response in console
        return Response(response_data)

class SetManagementReviewTaskTimeView(APIView):
    def post(self, request):
        time_str = request.data.get("time")
        data, error = schedule_task("daily_import_management_review", "app1.tasks.import_management_review_from_excel", time_str)
        if error:
            return Response({"error": error}, status=status.HTTP_400_BAD_REQUEST)
        
        response_data = {"success": True, **data}
        print(response_data)  # <-- Print the response in console
        return Response(response_data)

class SetAdvancedManpowerTaskTimeView(APIView):
    def post(self, request):
        time_str = request.data.get("time")
        data, error = schedule_task("daily_import_advanced_manpower", "app1.tasks.import_advanced_manpower_from_excel", time_str)
        if error:
            return Response({"error": error}, status=status.HTTP_400_BAD_REQUEST)
        
        response_data = {"success": True, **data}
        print(response_data)  # <-- Print the response in console
        return Response(response_data)












# import pandas as pd
# from rest_framework.views import APIView
# from rest_framework.response import Response
# from rest_framework import status
# from .models import EmployeeMaster, Station, OperatorSkill


# from django.db.models import Max

# class UploadOperatorSkillsAPIView(APIView):
#     def post(self, request):
#         file = request.FILES.get('file')
#         if not file:
#             return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             df = pd.read_excel(file)

#             fixed_columns = ["S.No.", "Operator Name", "Code", "Date of Joining"]
#             station_columns = [col for col in df.columns if col not in fixed_columns]

#             skipped_employees = []
#             skipped_stations = []

#             from django.db.models import Func

#             class Trim(Func):
#                 function = 'TRIM'

#             # --- Step 1: Remove existing duplicates ---
#             duplicates = (
#                 OperatorSkill.objects
#                 .values('operator_id', 'station_id')
#                 .annotate(latest_id=Max('id'))
#             )
#             for dup in duplicates:
#                 OperatorSkill.objects.filter(
#                     operator_id=dup['operator_id'],
#                     station_id=dup['station_id']
#                 ).exclude(id=dup['latest_id']).delete()

#             # --- Step 2: Process the file ---
#             for _, row in df.iterrows():
#                 operators = EmployeeMaster.objects.annotate(
#                     trimmed_name=Trim('name')
#                 ).filter(trimmed_name__iexact=str(row['Operator Name']).strip())

#                 if not operators.exists():
#                     skipped_employees.append(str(row['Operator Name']))
#                     continue

#                 for operator in operators:
#                     for station_name in station_columns:
#                         skill_level = row.get(station_name)
#                         if pd.isna(skill_level) or skill_level == '':
#                             continue

#                         stations = Station.objects.filter(skill__iexact=station_name.strip())
#                         if not stations.exists():
#                             skipped_stations.append(station_name)
#                             continue

#                         for station in stations:
#                             OperatorSkill.objects.update_or_create(
#                                 operator=operator,
#                                 station=station,
#                                 defaults={'skill_level': str(skill_level).strip()}
#                             )

#             return Response({
#                 "message": "Skills updated successfully!",
#                 "skipped_employees": skipped_employees,
#                 "skipped_stations": list(set(skipped_stations))
#             }, status=status.HTTP_200_OK)

#         except Exception as e:
#             return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


import pandas as pd
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Max
from .models import EmployeeMaster, Station, OperatorSkill


class UploadOperatorSkillsAPIView(APIView):
    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(file)

            fixed_columns = ["S.No.", "Operator Name", "Code", "Date of Joining"]
            station_columns = [col for col in df.columns if col not in fixed_columns]

            skipped_employees = []
            skipped_stations = []

            # --- Step 1: Remove existing duplicates ---
            duplicates = (
                OperatorSkill.objects
                .values('operator_id', 'station_id')
                .annotate(latest_id=Max('id'))
            )
            for dup in duplicates:
                OperatorSkill.objects.filter(
                    operator_id=dup['operator_id'],
                    station_id=dup['station_id']
                ).exclude(id=dup['latest_id']).delete()

            # --- Step 2: Process the file ---
            for _, row in df.iterrows():
                # Normalize Excel code (remove leading zeros)
                pay_code_excel = str(row['Code']).strip().lstrip("0")

                # Also normalize DB pay_code (remove leading zeros before comparing)
                operator = None
                for emp in EmployeeMaster.objects.all():
                    if emp.pay_code.strip().lstrip("0") == pay_code_excel:
                        operator = emp
                        break

                if not operator:
                    skipped_employees.append(pay_code_excel)
                    continue

                # Process station skill levels
                for station_name in station_columns:
                    skill_level = row.get(station_name)
                    if pd.isna(skill_level) or skill_level == '':
                        continue

                    stations = Station.objects.filter(skill__iexact=station_name.strip())
                    if not stations.exists():
                        skipped_stations.append(station_name)
                        continue

                    for station in stations:
                        OperatorSkill.objects.update_or_create(
                            operator=operator,
                            station=station,
                            defaults={'skill_level': str(skill_level).strip()}
                        )

            return Response({
                "message": "Skills updated successfully!",
                "skipped_employees": skipped_employees,
                "skipped_stations": list(set(skipped_stations))
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# level 0

from rest_framework import generics, status
from rest_framework.response import Response
from .models import UserInfo
from .serializers import UserInfoSerializer
from rest_framework.parsers import MultiPartParser, FormParser

class UserInfoListCreateView(generics.ListCreateAPIView):
    queryset = UserInfo.objects.all()
    parser_classes = [MultiPartParser, FormParser]

    def get_serializer_class(self):
        # Use create serializer for POST, update serializer for GET (list)
        if self.request.method == 'POST':
            return UserInfoSerializer
        return UserInfoUpdateSerializer  # Or a read-only serializer if you want

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        try:
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            print("❌ Validation or save error:", str(e))
            print("📝 Request data:", request.data)
            return Response(
                {"detail": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

from rest_framework import generics
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

class UserInfoUpdateView(generics.UpdateAPIView):
    queryset = UserInfo.objects.all()
    serializer_class = UserInfoUpdateSerializer
    lookup_field = 'temp_id'  # PATCH/PUT by temp_id
    parser_classes = [MultiPartParser, FormParser, JSONParser]  # <-- Add JSONParser
    

    
from rest_framework import generics, status
from rest_framework.response import Response
from .models import HumanBodyCheck
from .serializers import HumanBodyCheckSerializer

class HumanBodyCheckListCreateView(generics.ListCreateAPIView):
    queryset = HumanBodyCheck.objects.all()
    serializer_class = HumanBodyCheckSerializer
    
    def get_queryset(self):
        temp_id = self.request.query_params.get('temp_id')
        if temp_id:
            return HumanBodyCheck.objects.filter(temp_id=temp_id)
        return HumanBodyCheck.objects.all()
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Calculate overall status
        data = serializer.validated_data
        checks = [
            data.get('color_vision', 'pending'),
            data.get('eye_movement', 'pending'),
            data.get('fingers_functionality', 'pending'),
            data.get('hand_deformity', 'pending'),
            data.get('joint_mobility', 'pending'),
            data.get('hearing', 'pending'),
            data.get('bending_ability', 'pending')
        ]
        
        # Check additional checks if any
        additional_checks = data.get('additional_checks', [])
        for check in additional_checks:
            checks.append(check.get('status', 'pending'))
        
        # Determine overall status
        if 'fail' in checks:
            data['overall_status'] = 'fail'
        elif all(status == 'pass' for status in checks):
            data['overall_status'] = 'pass'
        else:
            data['overall_status'] = 'pending'
        
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
    
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import UserInfo, HumanBodyCheck
from .serializers import FetchUserInfoSerializer

class PassedUsersWithDetailsView(APIView):
    def get(self, request):
        name = request.GET.get('name', '').strip()
        if name:
            users = UserInfo.objects.filter(first_name__icontains=name)
        else:
            users = UserInfo.objects.all()

        passing_users = [user for user in users if HumanBodyCheck.objects.filter(temp_id=user.temp_id, overall_status='pass').exists()]
        serializer = FetchUserInfoSerializer(passing_users, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import UserInfo, HumanBodyCheck
from .serializers import FetchUserInfoSerializer

class PassedUsersWithDetailsView(APIView):
    def get(self, request):
        name = request.GET.get('name', '').strip()
        if not name:
            return Response({"error": "Name query parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

        users = UserInfo.objects.filter(
            first_name__icontains=name
        )

        # Filter users with a passing HumanBodyCheck
        passing_users = [user for user in users if HumanBodyCheck.objects.filter(temp_id=user.temp_id, overall_status='pass').exists()]

        serializer = FetchUserInfoSerializer(passing_users, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import UserInfo, HumanBodyCheck
from .serializers import FetchUserInfoSerializer

class AllPassedUsersView(APIView):
    def get(self, request):
        # Get temp_ids of all users who passed the human body check
        passed_temp_ids = HumanBodyCheck.objects.filter(overall_status='pass').values_list('temp_id', flat=True).distinct()

        # Filter user info for those passed temp_ids
        passed_users = UserInfo.objects.filter(temp_id__in=passed_temp_ids)

        serializer = FetchUserInfoSerializer(passed_users, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


from rest_framework import viewsets
from .models import SDCOrientationFeedback
from .serializers import SDCOrientationFeedbackSerializer

class SDCOrientationFeedbackViewSet(viewsets.ModelViewSet):
    queryset = SDCOrientationFeedback.objects.all()
    serializer_class = SDCOrientationFeedbackSerializer




from rest_framework.views import APIView
from rest_framework.response import Response
from .models import UserInfo
from .serializers import ListUserInfoWithBodyCheckSerializer

class UserInfoBodyCheckListView(APIView):
    def get(self, request):
        users = UserInfo.objects.all()
        serializer = ListUserInfoWithBodyCheckSerializer(users, many=True)
        return Response(serializer.data)





from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import SDCOrientationFeedback, UserInfo

class RemainingDepartmentsView(APIView):
    def get(self, request):
        temp_id = request.GET.get('temp_id')

        if not temp_id:
            return Response({"error": "temp_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        # Check if user exists
        try:
            user = UserInfo.objects.get(temp_id=temp_id)
        except UserInfo.DoesNotExist:
            return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)

        # List of all departments
        all_departments = [dept[0] for dept in SDCOrientationFeedback.DEPARTMENT_CHOICES]

        # Departments already submitted for this user
        submitted_departments = SDCOrientationFeedback.objects.filter(user=user).values_list('department', flat=True)

        # Remaining departments
        remaining_departments = list(set(all_departments) - set(submitted_departments))

        return Response({
            "temp_id": temp_id,
            "user": f"{user.first_name} {user.last_name}",
            "remaining_departments": remaining_departments
        }, status=status.HTTP_200_OK)
    

class ActiveTrainingBatchListView(generics.ListAPIView):
    """
    GET /api/training-batches/active/
    Returns a list of all training batches where `is_active` is True.
    """
    queryset = TrainingBatch.objects.filter(is_active=True)
    serializer_class = TrainingBatchSerializer

class BatchAttendanceDetailView(APIView):
    """
    GET /api/attendance-detail/{batch_id}/
    A powerful view that returns everything the frontend needs for the attendance page.
    """
    def get(self, request, batch_id, *args, **kwargs):
        # 1. Get all users for the batch
        users = UserInfo.objects.filter(batch_id=batch_id)
        if not users.exists():
            return Response({"error": "No users found for this batch."}, status=status.HTTP_404_NOT_FOUND)

        # 2. Determine the next training day to mark
        today = timezone.now().date()
        next_day_to_mark = None
        
        # Find the latest attendance record for this batch to see when they last marked
        latest_attendance = TrainingAttendance.objects.filter(batch_id=batch_id).order_by('-day_number').first()

        if not latest_attendance:
            # No attendance has ever been marked, so today is Day 1
            next_day_to_mark = 1
        else:
            last_marked_day = latest_attendance.day_number
            last_marked_date = latest_attendance.attendance_date
            
            if last_marked_day >= 6:
                # Training is complete
                next_day_to_mark = None
            elif today > last_marked_date:
                # It's a new day, so unlock the next training day
                next_day_to_mark = last_marked_day + 1
            else:
                # Attendance was already marked today, so no day is available to mark.
                next_day_to_mark = None
        
        # Check if the batch is completed (all 6 days are done)
        is_completed = TrainingAttendance.objects.filter(batch_id=batch_id, day_number=6).exists()

        # 3. Prepare the data payload
        data = {
            'batch_id': batch_id,
            'next_training_day_to_mark': next_day_to_mark,
            'is_completed': is_completed,
            'users': users
        }

        # 4. Serialize the data
        serializer = BatchAttendanceDetailSerializer(data, context={'batch_id': batch_id})
        return Response(serializer.data, status=status.HTTP_200_OK)
    


class PastTrainingBatchListView(generics.ListAPIView):
    """
    GET /api/training-batches/past/
    Returns a list of all training batches where `is_active` is False.
    """
    queryset = TrainingBatch.objects.filter(is_active=False)
    serializer_class = TrainingBatchSerializer


class BulkAttendanceUpdateView(APIView):
    """
    POST /api/attendance/
    Creates or updates attendance records, but now with date and day validation.
    """
    def post(self, request, *args, **kwargs):
        attendance_data = request.data
        if not isinstance(attendance_data, list) or not attendance_data:
            return Response({"error": "Expected a non-empty list of attendance objects."}, status=status.HTTP_400_BAD_REQUEST)

        # All items in a single submission must be for the same batch and day
        target_batch_id = attendance_data[0].get('batch')
        target_day = attendance_data[0].get('day_number')
        today = timezone.now().date()
        
        # --- Validation Logic ---
        # Check if this batch already has an attendance record for today
        if TrainingAttendance.objects.filter(batch_id=target_batch_id, attendance_date=today).exists():
            return Response({"error": f"Attendance for batch {target_batch_id} has already been submitted today."}, status=status.HTTP_400_BAD_REQUEST)

        # ... (You can add more validation here if needed) ...

        response_data = []
        errors = []
        for item in attendance_data:
            item['attendance_date'] = today  # Add today's date to each record
            
            try:
                # Use update_or_create to handle both cases smoothly
                attendance_obj, created = TrainingAttendance.objects.update_or_create(
                    user_id=item.get('user'),
                    batch_id=item.get('batch'),
                    day_number=item.get('day_number'),
                    defaults={'status': item.get('status'), 'attendance_date': today}
                )
                serializer = TrainingAttendanceSerializer(attendance_obj)
                response_data.append(serializer.data)
            except Exception as e:
                errors.append({"item": item, "error": str(e)})

        if errors:
            return Response({"success": response_data, "errors": errors}, status=status.HTTP_400_BAD_REQUEST)
        
        return Response(response_data, status=status.HTTP_201_CREATED)
    

class CompleteTrainingBatchView(APIView):
    """
    POST /api/batches/{batch_id}/complete/
    Marks a training batch as inactive (is_active=False).
    """
    def post(self, request, batch_id, *args, **kwargs):
        try:
            batch = TrainingBatch.objects.get(batch_id=batch_id)
            batch.is_active = False
            batch.save()
            serializer = TrainingBatchSerializer(batch)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except TrainingBatch.DoesNotExist:
            return Response({"error": "Batch not found."}, status=status.HTTP_404_NOT_FOUND)



# Refreshment Training


#views.py


from rest_framework import viewsets
from .models import Training_category, Curriculum, CurriculumContent, Trainer_name, Venues, Schedule, RescheduleLog,EmployeeAttendance
from .serializers import Training_categorySerializer, CurriculumSerializer, CurriculumContentSerializer, Trainer_nameSerializer, VenuesSerializer, ScheduleSerializer,EmployeeAttendanceSerializer, RescheduleLogSerializer

class Training_categoryViewSet(viewsets.ModelViewSet):
    queryset = Training_category.objects.all()
    serializer_class = Training_categorySerializer

class CurriculumViewSet(viewsets.ModelViewSet):
    serializer_class = CurriculumSerializer

    def get_queryset(self):
        queryset = Curriculum.objects.all()
        category_id = self.request.query_params.get('category_id')
        if category_id is not None:
            queryset = queryset.filter(category_id=category_id)
        return queryset

class CurriculumContentViewSet(viewsets.ModelViewSet):
    serializer_class = CurriculumContentSerializer

    def get_queryset(self):
        queryset = CurriculumContent.objects.all()
        curriculum_id = self.request.query_params.get('curriculum')
        if curriculum_id is not None:
            queryset = queryset.filter(curriculum_id=curriculum_id)
        return queryset

class Trainer_nameViewSet(viewsets.ModelViewSet):
    queryset = Trainer_name.objects.all()
    serializer_class = Trainer_nameSerializer

class VenueViewSet(viewsets.ModelViewSet):
    queryset = Venues.objects.all()
    serializer_class = VenuesSerializer

class ScheduleViewSet(viewsets.ModelViewSet):
    queryset = Schedule.objects.all()
    serializer_class = ScheduleSerializer

from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action

from .models import EmployeeAttendance, RescheduleLog
from .serializers import EmployeeAttendanceSerializer, RescheduleLogSerializer


class EmployeeAttendanceViewSet(viewsets.ModelViewSet):
    queryset = EmployeeAttendance.objects.all()
    serializer_class = EmployeeAttendanceSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        attendance_instance = serializer.save()

        print(f"Attendance created: status={attendance_instance.status}")

        if attendance_instance.status == 'rescheduled':
            if not (attendance_instance.reschedule_date and attendance_instance.reschedule_time and attendance_instance.reschedule_reason):
                print("Missing reschedule details; skipping RescheduleLog creation")
            else:
                try:
                    RescheduleLog.objects.create(
                        schedule=attendance_instance.schedule,
                        employee=attendance_instance.employee,
                        original_date=attendance_instance.schedule.date,
                        original_time=attendance_instance.schedule.time,
                        new_date=attendance_instance.reschedule_date,
                        new_time=attendance_instance.reschedule_time,
                        reason=attendance_instance.reschedule_reason,
                    )
                    print("RescheduleLog created")
                except Exception as e:
                    print("Error creating RescheduleLog:", e)

        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'])
    def by_schedule(self, request):
        schedule_id = request.query_params.get('schedule_id')
        if not schedule_id:
            return Response({"detail": "schedule_id query parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

        attendances = self.queryset.filter(schedule_id=schedule_id)
        serializer = self.get_serializer(attendances, many=True)
        return Response(serializer.data)




from rest_framework import viewsets
from rest_framework.response import Response

from .models import RescheduleLog
from .serializers import RescheduleLogSerializer


class RescheduleLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Viewset to retrieve Reschedule Logs (Read-only).
    Optional filtering by schedule or employee.
    """
    queryset = RescheduleLog.objects.all()
    serializer_class = RescheduleLogSerializer

    def list(self, request, *args, **kwargs):
        schedule_id = request.query_params.get('schedule_id')
        employee_id = request.query_params.get('employee_id')

        queryset = self.queryset

        if schedule_id:
            queryset = queryset.filter(schedule_id=schedule_id)

        if employee_id:
            queryset = queryset.filter(employee_id=employee_id)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

# # multiskilling

# class MultiSkillingViewSet(viewsets.ModelViewSet):
#     queryset = MultiSkilling.objects.all()
#     serializer_class = MultiSkillingSerializer



# @api_view(['GET'])
# def employees_with_active_skills(request):
#     name_query = request.GET.get('name', '')
#     operators = OperatorMaster.objects.filter(full_name__icontains=name_query)
#     result = []
#     for operator in operators:
#         active_skills = MultiSkilling.objects.filter(employee=operator, status='active')
#         skills = [
#             {
#                 "skill_level": skill.skill_level,
#                 "date": skill.date,
#                 "notes": skill.notes,
#                 "status": skill.status,
#                 "department": skill.department.department if skill.department else None,
#                 "section": skill.section.name if skill.section else None,
#                 "operation": skill.operation.name if skill.operation else None,
#             }
#             for skill in active_skills
#         ]
#         result.append({
#             "employee_id": operator.id,
#             "employee_code": operator.employee_code,
#             "full_name": operator.full_name,
#             "designation": operator.designation,
#             "department": operator.department,
#             "department_code": operator.department_code,
#             "date_of_join": operator.date_of_join,
#             "employee_pattern_category": operator.employment_pattern_category,
#             "skills": skills
#         })
#     return Response(result, status=status.HTTP_200_OK)    





from rest_framework import viewsets
from .models import NewMultiSkilling
from .serializers import NewMultiSkillingSerializer
from django.db.models import Count, Q



class NewMultiSkillingViewSet(viewsets.ModelViewSet):
    queryset = NewMultiSkilling.objects.all()
    serializer_class = NewMultiSkillingSerializer

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        Provides aggregate counts for multiskilling statuses.
        """
        queryset = self.get_queryset()

        # This code will now work because Count and Q are imported
        stats_data = queryset.aggregate(
            scheduled=Count('id', filter=Q(status='scheduled')),
            in_progress=Count('id', filter=Q(status='inprogress')),
            completed=Count('id', filter=Q(status='completed')),
            total=Count('id')
        )
        
        # Ensure all keys are present even if count is zero
        stats_data.setdefault('scheduled', 0)
        stats_data.setdefault('in_progress', 0)
        stats_data.setdefault('completed', 0)
        stats_data.setdefault('total', 0)

        return Response(stats_data)

# views.py
import logging
import traceback
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from django.db import transaction
from .models import EmployeeMaster, QuestionPaper, Question, Score, Station, TestSession

# Configure logger
logger = logging.getLogger(__name__)

# views.py
import logging
import traceback
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404
from django.db import transaction
from .models import (
    EmployeeMaster, QuestionPaper, Question, Score, Station, TestSession, OperatorSkill
)

# Configure logger
logger = logging.getLogger(__name__)

class SubmitWebTestAPIView(APIView):
    """
    Submit test answers from web/tablet exam (without remote).
    """
    def post(self, request):
        try:
            logger.info("SubmitWebTestAPIView called with data: %s", request.data)

            employee_id = request.data.get("employee_id")
            test_name = request.data.get("test_name")
            question_paper_id = request.data.get("question_paper_id")
            answers = request.data.get("answers", [])
            skill_id = request.data.get("skill_id")
            level_name = request.data.get("level_id")  # This is the level name string

            # Validate required fields
            if not employee_id or not test_name or not question_paper_id or not isinstance(answers, list):
                logger.warning("Validation failed: missing required fields.")
                return Response(
                    {"error": "employee_id, test_name, question_paper_id and answers[] are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Get employee and question paper
            logger.debug("Fetching employee_id=%s", employee_id)
            employee = get_object_or_404(EmployeeMaster, id=employee_id)

            logger.debug("Fetching question_paper_id=%s", question_paper_id)
            question_paper = get_object_or_404(QuestionPaper, id=question_paper_id)

            # Handle skill
            skill = None
            if skill_id:
                try:
                    skill_id_int = int(skill_id)
                    logger.debug("Fetching skill by id=%s", skill_id_int)
                    skill = get_object_or_404(Station, id=skill_id_int)
                except (ValueError, TypeError):
                    logger.debug("Fetching skill by name=%s", skill_id)
                    skill = get_object_or_404(Station, name=skill_id)

            # Handle level
            level_name_str = str(level_name) if level_name else None
            logger.info("Level received: %s", level_name_str)

            # Get all questions
            questions = list(Question.objects.filter(question_paper=question_paper).order_by("id"))
            total_questions = len(questions)
            logger.info("Total questions found: %s", total_questions)

            if total_questions == 0:
                logger.error("No questions found for paper id=%s", question_paper_id)
                return Response({"error": "No questions found for this paper."}, status=400)

            # Compare answers
            correct_count = 0
            for i, ans in enumerate(answers):
                if i < total_questions:
                    logger.debug("Q%s: submitted=%s, correct=%s",
                                 i+1, ans, questions[i].correct_index)
                    if ans == questions[i].correct_index:
                        correct_count += 1

            percentage = round((correct_count / total_questions) * 100, 2)
            passed = percentage >= 80
            logger.info("Scoring: %s/%s correct (%.2f%%), passed=%s",
                        correct_count, total_questions, percentage, passed)

            # ✅ Create or fetch a TestSession
            test_session, _ = TestSession.objects.get_or_create(
                test_name=test_name,
                key_id=f"{employee_id}-{question_paper_id}",  # unique session key
                employee=employee,
                defaults={
                    "level": level_name_str,
                    "skill": skill,
                    "question_paper": question_paper,
                }
            )

            # ✅ Save score linked to TestSession
            with transaction.atomic():
                score, created = Score.objects.get_or_create(
                    employee=employee,
                    test_name=test_name,
                    test=test_session,
                    defaults={
                        'marks': correct_count,
                        'percentage': percentage,
                        'passed': passed,
                        'skill': skill,
                        'level': level_name_str,
                    }
                )

                if not created:
                    if correct_count > score.marks:
                        logger.info("Updating existing score for employee=%s", employee.id)
                        score.marks = correct_count
                        score.percentage = percentage
                        score.passed = passed
                        score.skill = skill
                        score.level = level_name_str
                        score.save()

                # 🚀 Add to OperatorSkill if Level 1 passed
                if passed and level_name_str == "Level 1":
                    general_station, _ = Station.objects.get_or_create(skill="General")
                    OperatorSkill.objects.get_or_create(
                        operator=employee,
                        station=general_station,
                        defaults={"skill_level": "Level 1"}
                    )
                    logger.info("Employee %s added to General station in OperatorSkill", employee.id)

            return Response({
                "employee": employee.name,
                "marks": correct_count,
                "total_questions": total_questions,
                "percentage": percentage,
                "passed": passed,
                "level_received": level_name_str,
                "message": "Score saved successfully"
            }, status=200)

        except Exception as e:
            error_trace = traceback.format_exc()
            logger.error("Error in SubmitWebTestAPIView: %s\n%s", str(e), error_trace)
            return Response({
                "error": str(e),
                "traceback": error_trace
            }, status=500)

# hanchou & shokucho 



# your_app/views.py

from rest_framework import viewsets
from .models import HanContent, HanSubtopic, HanTrainingContent
from .serializers import (
    HanContentDetailSerializer,
    HanContentListSerializer,
    HanSubtopicSerializer,
    HanTrainingContentSerializer
)

class HanContentViewSet(viewsets.ModelViewSet):
    """
    Handles listing, creating, retrieving, updating, and deleting Main Topics.
    """
    # prefetch_related is a performance optimization that prevents many small database queries.
    queryset = HanContent.objects.prefetch_related('subtopics__materials').all()

    def get_serializer_class(self):
        """
        Chooses the serializer based on the action.
        - For retrieving a single item ('retrieve'), use the detailed serializer.
        - For listing all items ('list'), use the simple serializer.
        """
        if self.action == 'retrieve':
            return HanContentDetailSerializer
        return HanContentListSerializer


# --- MODIFIED VIEWSET FOR SUBTOPICS ---
class HanSubtopicViewSet(viewsets.ModelViewSet):
    serializer_class = HanSubtopicSerializer

    def get_queryset(self):
        """
        Filters subtopics based on a query parameter from the URL.
        Example: GET /api/subtopics/?han_content_id=5
        """
        queryset = HanSubtopic.objects.all()
        han_content_id = self.request.query_params.get('han_content_id')
        if han_content_id:
            queryset = queryset.filter(han_content_id=han_content_id)
        return queryset

    def perform_create(self, serializer):
        """
        When creating, the parent ID must be sent in the request body.
        Example: POST /api/subtopics/ with body {"title": "...", "han_content": 5}
        """
        # The serializer will handle associating the parent, since the ID is in the data.
        serializer.save()

from django.http import FileResponse


class HanTrainingContentViewSet(viewsets.ModelViewSet):
    serializer_class = HanTrainingContentSerializer

    def get_queryset(self):
        """
        Filters materials based on a query parameter from the URL.
        Example: GET /api/materials/?han_subtopic_id=12
        """
        queryset = HanTrainingContent.objects.all()
        subtopic_id = self.request.query_params.get('han_subtopic_id')
        if subtopic_id:
            queryset = queryset.filter(han_subtopic_id=subtopic_id)
        return queryset

    def perform_create(self, serializer):
        """
        When creating, the parent ID must be sent in the request body.
        Example: POST /api/materials/ with body {"description": "...", "han_subtopic": 12}
        """
        serializer.save()


def serve_han_material_file(request, pk):
    """
    Serves the protected media file for a HanTrainingContent object.
    """
    material = get_object_or_404(HanTrainingContent, pk=pk)

    if not material.training_file:
        raise Http404("No file found for this material.")

    try:
        return FileResponse(material.training_file.open('rb'), as_attachment=False)
    except FileNotFoundError:
        raise Http404("File does not exist on the server.")


from rest_framework import viewsets
from django.shortcuts import get_object_or_404
from django.http import FileResponse
from .models import ShoContent, ShoSubtopic, ShoTrainingContent
from .serializers import (
    ShoContentListSerializer,
    ShoContentDetailSerializer,
    ShoSubtopicSerializer,
    ShoTrainingContentSerializer
)


# --- SHO CONTENT VIEWSET ---
class ShoContentViewSet(viewsets.ModelViewSet):
    """
    Handles listing, creating, retrieving, updating, and deleting Main Topics.
    """
    queryset = ShoContent.objects.prefetch_related('sho_subtopics__sho_materials').all()

    def get_serializer_class(self):
        """
        Chooses the serializer based on the action.
        - For retrieving a single item ('retrieve'), use the detailed serializer.
        - For listing all items ('list'), use the simple serializer.
        """
        if self.action == 'retrieve':
            return ShoContentDetailSerializer
        return ShoContentListSerializer


# --- SHO SUBTOPIC VIEWSET ---
class ShoSubtopicViewSet(viewsets.ModelViewSet):
    serializer_class = ShoSubtopicSerializer

    def get_queryset(self):
        """
        Filters subtopics based on a query parameter from the URL.
        Example: GET /api/sho-subtopics/?sho_content_id=5
        """
        queryset = ShoSubtopic.objects.all()
        sho_content_id = self.request.query_params.get('sho_content_id')
        if sho_content_id:
            queryset = queryset.filter(sho_content_id=sho_content_id)
        return queryset

    def perform_create(self, serializer):
        """
        When creating, the parent ID must be sent in the request body.
        Example: POST /api/sho-subtopics/ with body {"title": "...", "sho_content": 5}
        """
        serializer.save()


# --- SHO TRAINING CONTENT VIEWSET ---
class ShoTrainingContentViewSet(viewsets.ModelViewSet):
    serializer_class = ShoTrainingContentSerializer

    def get_queryset(self):
        """
        Filters materials based on a query parameter from the URL.
        Example: GET /api/sho-materials/?sho_subtopic_id=12
        """
        queryset = ShoTrainingContent.objects.all()
        subtopic_id = self.request.query_params.get('sho_subtopic_id')
        if subtopic_id:
            queryset = queryset.filter(sho_subtopic_id=subtopic_id)
        return queryset

    def perform_create(self, serializer):
        """
        When creating, the parent ID must be sent in the request body.
        Example: POST /api/sho-materials/ with body {"sho_description": "...", "sho_subtopic": 12}
        """
        serializer.save()


def serve_sho_material_file(request, pk):
    """
    Serves the protected media file for a ShoTrainingContent object.
    """
    material = get_object_or_404(ShoTrainingContent, pk=pk)

    if not material.training_file:
        raise Http404("No file found for this material.")

    try:
        return FileResponse(material.training_file.open('rb'), as_attachment=False)
    except FileNotFoundError:
        raise Http404("File does not exist on the server.")







from rest_framework import viewsets
from .models import HanchouExamQuestion
from .serializers import HanchouExamQuestionSerializer

class HanchouExamQuestionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing Hanchou Exam Questions, including bulk upload.
    """
    queryset = HanchouExamQuestion.objects.all()
    serializer_class = HanchouExamQuestionSerializer
    parser_classes = [MultiPartParser, FormParser]

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request, *args, **kwargs):
        """
        Generates and serves an Excel template with sample data for Hanchou questions.
        """
        sample_data = [
            {
                'question': 'What is the largest mammal in the world?',
                'option_a': 'Elephant',
                'option_b': 'Blue Whale',
                'option_c': 'Giraffe',
                'option_d': 'Great White Shark',
                'correct_answer': 'Blue Whale'
            },
            {
                'question': 'Which element has the atomic number 1?',
                'option_a': 'Helium',
                'option_b': 'Oxygen',
                'option_c': 'Hydrogen',
                'option_d': 'Carbon',
                'correct_answer': 'Hydrogen'
            }
        ]
        
        df = pd.DataFrame(sample_data)
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Using a specific sheet name is good practice
            df.to_excel(writer, index=False, sheet_name='Hanchou_Questions')
            
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Hanchou_Question_Upload_Template.xlsx"'
        
        return response

    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request, *args, **kwargs):
        """
        Handles bulk creation of Hanchou questions from an Excel file.
        """
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'detail': 'No file was uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # The sheet_name must match the one used in download_template
            df = pd.read_excel(file_obj, sheet_name='Hanchou_Questions', engine='openpyxl')
            df = df.where(pd.notnull(df), None)
        except Exception as e:
            return Response({'detail': f"Error reading the Excel file. Ensure it contains a sheet named 'Hanchou_Questions'. Error: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        required_columns = {
            'question', 'option_a', 'option_b', 'option_c',
            'option_d', 'correct_answer'
        }
        
        if not required_columns.issubset(df.columns):
            missing_cols = required_columns - set(df.columns)
            return Response({'detail': f'File is missing required columns: {", ".join(missing_cols)}'}, status=status.HTTP_400_BAD_REQUEST)

        questions_to_create = []
        errors = []

        for index, row in df.iterrows():
            row_data = row.to_dict()
            serializer = self.get_serializer(data=row_data)
            
            try:
                # Validate using the model's clean() method
                instance = HanchouExamQuestion(**row_data)
                instance.clean()
                
                # Also run standard serializer validation
                serializer.is_valid(raise_exception=True)
                questions_to_create.append(serializer.validated_data)

            except (DRFValidationError, DjangoValidationError, TypeError) as e:
                error_detail = serializer.errors if hasattr(serializer, 'errors') and serializer.errors else str(e)
                errors.append({'row': index + 2, 'errors': error_detail})
        
        if questions_to_create:
            model_instances = [HanchouExamQuestion(**data) for data in questions_to_create]
            HanchouExamQuestion.objects.bulk_create(model_instances)

        response_data = {
            'status': 'Upload complete.',
            'created_count': len(questions_to_create),
            'error_count': len(errors),
            'errors': errors
        }

        return Response(response_data, status=status.HTTP_201_CREATED)
    

from rest_framework import viewsets, permissions, filters as drf_filters
from django_filters import rest_framework as filters
from .models import HanchouExamResult
from .serializers import HanchouExamResultSerializer

class HanchouExamResultFilter(filters.FilterSet):
    pay_code = filters.CharFilter(field_name="employee__pay_code", lookup_expr="iexact")
    name = filters.CharFilter(field_name="employee__name", lookup_expr="icontains")
    exam_date = filters.DateFromToRangeFilter(field_name="exam_date")
    submitted_at = filters.DateFromToRangeFilter(field_name="submitted_at")
    passed = filters.BooleanFilter()

    class Meta:
        model = HanchouExamResult
        fields = ["employee", "passed", "exam_date"]

class HanchouExamResultViewSet(viewsets.ModelViewSet):
    queryset = HanchouExamResult.objects.select_related("employee").all().order_by("-submitted_at", "-started_at")
    serializer_class = HanchouExamResultSerializer
    # permission_classes = [permissions.IsAuthenticated]  # adjust as needed

    filter_backends = [filters.DjangoFilterBackend, drf_filters.SearchFilter, drf_filters.OrderingFilter]
    filterset_class = HanchouExamResultFilter
    search_fields = ["employee__name", "employee__pay_code", "employee__card_no", "remarks"]
    ordering_fields = ["submitted_at", "exam_date", "score", "total_questions", "duration_seconds"]
    ordering = ["-submitted_at", "-started_at"]

from rest_framework import viewsets
from .models import ShokuchouExamQuestion,ShokuchouExamResult
from .serializers import ShokuchouExamQuestionSerializer,ShokuchouExamResultSerializer

# Add these imports at the top of your views.py
import pandas as pd
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django.core.exceptions import ValidationError as DjangoValidationError
from rest_framework.exceptions import ValidationError as DRFValidationError

# Your existing model and serializer
from .models import ShokuchouExamQuestion
from .serializers import ShokuchouExamQuestionSerializer


class ShokuchouExamQuestionViewSet(viewsets.ModelViewSet):
    queryset = ShokuchouExamQuestion.objects.all()
    serializer_class = ShokuchouExamQuestionSerializer
    # Add parser classes for the ViewSet to handle file uploads
    parser_classes = [MultiPartParser, FormParser]



    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request, *args, **kwargs):
        """
        Generates and serves an Excel template with sample data for bulk uploading questions.
        """
        # Define the sample data
        sample_data = [
            {
                'sho_question': 'What is the capital of Japan?',
                'sho_option_a': 'Seoul',
                'sho_option_b': 'Beijing',
                'sho_option_c': 'Tokyo',
                'sho_option_d': 'Bangkok',
                'sho_correct_answer': 'Tokyo'
            },
            {
                'sho_question': 'Which planet is known as the Red Planet?',
                'sho_option_a': 'Earth',
                'sho_option_b': 'Mars',
                'sho_option_c': 'Jupiter',
                'sho_option_d': 'Saturn',
                'sho_correct_answer': 'Mars'
            }
        ]
        
        # Create a pandas DataFrame
        df = pd.DataFrame(sample_data)
        
        # Use an in-memory buffer
        buffer = io.BytesIO()
        
        # Write the DataFrame to the buffer as an Excel file
        # index=False prevents pandas from writing row indices to the file
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Questions')
            
        # Set the buffer's pointer to the beginning
        buffer.seek(0)
        
        # Create the HTTP response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Question_Upload_Template.xlsx"'
        
        return response







    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request, *args, **kwargs):
        """
        Handles bulk creation of questions from an Excel file upload.
        The Excel file must have a sheet named 'Questions' and columns:
        'sho_question', 'sho_option_a', 'sho_option_b', 'sho_option_c',
        'sho_option_d', 'sho_correct_answer'.
        """
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'detail': 'No file was uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not file_obj.name.endswith(('.xlsx', '.xls')):
            return Response({'detail': 'Invalid file format. Please upload an Excel file (.xlsx, .xls).'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Explicitly read the 'Questions' sheet
            df = pd.read_excel(file_obj, sheet_name='Questions', engine='openpyxl')
            # Replace NaN values with None for proper serialization
            df = df.where(pd.notnull(df), None)
        except Exception as e:
            # Catches errors like missing sheet or unreadable file
            return Response(
                {'detail': f"Error reading the Excel file. Make sure it contains a sheet named 'Questions'. Error: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )

        required_columns = {
            'sho_question', 'sho_option_a', 'sho_option_b', 'sho_option_c',
            'sho_option_d', 'sho_correct_answer'
        }
        
        if not required_columns.issubset(df.columns):
            missing_cols = required_columns - set(df.columns)
            return Response(
                {'detail': f'File is missing required columns: {", ".join(missing_cols)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        questions_to_create = []
        errors = []

        for index, row in df.iterrows():
            row_data = row.to_dict()
            serializer = self.get_serializer(data=row_data)
            
            try:
                # The model's clean() method is not called by default in DRF serializers.
                # We can simulate it by creating a model instance and calling full_clean().
                # This ensures our custom validation rule is checked.
                instance = ShokuchouExamQuestion(**row_data)
                instance.clean() # This will call our custom validation
                
                # We also run serializer validation for data types, max_length, etc.
                serializer.is_valid(raise_exception=True)
                questions_to_create.append(serializer.validated_data)

            except (DRFValidationError, DjangoValidationError, TypeError) as e:
                error_detail = serializer.errors if hasattr(serializer, 'errors') and serializer.errors else str(e)
                errors.append({'row': index + 2, 'errors': error_detail}) # +2 because index is 0-based and header is row 1
        
        # Now, create the actual model instances for bulk_create
        if questions_to_create:
            model_instances = [ShokuchouExamQuestion(**data) for data in questions_to_create]
            ShokuchouExamQuestion.objects.bulk_create(model_instances)

        response_data = {
            'status': 'Upload complete.',
            'created_count': len(questions_to_create),
            'error_count': len(errors),
            'errors': errors
        }

        return Response(response_data, status=status.HTTP_201_CREATED)






class ShokuchouExamResultFilter(filters.FilterSet):
    pay_code = filters.CharFilter(field_name="employee__pay_code", lookup_expr="iexact")
    name = filters.CharFilter(field_name="employee__name", lookup_expr="icontains")
    sho_submitted_at = filters.DateFromToRangeFilter(field_name="sho_submitted_at")
    sho_passed = filters.BooleanFilter()

    class Meta:
        model = ShokuchouExamResult
        fields = ["employee", "sho_passed", "sho_submitted_at"]


# --- SHO RESULT VIEWSET ---
class ShokuchouExamResultViewSet(viewsets.ModelViewSet):
    queryset = (
        ShokuchouExamResult.objects.select_related("employee")
        .all()
        .order_by("-sho_submitted_at", "-sho_started_at")
    )
    serializer_class = ShokuchouExamResultSerializer
    # permission_classes = [permissions.IsAuthenticated]  # enable if needed

    filter_backends = [filters.DjangoFilterBackend, drf_filters.SearchFilter, drf_filters.OrderingFilter]
    filterset_class = ShokuchouExamResultFilter
    search_fields = ["employee__name", "employee__pay_code", "employee__card_no", "sho_remarks"]
    ordering_fields = [
        "sho_submitted_at",
        "sho_score",
        "sho_total_questions",
        "sho_duration_seconds",
    ]
    ordering = ["-sho_submitted_at", "-sho_started_at"]



# your_app/views.py
import io
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.lib.colors import blue
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Import your models
from .models import OperatorSkill, HanchouExamResult



# your_app/views.py

import io
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from .models import OperatorSkill, HanchouExamResult # and other models...

# ★ 1. DEFINE A HELPER FUNCTION FOR THE REFLECTION EFFECT
def draw_reflected_text(p, x, y, text, font_name, font_size, main_color, shadow_color, y_offset=2):
    """
    Draws text with a soft shadow/reflection beneath it.
    """
    # Draw the shadow/reflection first in a light grey
    p.setFont(font_name, font_size)
    p.setFillColorRGB(*shadow_color)
    p.drawCentredString(x, y - y_offset, text)
    
    # Draw the main text on top in the primary color
    p.setFillColorRGB(*main_color)
    p.drawCentredString(x, y, text)


class OperatorSkillCertificatePDF(APIView):
    def get(self, request, pk):
        try:
            skill = OperatorSkill.objects.select_related('operator', 'station').get(pk=pk)
            
            employee_name = skill.operator.name.strip()
            skill_name = str(skill.station).strip()
            skill_level_value = skill.skill_level.strip()

            buffer = io.BytesIO()
            p = canvas.Canvas(buffer, pagesize=landscape(letter))
            width, height = landscape(letter)

            # ★ 2. DEFINE COLORS AND MARGINS
            main_color = (85/255, 26/255, 139/255)  # Deep Purple
            shadow_color = (200/255, 200/255, 200/255) # Light Grey
            border_color = (244/255, 145/255, 34/255) # Orange
            margin = 0.5 * inch

            # ★ 3. DRAW THE DECORATIVE BORDERS
            # Outer Border
            p.setStrokeColorRGB(*border_color)
            p.setLineWidth(3)
            p.rect(margin, margin, width - 2 * margin, height - 2 * margin)
            
            # Inner Border (subtle white/grey)
            p.setStrokeColorRGB(0.9, 0.9, 0.9)
            p.setLineWidth(1)
            p.rect(margin + 5, margin + 5, width - 2 * (margin + 5), height - 2 * (margin + 5))

            # --- DRAW CERTIFICATE CONTENT USING THE HELPER FUNCTION ---
            
            draw_reflected_text(p, width / 2.0, height - 1.7*inch, "KML SEATING", "Times-Bold", 36, main_color, shadow_color)
            draw_reflected_text(p, width / 2.0, height - 2.4*inch, "DOJO TRAINING CERTIFICATE", "Times-Bold", 28, main_color, shadow_color)
            
            # Body Text
            p.setFont("Times-Roman", 16)
            p.setFillColorRGB(0.1, 0.1, 0.1) # Dark grey for body
            p.drawCentredString(width / 2.0, height - 3.7*inch, "THIS IS TO CERTIFY THAT")
            
            # Employee Name
            draw_reflected_text(p, width / 2.0, height - 4.4*inch, f'“{employee_name.upper()}”', "Times-Bold", 24, main_color, shadow_color, y_offset=1.5)
            
            p.setFont("Times-Roman", 16)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawCentredString(width / 2.0, height - 5.2*inch, "HAS SUCCESSFULLY COMPLETED")
            p.drawCentredString(width / 2.0, height - 5.7*inch, "HIS TRAINING IN")

            # Skill Name
            draw_reflected_text(p, width / 2.0, height - 6.4*inch, skill_name.upper(), "Times-Bold", 22, main_color, shadow_color, y_offset=1.5)

            # ★ NEW: Add the skill level to the certificate
            p.setFont("Times-Roman", 16)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawCentredString(width / 2.0, height - 7.0*inch, "AT THE SKILL LEVEL OF")
            draw_reflected_text(p, width / 2.0, height - 7.5*inch, skill_level_value.upper(), "Times-Bold", 22, main_color, shadow_color, y_offset=1.5)

            # Signature
            p.setFont("Times-Roman", 12)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawRightString(width - margin - 0.5*inch, margin + 0.8*inch, "TRAINER SIGNATURE")
            p.line(width - margin - 2.5*inch, margin + 0.7*inch, width - margin - 0.5*inch, margin + 0.7*inch)

            p.showPage()
            p.save()

            buffer.seek(0)
            return HttpResponse(buffer, content_type='application/pdf', headers={'Content-Disposition': f'attachment; filename="Dojo_Certificate_{employee_name}.pdf"'})

        except OperatorSkill.DoesNotExist:
            return Response({"error": "Skill not found"}, status=status.HTTP_404_NOT_FOUND)
        except AttributeError:
            return Response({"error": "Associated operator or station could not be found."}, status=status.HTTP_404_NOT_FOUND)


class HanchouResultCertificatePDF(APIView):
    def get(self, request, pk):
        try:
            result = HanchouExamResult.objects.select_related('employee').get(pk=pk)
            
            employee_name = result.employee.name.strip()
            exam_name = result.exam_name.strip()

            buffer = io.BytesIO()
            p = canvas.Canvas(buffer, pagesize=landscape(letter))
            width, height = landscape(letter)

            # Define colors and margins
            main_color = (85/255, 26/255, 139/255)
            shadow_color = (200/255, 200/255, 200/255)
            border_color = (244/255, 145/255, 34/255)
            margin = 0.5 * inch

            # Draw the decorative borders
            p.setStrokeColorRGB(*border_color)
            p.setLineWidth(3)
            p.rect(margin, margin, width - 2 * margin, height - 2 * margin)
            p.setStrokeColorRGB(0.9, 0.9, 0.9)
            p.setLineWidth(1)
            p.rect(margin + 5, margin + 5, width - 2 * (margin + 5), height - 2 * (margin + 5))

            # --- DRAW CERTIFICATE CONTENT ---
            
            draw_reflected_text(p, width / 2.0, height - 1.7*inch, "KML SEATING", "Times-Bold", 36, main_color, shadow_color)
            draw_reflected_text(p, width / 2.0, height - 2.4*inch, "HANCHOUE EXAM CERTIFICATE", "Times-Bold", 28, main_color, shadow_color)
            
            p.setFont("Times-Roman", 16)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawCentredString(width / 2.0, height - 3.7*inch, "THIS IS TO CERTIFY THAT")
            
            draw_reflected_text(p, width / 2.0, height - 4.4*inch, f'“{employee_name.upper()}”', "Times-Bold", 24, main_color, shadow_color, y_offset=1.5)
            
            p.setFont("Times-Roman", 16)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawCentredString(width / 2.0, height - 5.2*inch, "HAS SUCCESSFULLY PASSED THE")
            
            draw_reflected_text(p, width / 2.0, height - 5.9*inch, exam_name.upper(), "Times-Bold", 22, main_color, shadow_color, y_offset=1.5)

            # Signature
            p.setFont("Times-Roman", 12)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawRightString(width - margin - 0.5*inch, margin + 0.8*inch, "TRAINER SIGNATURE")
            p.line(width - margin - 2.5*inch, margin + 0.7*inch, width - margin - 0.5*inch, margin + 0.7*inch)

            p.showPage()
            p.save()

            buffer.seek(0)
            return HttpResponse(buffer, content_type='application/pdf', headers={'Content-Disposition': f'attachment; filename="Hanchou_Certificate_{employee_name}.pdf"'})

        except HanchouExamResult.DoesNotExist:
            return Response({"error": "Hanchou exam result not found"}, status=status.HTTP_404_NOT_FOUND)
        except AttributeError:
            return Response({"error": "Associated employee for this result could not be found."}, status=status.HTTP_404_NOT_FOUND)



# Make sure you have all necessary imports at the top of your views.py file
import io
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch

from .models import Score # Ensure Score model is imported

# This helper function should be defined or imported in your views.py
def draw_reflected_text(p, x, y, text, font_name, font_size, main_color, shadow_color, y_offset=2):
    """Draws centered text with a subtle shadow/reflection effect."""
    # Shadow
    p.setFont(font_name, font_size)
    p.setFillColorRGB(*shadow_color)
    p.drawCentredString(x + y_offset, y - y_offset, text)
    # Main Text
    p.setFillColorRGB(*main_color)
    p.drawCentredString(x, y, text)


class ScoreCertificatePDF(APIView):
    """
    Generates a PDF certificate for a given Score.
    The content of the certificate is dynamically changed based on the score's level.
    """
    def get(self, request, pk):
        try:
            # Fetch the Score object with related employee and skill for efficiency
            score = Score.objects.select_related('employee', 'skill').get(pk=pk)

            # --- 1. EXTRACT DATA FROM THE MODEL ---
            employee_name = score.employee.name.strip()
            skill_name = str(score.skill).strip() if score.skill else "General Assessment"
            level_name = score.level.strip() if score.level else "" 
            percentage = score.percentage
            date_completed = score.created_at.strftime("%B %d, %Y")

            # --- 2. SETUP PDF CANVAS AND STYLES ---
            buffer = io.BytesIO()
            p = canvas.Canvas(buffer, pagesize=landscape(letter))
            width, height = landscape(letter)

            main_color = (85/255, 26/255, 139/255)  # Deep Purple
            shadow_color = (200/255, 200/255, 200/255) # Light Grey
            border_color = (244/255, 145/255, 34/255) # Orange
            margin = 0.5 * inch

            # --- 3. DRAW DECORATIVE BORDERS ---
            p.setStrokeColorRGB(*border_color)
            p.setLineWidth(3)
            p.rect(margin, margin, width - 2 * margin, height - 2 * margin)
            p.setStrokeColorRGB(0.9, 0.9, 0.9)
            p.setLineWidth(1)
            p.rect(margin + 5, margin + 5, width - 2 * (margin + 5), height - 2 * (margin + 5))

            # --- 4. DRAW CONTENT COMMON TO ALL CERTIFICATES ---
            draw_reflected_text(p, width / 2.0, height - 1.7*inch, "KML SEATING", "Times-Bold", 36, main_color, shadow_color)
            p.setFont("Times-Roman", 16)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawCentredString(width / 2.0, height - 3.7*inch, "THIS IS TO CERTIFY THAT")
            draw_reflected_text(p, width / 2.0, height - 4.4*inch, f'“{employee_name.upper()}”', "Times-Bold", 24, main_color, shadow_color, y_offset=1.5)

            # --- 5. DRAW CONDITIONAL CONTENT BASED ON LEVEL ---
            
            # Using .lower() makes the comparison case-insensitive (handles "Level 1", "level 1", etc.)
            if level_name.lower() == 'level 1':
                # --- CONTENT FOR LEVEL 1: DOJO TRAINING CERTIFICATE ---
                draw_reflected_text(p, width / 2.0, height - 2.4*inch, "DOJO TRAINING CERTIFICATE", "Times-Bold", 28, main_color, shadow_color)
                
                p.setFont("Times-Roman", 16)
                p.setFillColorRGB(0.1, 0.1, 0.1)
                p.drawCentredString(width / 2.0, height - 5.2*inch, "HAS SUCCESSFULLY COMPLETED")
                p.drawCentredString(width / 2.0, height - 5.7*inch, "HIS TRAINING IN")

                draw_reflected_text(p, width / 2.0, height - 6.4*inch, "DOJO", "Times-Bold", 22, main_color, shadow_color, y_offset=1.5)
            
            else:
                # --- CONTENT FOR ALL OTHER LEVELS: CERTIFICATE OF ACHIEVEMENT ---
                draw_reflected_text(p, width / 2.0, height - 2.4*inch, "CERTIFICATE OF ACHIEVEMENT", "Times-Bold", 28, main_color, shadow_color)

                p.setFont("Times-Roman", 16)
                p.setFillColorRGB(0.1, 0.1, 0.1)
                p.drawCentredString(width / 2.0, height - 5.2*inch, "HAS DEMONSTRATED PROFICIENCY IN")

                draw_reflected_text(p, width / 2.0, height - 5.9*inch, skill_name.upper(), "Times-Bold", 22, main_color, shadow_color, y_offset=1.5)
                p.drawCentredString(width / 2.0, height - 6.5*inch, "ACHIEVING")
                draw_reflected_text(p, width / 2.0, height - 7.0*inch, level_name.upper() if level_name else "ADVANCED STANDING", "Times-Bold", 22, main_color, shadow_color, y_offset=1.5)
                
                p.setFont("Times-Roman", 14)
                p.drawCentredString(width / 2.0, height - 7.6*inch, f"COMPLETED ON {date_completed} WITH A PASSING SCORE")

            # --- 6. DRAW SIGNATURE AND FINALIZE PDF ---
            p.setFont("Times-Roman", 12)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawRightString(width - margin - 0.5*inch, margin + 0.8*inch, "ASSESSOR SIGNATURE")
            p.line(width - margin - 2.5*inch, margin + 0.7*inch, width - margin - 0.5*inch, margin + 0.7*inch)

            p.showPage()
            p.save()

            # --- 7. PREPARE AND RETURN HTTP RESPONSE ---
            buffer.seek(0)
            clean_skill_name = "".join(c if c.isalnum() else "_" for c in skill_name)
            filename = f"Cert_{employee_name.replace(' ', '_')}_{clean_skill_name}.pdf"
            
            return HttpResponse(
                buffer, 
                content_type='application/pdf', 
                headers={'Content-Disposition': f'attachment; filename="{filename}"'}
            )

        except Score.DoesNotExist:
            return Response({"error": "Score record not found."}, status=status.HTTP_404_NOT_FOUND)
        except AttributeError:
            return Response({"error": "Associated employee or skill data could not be found."}, status=status.HTTP_404_NOT_FOUND)




import io
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

# Make sure to import your models
from .models import ShokuchouExamResult, EmployeeMaster 

# --- Place the draw_reflected_text function here ---
def draw_reflected_text(p, x, y, text, font, size, main_color, shadow_color, x_offset=1.5, y_offset=1.5):
    """Draws text with a simple shadow/reflection underneath."""
    # Draw shadow/reflection
    p.setFont(font, size)
    p.setFillColorRGB(*shadow_color)
    p.drawCentredString(x + x_offset, y - y_offset, text)
    
    # Draw main text
    p.setFillColorRGB(*main_color)
    p.drawCentredString(x, y, text)
# ----------------------------------------------------


class ShokuchouResultCertificatePDF(APIView):
    def get(self, request, pk):
        try:
            # CHANGED: Query ShokuchouExamResult instead of HanchouExamResult
            result = ShokuchouExamResult.objects.select_related('employee').get(pk=pk)
            
            employee_name = result.employee.name.strip()
            # CHANGED: Use sho_exam_name field
            exam_name = result.sho_exam_name.strip()

            buffer = io.BytesIO()
            p = canvas.Canvas(buffer, pagesize=landscape(letter))
            width, height = landscape(letter)

            # Define colors and margins (kept the same for consistent branding)
            main_color = (85/255, 26/255, 139/255)
            shadow_color = (200/255, 200/255, 200/255)
            border_color = (244/255, 145/255, 34/255)
            margin = 0.5 * inch

            # Draw the decorative borders
            p.setStrokeColorRGB(*border_color)
            p.setLineWidth(3)
            p.rect(margin, margin, width - 2 * margin, height - 2 * margin)
            p.setStrokeColorRGB(0.9, 0.9, 0.9)
            p.setLineWidth(1)
            p.rect(margin + 5, margin + 5, width - 2 * (margin + 5), height - 2 * (margin + 5))

            # --- DRAW CERTIFICATE CONTENT ---
            
            draw_reflected_text(p, width / 2.0, height - 1.7*inch, "KML SEATING", "Times-Bold", 36, main_color, shadow_color)
            # CHANGED: Updated certificate title
            draw_reflected_text(p, width / 2.0, height - 2.4*inch, "SHOKUCHOU EXAM CERTIFICATE", "Times-Bold", 28, main_color, shadow_color)
            
            p.setFont("Times-Roman", 16)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawCentredString(width / 2.0, height - 3.7*inch, "THIS IS TO CERTIFY THAT")
            
            draw_reflected_text(p, width / 2.0, height - 4.4*inch, f'“{employee_name.upper()}”', "Times-Bold", 24, main_color, shadow_color, y_offset=1.5)
            
            p.setFont("Times-Roman", 16)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawCentredString(width / 2.0, height - 5.2*inch, "HAS SUCCESSFULLY PASSED THE")
            
            draw_reflected_text(p, width / 2.0, height - 5.9*inch, exam_name.upper(), "Times-Bold", 22, main_color, shadow_color, y_offset=1.5)

            # Signature
            p.setFont("Times-Roman", 12)
            p.setFillColorRGB(0.1, 0.1, 0.1)
            p.drawRightString(width - margin - 0.5*inch, margin + 0.8*inch, "TRAINER SIGNATURE")
            p.line(width - margin - 2.5*inch, margin + 0.7*inch, width - margin - 0.5*inch, margin + 0.7*inch)

            p.showPage()
            p.save()

            buffer.seek(0)
            # CHANGED: Updated the filename for the download
            return HttpResponse(buffer, content_type='application/pdf', headers={'Content-Disposition': f'attachment; filename="Shokuchou_Certificate_{employee_name}.pdf"'})

        # CHANGED: Catch DoesNotExist for the correct model
        except ShokuchouExamResult.DoesNotExist:
            return Response({"error": "Shokuchou exam result not found"}, status=status.HTTP_404_NOT_FOUND)
        except AttributeError:
            return Response({"error": "Associated employee for this result could not be found."}, status=status.HTTP_404_NOT_FOUND)




# ------------ojlistoutview----------------#
# views.py
from itertools import chain
from rest_framework import viewsets
from rest_framework.response import Response

from .models import LevelTwoTraineeInfo, LevelTwoQATraineeInfo
from .serializers import UnifiedOJTStatusSerializer


class OJTStatusViewSet(viewsets.ViewSet):
    """
    Read-only API for OJT Status (Production + QA trainees).
    """
    def list(self, request):
        # ✅ Get both trainee sets
        production_trainees = LevelTwoTraineeInfo.objects.all()
        qa_trainees = LevelTwoQATraineeInfo.objects.all()

        # ✅ Combine into one list
        all_trainees = list(chain(production_trainees, qa_trainees))

        # ✅ Serialize them together
        serializer = UnifiedOJTStatusSerializer(all_trainees, many=True)

        return Response(serializer.data)


# views.py
from rest_framework import viewsets
from rest_framework.response import Response
from .models import LevelThreeTraineeInfo, LevelThreeQATraineeInfo
from .serializers import UnifiedLevelThreeOJTStatusSerializer


class LevelThreeOJTStatusViewSet(viewsets.ViewSet):
    """
    Unified OJT Status for Level 3 (Production + QA)
    """

    def list(self, request):
        # Fetch all trainees (Production + QA)
        production_trainees = LevelThreeTraineeInfo.objects.all()
        qa_trainees = LevelThreeQATraineeInfo.objects.all()

        # Merge into single list
        all_trainees = list(production_trainees) + list(qa_trainees)

        # Serialize
        serializer = UnifiedLevelThreeOJTStatusSerializer(all_trainees, many=True)
        return Response(serializer.data)





#-----------------------------------------
from django.http import HttpResponse, JsonResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import json
import traceback
from collections import defaultdict
from .models import (
    EmployeeMaster, OperatorSkill, Station, SubLine, MainLine, 
    MainDepartment, Score, MultiSkilling
)

@method_decorator(csrf_exempt, name='dispatch')
class SkillMatrixExcelView(View):
    def post(self, request, *args, **kwargs):
        """
        Generate Skill Matrix Excel Report
        Accepts filters for department, main_line, sub_line
        """
        try:
            print("\n=== Received Skill Matrix Excel generation request ===")
            
            # 1. Parse input data
            filters = self._get_filters(request)
            print(f"Processing filters: {filters}")

            # 2. Get filtered data
            skill_matrix_data = self._get_skill_matrix_data(filters)
            if not skill_matrix_data['operators']:
                return JsonResponse({'error': 'No operators found for the given criteria'}, status=404)

            # 3. Generate Excel content
            print("Generating Skill Matrix Excel content...")
            buffer = BytesIO()
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Skill Matrix Report"
            
            self.create_skill_matrix_content(ws, skill_matrix_data, filters)
            
            # 4. Save to buffer
            print("Building Skill Matrix Excel document...")
            wb.save(buffer)
            buffer.seek(0)
            print("Skill Matrix Excel generation completed successfully")

            # 5. Return Excel response
            response = HttpResponse(
                buffer.getvalue(), 
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Create filename based on filters
            filename = self._generate_filename(filters)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            print("\n!!! Skill Matrix Excel generation failed !!!")
            traceback.print_exc()
            return JsonResponse(
                {
                    'error': 'Internal server error',
                    'detail': str(e),
                    'traceback': traceback.format_exc()
                }, 
                status=500
            )

    def _get_filters(self, request):
        """Helper method to extract filters from request"""
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                return {
                    'department_id': data.get('department_id'),
                    'main_line_id': data.get('main_line_id'),
                    'sub_line_id': data.get('sub_line_id'),
                }
            except json.JSONDecodeError:
                return {}
        return {
            'department_id': request.POST.get('department_id'),
            'main_line_id': request.POST.get('main_line_id'),
            'sub_line_id': request.POST.get('sub_line_id'),
        }

    def _get_skill_matrix_data(self, filters):
        """Get organized skill matrix data based on filters"""
        
        # Build query filters
        station_filters = {}
        if filters.get('sub_line_id'):
            station_filters['sub_line_id'] = filters['sub_line_id']
        elif filters.get('main_line_id'):
            station_filters['sub_line__main_line_id'] = filters['main_line_id']
        elif filters.get('department_id'):
            station_filters['sub_line__main_line__department_id'] = filters['department_id']

        # Get stations based on filters
        stations = Station.objects.filter(**station_filters).select_related(
            'sub_line__main_line__department'
        ).order_by('station_number')

        if not stations.exists():
            return {'operators': [], 'stations': [], 'matrix': {}}

        # Get operators who have skills in these stations
        operator_skills = OperatorSkill.objects.filter(
            station__in=stations
        ).select_related('operator', 'station').order_by('operator__name')

        # Build matrix data structure
        operators = list(set([skill.operator for skill in operator_skills]))
        operators.sort(key=lambda x: x.name)

        # Create skill matrix
        matrix = defaultdict(dict)
        for skill in operator_skills:
            matrix[skill.operator.card_no][skill.station.station_number] = {
                'skill_level': skill.skill_level,
                'sequence': skill.sequence or 0
            }

        return {
            'operators': operators,
            'stations': stations,
            'matrix': dict(matrix),
            'hierarchy_info': self._get_hierarchy_info(filters)
        }

    def _get_hierarchy_info(self, filters):
        """Get hierarchy information for report header"""
        info = {}
        
        if filters.get('department_id'):
            try:
                dept = MainDepartment.objects.get(id=filters['department_id'])
                info['department'] = dept.name
            except MainDepartment.DoesNotExist:
                pass
                
        if filters.get('main_line_id'):
            try:
                line = MainLine.objects.get(id=filters['main_line_id'])
                info['main_line'] = line.name
                if not info.get('department'):
                    info['department'] = line.department.name
            except MainLine.DoesNotExist:
                pass
                
        if filters.get('sub_line_id'):
            try:
                sub_line = SubLine.objects.get(id=filters['sub_line_id'])
                info['sub_line'] = sub_line.name
                if not info.get('main_line'):
                    info['main_line'] = sub_line.main_line.name
                if not info.get('department'):
                    info['department'] = sub_line.main_line.department.name
            except SubLine.DoesNotExist:
                pass
                
        return info

    def _generate_filename(self, filters):
        """Generate appropriate filename based on filters"""
        parts = ["skill_matrix"]
        
        if filters.get('sub_line_id'):
            parts.append("sub_line")
        elif filters.get('main_line_id'):
            parts.append("main_line")
        elif filters.get('department_id'):
            parts.append("department")
        else:
            parts.append("all")
            
        return f"{'_'.join(parts)}_report.xlsx"

    def _is_qualified(self, skill_level, min_required):
        """Determine if a skill level is considered qualified based on minimum required"""
        skill_levels = {'Level 1': 1, 'Level 2': 2, 'Level 3': 3, 'Level 4': 4}
        current_level = skill_levels.get(skill_level, 0)
        min_level = skill_levels.get(min_required, 0)
        return current_level >= min_level

    def create_skill_matrix_content(self, ws, data, filters):
        """Generate the Excel content structure for skill matrix"""
        current_row = 1
        
        # Title with hierarchy information
        title_parts = ["Skill Matrix Report"]
        hierarchy = data['hierarchy_info']
        
        if hierarchy.get('department'):
            title_parts.append(f"Department: {hierarchy['department']}")
        if hierarchy.get('main_line'):
            title_parts.append(f"Line: {hierarchy['main_line']}")
        if hierarchy.get('sub_line'):
            title_parts.append(f"Sub-Line: {hierarchy['sub_line']}")
            
        # Add title
        ws.merge_cells(f'A{current_row}:E{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = " - ".join(title_parts)
        title_cell.font = Font(size=16, bold=True, color='2E4D6B')
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 2
        
        # Summary section
        current_row = self._add_summary_section(ws, data, current_row)
        current_row += 2
        
        # Main skill matrix table
        current_row = self._add_skill_matrix_table(ws, data, current_row)
        current_row += 2
        
        # Legend
        self._add_legend(ws, current_row)
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws)

    def _add_summary_section(self, ws, data, start_row):
        """Add summary statistics section"""
        # Section title
        ws[f'A{start_row}'].value = "Summary Statistics"
        ws[f'A{start_row}'].font = Font(size=12, bold=True)
        start_row += 1
        
        total_operators = len(data['operators'])
        total_stations = len(data['stations'])
        
        # Calculate qualification statistics
        qualified_counts = defaultdict(int)
        for operator in data['operators']:
            operator_skills = data['matrix'].get(operator.card_no, {})
            for station_num, skill_info in operator_skills.items():
                level = skill_info['skill_level']
                if self._is_qualified(level, 'Level 1'):  # Assuming Level 1 as baseline for counting
                    qualified_counts[level] += 1

        # Summary data
        summary_data = [
            ["Metric", "Count"],
            ["Total Operators", total_operators],
            ["Total Stations", total_stations],
            ["Level 1 Qualifications", qualified_counts.get('Level 1', 0)],
            ["Level 2 Qualifications", qualified_counts.get('Level 2', 0)],
            ["Level 3 Qualifications", qualified_counts.get('Level 3', 0)],
            ["Level 4 Qualifications", qualified_counts.get('Level 4', 0)],
        ]
        
        # Add summary table
        for i, row_data in enumerate(summary_data):
            for j, value in enumerate(row_data):
                cell = ws.cell(row=start_row + i, column=j + 1, value=value)
                
                # Style header row
                if i == 0:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='4682B4', end_color='4682B4', fill_type='solid')
                else:
                    cell.font = Font(size=10)
                
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        return start_row + len(summary_data)

    def _add_skill_matrix_table(self, ws, data, start_row):
        """Add the main skill matrix table"""
        # Section title
        ws[f'A{start_row}'].value = "Operator Skill Matrix"
        ws[f'A{start_row}'].font = Font(size=12, bold=True)
        start_row += 1
        
        operators = data['operators']
        stations = data['stations']
        matrix = data['matrix']
        
        # Build table headers
        headers = ["S.No", "Operator Name", "Card No", "DOJ"]
        
        # Add station headers dynamically based on filtered stations
        for station in stations:
            station_text = f"Station {station.station_number}"
            if hasattr(station, 'skill') and station.skill:
                station_text += f"\n{station.skill}"
            headers.append(station_text)
        
        # Add headers to worksheet
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4682B4', end_color='4682B4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Add minimum skill required and operator required row
        min_skill_row = start_row + 1
        ws.cell(row=min_skill_row, column=1, value="Minimum Skill Required")
        ws.cell(row=min_skill_row, column=2, value="")
        ws.cell(row=min_skill_row, column=3, value="")
        ws.cell(row=min_skill_row, column=4, value="")
        for col, station in enumerate(stations, 5):
            cell = ws.cell(row=min_skill_row, column=col, value=station.minimum_skill_required)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        op_required_row = start_row + 2
        ws.cell(row=op_required_row, column=1, value="Station Wise Minimum Operator Required")
        ws.cell(row=op_required_row, column=2, value="")
        ws.cell(row=op_required_row, column=3, value="")
        ws.cell(row=op_required_row, column=4, value="")
        for col, station in enumerate(stations, 5):
            cell = ws.cell(row=op_required_row, column=col, value=station.min_operator_required)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Add data rows
        for idx, operator in enumerate(operators, 1):
            row_num = start_row + 2 + idx
            
            # Basic operator info
            row_data = [
                idx,
                operator.name,
                operator.card_no,
                operator.joining_date.strftime('%d/%m/%Y') if operator.joining_date else "N/A"
            ]
            
            # Add skill level and qualification status for each station
            operator_skills = matrix.get(operator.card_no, {})
            for station in stations:
                skill_info = operator_skills.get(station.station_number)
                if skill_info:
                    level = skill_info['skill_level'].replace('Level ', 'L')
                    min_required = station.minimum_skill_required
                    qualified = "Qualified" if self._is_qualified(skill_info['skill_level'], min_required) else "Not Qualified"
                    row_data.append(f"{level} ({qualified})")
                else:
                    row_data.append("-")
            
            # Add row data to worksheet with styling
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Color code skill levels
                if col > 4:  # Station columns
                    if "L1" in value:
                        cell.font = Font(bold=True, color='FF0000')
                    elif "L2" in value:
                        cell.font = Font(bold=True, color='FFA500')
                    elif "L3" in value:
                        cell.font = Font(bold=True, color='FFC107')
                    elif "L4" in value:
                        cell.font = Font(bold=True, color='008000')
                
                # Zebra striping
                if idx % 2 == 0:
                    cell.fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
        
        return row_num + 1

    def _add_legend(self, ws, start_row):
        """Add legend for skill levels and symbols"""
        # Section title
        ws[f'A{start_row}'].value = "Legend"
        ws[f'A{start_row}'].font = Font(size=12, bold=True)
        start_row += 1
        
        legend_data = [
            ["Symbol/Code", "Meaning"],
            ["L1 (Qualified)", "Level 1 - Basic (Qualified) "],
            ["L1 (Not Qualified)", "Level 1 - Below Minimum Required "],
            ["L2 (Qualified)", "Level 2 - Intermediate (Qualified) "],
            ["L2 (Not Qualified)", "Level 2 - Below Minimum Required "],
            ["L3 (Qualified)", "Level 3 - Advanced (Qualified) "],
            ["L3 (Not Qualified)", "Level 3 - Below Minimum Required "],
            ["L4 (Qualified)", "Level 4 - Expert (Qualified) "],
            ["L4 (Not Qualified)", "Level 4 - Below Minimum Required "],
            ["-", "No Skill/Not Assigned"],
        ]
        
        # Add legend table
        for i, row_data in enumerate(legend_data):
            for j, value in enumerate(row_data):
                cell = ws.cell(row=start_row + i, column=j + 1, value=value)
                
                # Style header row
                if i == 0:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='666666', end_color='666666', fill_type='solid')
                else:
                    cell.font = Font(size=10)
                    # Apply color to Symbol/Code column (j == 0)
                    if j == 0:
                        if "L1" in value:
                            cell.font = Font(size=10, color='FF0000', bold=True)
                        elif "L2" in value:
                            cell.font = Font(size=10, color='FFA500', bold=True)
                        elif "L3" in value:
                            cell.font = Font(size=10, color='FFC107', bold=True)
                        elif "L4" in value:
                            cell.font = Font(size=10, color='008000', bold=True)
                        elif "-" in value:
                            cell.font = Font(size=10, color='000000', bold=True)
                
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, 8), 20)
            ws.column_dimensions[column_letter].width = adjusted_width















#------------------------------------------------------------------------------------------------------
from django.http import HttpResponse, JsonResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
from io import BytesIO
import traceback
from .models import (
    LevelTwoTraineeInfo, LevelTwoTrainingTopic, LevelTwoOJTDay, 
    LevelTwoOJTScore, LevelTwoSubStation, Line
)

@method_decorator(csrf_exempt, name='dispatch')
class LevelTwoOJTReportPDFView(View):
    def post(self, request, *args, **kwargs):
        """
        Handle PDF generation requests for Level Two OJT production reports
        """
        try:
            print("\n=== Received OJT PDF generation request ===")
            
            # 1. Parse input data
            trainee_id = self._get_trainee_id(request)
            station_id = self._get_station_id(request)
            
            if not trainee_id:
                return JsonResponse({'error': 'trainee_id is required'}, status=400)
            if not station_id:
                return JsonResponse({'error': 'station_id is required'}, status=400)
            
            print(f"Processing trainee_id: {trainee_id}, station_id: {station_id}")

            # 2. Get trainee record
            trainee = self._get_trainee_record(trainee_id, station_id)
            if isinstance(trainee, JsonResponse):
                return trainee
            if not trainee:
                return JsonResponse({'error': 'Trainee not found for the specified trainee_id and station_id'}, status=404)

            print(f"Found trainee: {trainee.trainee_name}")

            # 3. Generate PDF content
            print("Generating OJT PDF content...")
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), 
                                  leftMargin=0.5*inch, rightMargin=0.5*inch,
                                  topMargin=0.5*inch, bottomMargin=0.5*inch)
            story = self.create_ojt_pdf_content(trainee)
            
            # 4. Build PDF document
            print("Building OJT PDF document...")
            doc.build(story)
            buffer.seek(0)
            print("OJT PDF generation completed successfully")

            # 5. Return PDF response
            response = HttpResponse(
                buffer.getvalue(), 
                content_type='application/pdf'
            )
            response['Content-Disposition'] = (
                f'attachment; filename="ojt_report_{trainee_id}_production.pdf"'
            )
            return response
            
        except Exception as e:
            print("\n!!! OJT PDF generation failed !!!")
            traceback.print_exc()
            return JsonResponse(
                {
                    'error': 'Internal server error',
                    'detail': str(e),
                    'traceback': traceback.format_exc()
                }, 
                status=500
            )

    def _get_trainee_id(self, request):
        """Helper method to extract trainee_id from request"""
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                return data.get('trainee_id')
            except json.JSONDecodeError:
                return None
        return request.POST.get('trainee_id')

    def _get_station_id(self, request):
        """Helper method to extract station_id from request"""
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                return data.get('station_id')
            except json.JSONDecodeError:
                return None
        return request.POST.get('station_id')

    def _get_trainee_record(self, trainee_id, station_id):
        """Get trainee record for production"""
        try:
            return LevelTwoTraineeInfo.objects.get(traineeId=trainee_id, station_id=station_id)
        except LevelTwoTraineeInfo.DoesNotExist:
            return None
        except LevelTwoTraineeInfo.MultipleObjectsReturned:
            try:
                return LevelTwoTraineeInfo.objects.filter(traineeId=trainee_id, station_id=station_id).order_by('-id').first()
            except Exception as e:
                return JsonResponse(
                    {'error': f'Multiple trainee records found for trainee_id {trainee_id} and station_id {station_id}, and fallback selection failed: {str(e)}'},
                    status=400
                )

    def create_ojt_pdf_content(self, trainee):
        """Generate the OJT PDF content structure for production"""
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = "Level-2 OJT MONITORING SHEET - PRODUCTION"
        story.append(Paragraph(title, styles['Title']))
        story.append(Spacer(1, 12))
        
        # Add all sections
        self._add_trainee_info(story, styles, trainee)
        self._add_ojt_monitoring_table(story, styles, trainee)
        self._add_judgement_criteria(story, styles)
        
        return story

    def _add_trainee_info(self, story, styles, trainee):
        """Add trainee basic information for production"""
        # Get station_name from the ForeignKey
        station_name = trainee.station.name if trainee.station else "Not Assigned"

        # Get line_name from the ForeignKey
        line_name = trainee.line.name if hasattr(trainee, 'line') and trainee.line else "Not Assigned"

        info_data = [
            ["Trainee Name:", trainee.trainee_name, "Trainee Id:", trainee.traineeId],
            ["Station Name:", station_name, "Line Name:", line_name],
            ["Trainer Name:", trainee.trainer_name, "Training Status:", trainee.training_status]
        ]
        
        info_table = Table(info_data, colWidths=[100, 200, 100, 200])
        info_style = TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('ALIGN', (0,0), (0,-1), 'LEFT'),
            ('ALIGN', (2,0), (2,-1), 'LEFT'),
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ])
        info_table.setStyle(info_style)
        story.append(info_table)
        story.append(Spacer(1, 16))

    def _add_ojt_monitoring_table(self, story, styles, trainee):
        """Add the main OJT monitoring table for production"""
        topics = LevelTwoTrainingTopic.objects.all().order_by('sl_no')
        days = LevelTwoOJTDay.objects.all().order_by('id')
        scores = LevelTwoOJTScore.objects.filter(trainee=trainee)

        header = ["S.NO.", "TRAINING TOPIC", "Date"]
        for day in days:
            header.append(day.name)
        
        table_data = [header]
        
        for topic in topics:
            row = [str(topic.sl_no), topic.topic, topic.date]
            for day in days:
                score_obj = scores.filter(topic=topic, day=day).first()
                score_display = str(score_obj.score) if score_obj else ""
                row.append(score_display)
            table_data.append(row)

        base_widths = [40, 300, 80]
        day_width = (720 - sum(base_widths)) / len(days) if days else 60
        col_widths = base_widths + [day_width] * len(days)

        monitoring_table = Table(table_data, colWidths=col_widths)
        monitoring_style = self._get_monitoring_table_style(len(topics), len(days))
        monitoring_table.setStyle(monitoring_style)
        
        story.append(monitoring_table)
        story.append(Spacer(1, 16))

    def _get_monitoring_table_style(self, topic_count, day_count):
        """Get styling for the monitoring table"""
        style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4682B4')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('TOPPADDING', (0,0), (-1,0), 8),
            ('ALIGN', (0,1), (0,-1), 'CENTER'),
            ('ALIGN', (1,1), (1,-1), 'LEFT'),
            ('ALIGN', (2,1), (2,-1), 'CENTER'),
            ('ALIGN', (3,1), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F8F8')]),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ])
        
        return style

    def _add_judgement_criteria(self, story, styles):
        """Add judgement criteria section"""
        story.append(Paragraph("Judgement Criteria", styles['Heading3']))
        story.append(Spacer(1, 8))
        
        criteria_data = [
            ["Criteria", "Description"],
            ["10 Marks", "OK - Trainee demonstrates competency"],
            ["0 Marks", "Not OK - Trainee needs improvement"]
        ]
        
        criteria_table = Table(criteria_data, colWidths=[100, 400])
        criteria_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4682B4')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('ALIGN', (0,1), (0,-1), 'CENTER'),
            ('ALIGN', (1,1), (1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ])
        criteria_table.setStyle(criteria_style)
        story.append(criteria_table)
        story.append(Spacer(1, 16))

        assessment_text = """
        <b>Day-6 Assessment should be 100%</b><br/>
        <b>Passing Criteria: 100%</b><br/>
        <i>If failed in evaluation re-training is required</i>
        """
        story.append(Paragraph(assessment_text, styles['Normal']))




from django.http import HttpResponse, JsonResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
from io import BytesIO
import json
import traceback
from .models import (
    LevelTwoQATraineeInfo, LevelTwoQATrainingTopic, LevelTwoQAOJTDay, 
    LevelTwoQAOJTScore, LevelTwoQualityLine
)

@method_decorator(csrf_exempt, name='dispatch')
class LevelTwoQAReportPDFView(View):
    def post(self, request, *args, **kwargs):
        """
        Handle PDF generation requests for Level Two QA reports
        """
        try:
            print("\n=== Received QA PDF generation request ===")
            
            # 1. Parse input data
            trainee_id = self._get_trainee_id(request)
            line_id = self._get_line_id(request)
            
            if not trainee_id:
                return JsonResponse({'error': 'trainee_id is required'}, status=400)
            if not line_id:
                return JsonResponse({'error': 'line_id is required'}, status=400)
            
            print(f"Processing trainee_id: {trainee_id}, line_id: {line_id}")

            # 2. Get trainee record
            trainee = self._get_trainee_record(trainee_id, line_id)
            if isinstance(trainee, JsonResponse):
                return trainee
            if not trainee:
                return JsonResponse({'error': 'Trainee not found for the specified trainee_id and line_id'}, status=404)

            print(f"Found trainee: {trainee.trainee_name}")

            # 3. Generate PDF content
            print("Generating QA PDF content...")
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), 
                                  leftMargin=0.5*inch, rightMargin=0.5*inch,
                                  topMargin=0.5*inch, bottomMargin=0.5*inch)
            story = self.create_qa_pdf_content(trainee)
            
            # 4. Build PDF document
            print("Building QA PDF document...")
            doc.build(story)
            buffer.seek(0)
            print("QA PDF generation completed successfully")

            # 5. Return PDF response
            response = HttpResponse(
                buffer.getvalue(), 
                content_type='application/pdf'
            )
            response['Content-Disposition'] = (
                f'attachment; filename="qa_report_{trainee_id}_quality.pdf"'
            )
            return response
            
        except Exception as e:
            print("\n!!! QA PDF generation failed !!!")
            traceback.print_exc()
            return JsonResponse(
                {
                    'error': 'Internal server error',
                    'detail': str(e),
                    'traceback': traceback.format_exc()
                }, 
                status=500
            )

    def _get_trainee_id(self, request):
        """Helper method to extract trainee_id from request"""
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                return data.get('trainee_id')
            except json.JSONDecodeError:
                return None
        return request.POST.get('trainee_id')

    def _get_line_id(self, request):
        """Helper method to extract line_id from request"""
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                return data.get('line_id')
            except json.JSONDecodeError:
                return None
        return request.POST.get('line_id')

    def _get_trainee_record(self, trainee_id, line_id):
        """Get trainee record for quality"""
        try:
            return LevelTwoQATraineeInfo.objects.get(traineeId=trainee_id, line_id=line_id)
        except LevelTwoQATraineeInfo.DoesNotExist:
            return None
        except LevelTwoQATraineeInfo.MultipleObjectsReturned:
            try:
                return LevelTwoQATraineeInfo.objects.filter(traineeId=trainee_id, line_id=line_id).order_by('-id').first()
            except Exception as e:
                return JsonResponse(
                    {'error': f'Multiple trainee records found for trainee_id {trainee_id} and line_id {line_id}, and fallback selection failed: {str(e)}'},
                    status=400
                )

    def create_qa_pdf_content(self, trainee):
        """Generate the QA PDF content structure for quality"""
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = "Level-2 QA MONITORING SHEET - QUALITY"
        story.append(Paragraph(title, styles['Title']))
        story.append(Spacer(1, 12))
        
        # Add all sections
        self._add_trainee_info(story, styles, trainee)
        self._add_qa_monitoring_table(story, styles, trainee)
        self._add_judgement_criteria(story, styles)
        
        return story

    def _add_trainee_info(self, story, styles, trainee):
        """Add trainee basic information for quality"""
        # Get line_name from the ForeignKey
        line_name = trainee.line.name if trainee.line else "Not Assigned"

        info_data = [
            ["Trainee Name:", trainee.trainee_name, "Trainee Id:", trainee.traineeId],
            ["Line Name:", line_name, "", ""],
            ["Trainer Name:", trainee.trainer_name, "Training Status:", trainee.training_status]
        ]
        
        info_table = Table(info_data, colWidths=[100, 200, 100, 200])
        info_style = TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('ALIGN', (0,0), (0,-1), 'LEFT'),
            ('ALIGN', (2,0), (2,-1), 'LEFT'),
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ])
        info_table.setStyle(info_style)
        story.append(info_table)
        story.append(Spacer(1, 16))

    def _add_qa_monitoring_table(self, story, styles, trainee):
        """Add the main QA monitoring table for quality"""
        topics = LevelTwoQATrainingTopic.objects.all().order_by('sl_no')
        days = LevelTwoQAOJTDay.objects.all().order_by('id')
        scores = LevelTwoQAOJTScore.objects.filter(trainee=trainee)

        header = ["S.NO.", "TRAINING TOPIC", "Date"]
        for day in days:
            header.append(day.name)
        
        table_data = [header]
        
        for topic in topics:
            row = [str(topic.sl_no), topic.topic, topic.date]
            for day in days:
                score_obj = scores.filter(topic=topic, day=day).first()
                score_display = str(score_obj.score) if score_obj else ""
                row.append(score_display)
            table_data.append(row)

        base_widths = [40, 300, 80]
        day_width = (720 - sum(base_widths)) / len(days) if days else 60
        col_widths = base_widths + [day_width] * len(days)

        monitoring_table = Table(table_data, colWidths=col_widths)
        monitoring_style = self._get_monitoring_table_style(len(topics), len(days))
        monitoring_table.setStyle(monitoring_style)
        
        story.append(monitoring_table)
        story.append(Spacer(1, 16))

    def _get_monitoring_table_style(self, topic_count, day_count):
        """Get styling for the monitoring table"""
        style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4682B4')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('TOPPADDING', (0,0), (-1,0), 8),
            ('ALIGN', (0,1), (0,-1), 'CENTER'),
            ('ALIGN', (1,1), (1,-1), 'LEFT'),
            ('ALIGN', (2,1), (2,-1), 'CENTER'),
            ('ALIGN', (3,1), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F8F8')]),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ])
        
        return style

    def _add_judgement_criteria(self, story, styles):
        """Add judgement criteria section"""
        story.append(Paragraph("Judgement Criteria", styles['Heading3']))
        story.append(Spacer(1, 8))
        
        criteria_data = [
            ["Criteria", "Description"],
            ["0-10 Marks", "Score based on trainee performance"],
            ["Pass", "Achieves 100% on final assessment"],
            ["Fail", "Below 100% on final assessment"]
        ]
        
        criteria_table = Table(criteria_data, colWidths=[100, 400])
        criteria_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4682B4')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('ALIGN', (0,1), (0,-1), 'CENTER'),
            ('ALIGN', (1,1), (1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ])
        criteria_table.setStyle(criteria_style)
        story.append(criteria_table)
        story.append(Spacer(1, 16))

        assessment_text = """
        <b>Day-6 Assessment should be 100%</b><br/>
        <b>Passing Criteria: 100%</b><br/>
        <i>If failed in evaluation re-training is required</i>
        """
        story.append(Paragraph(assessment_text, styles['Normal']))


from django.http import HttpResponse, JsonResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
from io import BytesIO
import json
import traceback
from .models import (
    LevelThreeTraineeInfo, LevelThreeTrainingTopic, LevelThreeOJTDay, 
    LevelThreeOJTScore, LevelThreeSubStation, LevelThreeLine
)

@method_decorator(csrf_exempt, name='dispatch')
class LevelThreeOJTReportPDFView(View):
    def post(self, request, *args, **kwargs):
        """
        Handle PDF generation requests for Level Three OJT production reports
        """
        try:
            print("\n=== Received Level Three OJT PDF generation request ===")
            
            # 1. Parse input data
            trainee_id = self._get_trainee_id(request)
            station_id = self._get_station_id(request)
            
            if not trainee_id:
                return JsonResponse({'error': 'trainee_id is required'}, status=400)
            if not station_id:
                return JsonResponse({'error': 'station_id is required'}, status=400)
            
            print(f"Processing trainee_id: {trainee_id}, station_id: {station_id}")

            # 2. Get trainee record
            trainee = self._get_trainee_record(trainee_id, station_id)
            if isinstance(trainee, JsonResponse):
                return trainee
            if not trainee:
                return JsonResponse({'error': 'Trainee not found for the specified trainee_id and station_id'}, status=404)

            print(f"Found trainee: {trainee.trainee_name}")

            # 3. Generate PDF content
            print("Generating Level Three OJT PDF content...")
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), 
                                  leftMargin=0.5*inch, rightMargin=0.5*inch,
                                  topMargin=0.5*inch, bottomMargin=0.5*inch)
            story = self.create_ojt_pdf_content(trainee)
            
            # 4. Build PDF document
            print("Building Level Three OJT PDF document...")
            doc.build(story)
            buffer.seek(0)
            print("Level Three OJT PDF generation completed successfully")

            # 5. Return PDF response
            response = HttpResponse(
                buffer.getvalue(), 
                content_type='application/pdf'
            )
            response['Content-Disposition'] = (
                f'attachment; filename="level_three_ojt_report_{trainee_id}_production.pdf"'
            )
            return response
            
        except Exception as e:
            print("\n!!! Level Three OJT PDF generation failed !!!")
            traceback.print_exc()
            return JsonResponse(
                {
                    'error': 'Internal server error',
                    'detail': str(e),
                    'traceback': traceback.format_exc()
                }, 
                status=500
            )

    def _get_trainee_id(self, request):
        """Helper method to extract trainee_id from request"""
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                return data.get('trainee_id')
            except json.JSONDecodeError:
                return None
        return request.POST.get('trainee_id')

    def _get_station_id(self, request):
        """Helper method to extract station_id from request"""
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                return data.get('station_id')
            except json.JSONDecodeError:
                return None
        return request.POST.get('station_id')

    def _get_trainee_record(self, trainee_id, station_id):
        """Get trainee record for Level Three production"""
        try:
            return LevelThreeTraineeInfo.objects.get(trainee_id=trainee_id, station_id=station_id)
        except LevelThreeTraineeInfo.DoesNotExist:
            return None
        except LevelThreeTraineeInfo.MultipleObjectsReturned:
            try:
                return (LevelThreeTraineeInfo.objects
                        .filter(trainee_Id=trainee_id, station_id=station_id)
                        .order_by('-id')
                        .first())
            except Exception as e:
                return JsonResponse(
                    {
                        'error': f'Multiple trainee records found for trainee_id {trainee_id} and station_id {station_id}, '
                                f'and fallback selection failed: {str(e)}'
                    },
                    status=400
                )

    def create_ojt_pdf_content(self, trainee):
        """Generate the Level Three OJT PDF content structure for production"""
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = "Level-3 OJT MONITORING SHEET - PRODUCTION"
        story.append(Paragraph(title, styles['Title']))
        story.append(Spacer(1, 12))
        
        # Add all sections
        self._add_trainee_info(story, styles, trainee)
        self._add_ojt_monitoring_table(story, styles, trainee)
        self._add_judgement_criteria(story, styles)
        
        return story

    def _add_trainee_info(self, story, styles, trainee):
        """Add trainee basic information for Level Three production"""
        # Get station_name from the ForeignKey
        station_name = trainee.station.name if trainee.station else "Not Assigned"

        # Get line_name from the ForeignKey
        line_name = trainee.line.name if trainee.line else "Not Assigned"

        info_data = [
            ["Trainee Name:", trainee.trainee_name, "Trainee Id:", trainee.trainee_id],
            ["Station Name:", station_name, "Line Name:", line_name],
            ["Trainer Name:", trainee.trainer_name, "Training Status:", trainee.training_status]
        ]
        
        info_table = Table(info_data, colWidths=[100, 200, 100, 200])
        info_style = TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('ALIGN', (0,0), (0,-1), 'LEFT'),
            ('ALIGN', (2,0), (2,-1), 'LEFT'),
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ])
        info_table.setStyle(info_style)
        story.append(info_table)
        story.append(Spacer(1, 16))

    def _add_ojt_monitoring_table(self, story, styles, trainee):
        """Add the main OJT monitoring table for Level Three production"""
        topics = LevelThreeTrainingTopic.objects.all().order_by('sl_no')
        days = LevelThreeOJTDay.objects.all().order_by('id')
        scores = LevelThreeOJTScore.objects.filter(trainee=trainee)

        header = ["S.NO.", "TRAINING TOPIC", "Date"]
        for day in days:
            header.append(day.name)
        
        table_data = [header]
        
        for topic in topics:
            row = [str(topic.sl_no), topic.topic, topic.date]
            for day in days:
                score_obj = scores.filter(topic=topic, day=day).first()
                score_display = str(score_obj.score) if score_obj else ""
                row.append(score_display)
            table_data.append(row)

        base_widths = [40, 300, 80]
        day_width = (720 - sum(base_widths)) / len(days) if days else 60
        col_widths = base_widths + [day_width] * len(days)

        monitoring_table = Table(table_data, colWidths=col_widths)
        monitoring_style = self._get_monitoring_table_style(len(topics), len(days))
        monitoring_table.setStyle(monitoring_style)
        
        story.append(monitoring_table)
        story.append(Spacer(1, 16))

    def _get_monitoring_table_style(self, topic_count, day_count):
        """Get styling for the monitoring table"""
        style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4682B4')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('TOPPADDING', (0,0), (-1,0), 8),
            ('ALIGN', (0,1), (0,-1), 'CENTER'),
            ('ALIGN', (1,1), (1,-1), 'LEFT'),
            ('ALIGN', (2,1), (2,-1), 'CENTER'),
            ('ALIGN', (3,1), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F8F8')]),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ])
        
        return style

    def _add_judgement_criteria(self, story, styles):
        """Add judgement criteria section for Level Three"""
        story.append(Paragraph("Judgement Criteria", styles['Heading3']))
        story.append(Spacer(1, 8))
        
        criteria_data = [
            ["Criteria", "Description"],
            ["10 Marks", "OK - Trainee demonstrates competency"],
            ["0 Marks", "Not OK - Trainee needs improvement"]
        ]
        
        criteria_table = Table(criteria_data, colWidths=[100, 400])
        criteria_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4682B4')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('ALIGN', (0,1), (0,-1), 'CENTER'),
            ('ALIGN', (1,1), (1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ])
        criteria_table.setStyle(criteria_style)
        story.append(criteria_table)
        story.append(Spacer(1, 16))

        assessment_text = """
        <b>Assessment should be 100%</b><br/>
        <b>Passing Criteria: 100%</b><br/>
        <i>If failed in evaluation re-training is required</i>
        """
        story.append(Paragraph(assessment_text, styles['Normal']))


        #The END





from rest_framework import viewsets
from .models import FactoryStructure
from .serializers import FactoryStructureSerializer

class FactoryStructureViewSet(viewsets.ModelViewSet):
    queryset = FactoryStructure.objects.all()
    serializer_class = FactoryStructureSerializer

    



from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import ProductionPlan
from .serializers import ProductionPlanSerializer

# class ProductionPlanViewSet(viewsets.ModelViewSet):
#     queryset = ProductionPlan.objects.all()
#     serializer_class = ProductionPlanSerializer

#     def create(self, request, *args, **kwargs):
#         # Handle bulk creation
#         if isinstance(request.data, list):
#             serializer = self.get_serializer(data=request.data, many=True)
#             serializer.is_valid(raise_exception=True)
#             self.perform_bulk_create(serializer)
#             return Response(serializer.data, status=status.HTTP_201_CREATED)
#         else:
#             return super().create(request, *args, **kwargs)

#     def perform_bulk_create(self, serializer):
#         serializer.save()

#     def get_queryset(self):
#         queryset = super().get_queryset()
#         factory = self.request.query_params.get('factory')
#         shop_floor = self.request.query_params.get('shop_floor')
#         line = self.request.query_params.get('line')
#         station = self.request.query_params.get('station')
#         year = self.request.query_params.get('year')
#         month = self.request.query_params.get('month')

#         if factory:
#             queryset = queryset.filter(factory_id=factory)
#         if shop_floor:
#             queryset = queryset.filter(shop_floor_id=shop_floor)
#         if line:
#             queryset = queryset.filter(line_id=line)
#         if station:
#             queryset = queryset.filter(station_id=station)
#         if year:
#             queryset = queryset.filter(year=year)
#         if month:
#             queryset = queryset.filter(month=month)

#         return queryset


from rest_framework import viewsets, status
from rest_framework.response import Response
from .models import ProductionPlan
from .serializers import ProductionPlanSerializer

from rest_framework import viewsets, status
from rest_framework.response import Response
from decimal import Decimal, ROUND_HALF_UP
from .models import ProductionPlan
from .serializers import ProductionPlanSerializer

class ProductionPlanViewSet(viewsets.ModelViewSet):
    queryset = ProductionPlan.objects.all()
    serializer_class = ProductionPlanSerializer

    def create(self, request, *args, **kwargs):
        # Handle bulk creation
        if isinstance(request.data, list):
            serializer = self.get_serializer(data=request.data, many=True)
            serializer.is_valid(raise_exception=True)
            self.perform_bulk_create(serializer)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return super().create(request, *args, **kwargs)

    def perform_bulk_create(self, serializer):
        serializer.save()

    def get_queryset(self):
        queryset = super().get_queryset()
        factory = self.request.query_params.get('factory')
        shop_floor = self.request.query_params.get('shop_floor')
        line = self.request.query_params.get('line')
        station = self.request.query_params.get('station')
        year = self.request.query_params.get('year')
        month = self.request.query_params.get('month')

        if factory:
            queryset = queryset.filter(factory_id=factory)
        if shop_floor:
            queryset = queryset.filter(shop_floor_id=shop_floor)
        if line:
            queryset = queryset.filter(line_id=line)
        if station:
            queryset = queryset.filter(station_id=station)
        if year:
            queryset = queryset.filter(year=year)
        if month:
            queryset = queryset.filter(month=month)

        return queryset
    
    def get_previous_month_plan(self, current_plan):
        """Get previous month's production plan"""
        month_order = [
            'January', 'February', 'March', 'April', 'May', 'June',
            'July', 'August', 'September', 'October', 'November', 'December'
        ]
        
        try:
            current_index = month_order.index(current_plan.month)
            if current_index == 0:  # January
                prev_month = 'December'
                prev_year = current_plan.year - 1
            else:
                prev_month = month_order[current_index - 1]
                prev_year = current_plan.year
            
            return ProductionPlan.objects.filter(
                month=prev_month,
                year=prev_year,
                factory=current_plan.factory,
                shop_floor=current_plan.shop_floor,
                line=current_plan.line,
                station=current_plan.station
            ).first()
        except (ValueError, ProductionPlan.DoesNotExist):
            return None

    def calculate_4_block_gap_data(self, current_plan):
        """
        Calculate the 4-block graph data exactly as shown in Excel:
        1. Production Plan (A) - From Plan input
        2. Operators Required (B) - From Plan input  
        3. Operators Available (C) - From Last month Actual
        4. Gap (B-C) - Difference between Required and Available
        """
        previous_plan = self.get_previous_month_plan(current_plan)
        
        # Block A: Production Plan (from current month plan input)
        production_plan = current_plan.total_production_plan
        
        # Block B: Operators Required (from current month plan input)
        operators_required = current_plan.total_operators_required_plan
        
        # Block C: Operators Available (from previous month actual, adjusted for attrition)
        if previous_plan and previous_plan.total_operators_required_actual > 0:
            attrition_decimal = current_plan.attrition_rate / Decimal('100')
            operators_available = int(
                Decimal(previous_plan.total_operators_required_actual) * (Decimal('1') - attrition_decimal)
            )
        else:
            operators_available = 0
        
        # Block D: Gap = Operators Required - Operators Available (B - C)
        gap = operators_required - operators_available
        
        return {
            'month': current_plan.month,
            'year': current_plan.year,
            'production_plan': {
                'value': production_plan,
                'label': 'Production Plan',
                'description': 'From Plan input',
                'code': 'A'
            },
            'operators_required': {
                'value': operators_required,
                'label': 'Operators Required',
                'description': 'From Plan input', 
                'code': 'B'
            },
            'operators_available': {
                'value': operators_available,
                'label': 'Operators Available',
                'description': 'From Last month Actual',
                'code': 'C'
            },
            'gap': {
                'value': gap,
                'label': 'Gap',
                'description': 'B-C',
                'code': 'B-C',
                'is_shortage': gap > 0  # Positive gap means shortage (need to hire)
            },
            'has_previous_data': previous_plan is not None,
            'previous_month_data': {
                'month': previous_plan.month if previous_plan else None,
                'actual_operators': previous_plan.total_operators_required_actual if previous_plan else 0,
                'actual_production': previous_plan.total_production_actual if previous_plan else 0
            } if previous_plan else None,
            'attrition_rate': float(current_plan.attrition_rate)
        }

    def calculate_level_wise_hiring(self, current_plan):
        """Calculate level-wise hiring requirements (L1, L2, L3, L4)"""
        previous_plan = self.get_previous_month_plan(current_plan)
        
        levels = ['l1', 'l2', 'l3', 'l4']
        level_data = []
        
        for level in levels:
            # Current month required for this level
            current_required = (
                getattr(current_plan, f'ctq_plan_{level}', 0) +
                getattr(current_plan, f'pdi_plan_{level}', 0) +
                getattr(current_plan, f'other_plan_{level}', 0) +
                getattr(current_plan, f'bifurcation_plan_{level}', 0)
            )
            
            # Previous month actual for this level
            if previous_plan:
                previous_actual = (
                    getattr(previous_plan, f'ctq_actual_{level}', 0) +
                    getattr(previous_plan, f'pdi_actual_{level}', 0) +
                    getattr(previous_plan, f'other_actual_{level}', 0) +
                    getattr(previous_plan, f'bifurcation_actual_{level}', 0)
                )
                # Adjust for attrition
                attrition_decimal = current_plan.attrition_rate / Decimal('100')
                available = int(Decimal(previous_actual) * (Decimal('1') - attrition_decimal))
            else:
                available = 0
            
            # Calculate gap for this level
            level_gap = current_required - available
            
            level_data.append({
                'level': level.upper(),
                'required': current_required,
                'available': available,
                'gap': level_gap,
                'hiring_required': max(0, level_gap)  # Only positive gaps indicate hiring needed
            })
        
        return level_data

    @action(detail=False, methods=['get'])
    def gap_graph_data(self, request):
        """
        Get gap data formatted for the 4-block graph visualization
        URL: /api/production-plans/gap_graph_data/
        """
        factory = request.query_params.get('factory')
        month = request.query_params.get('month')
        year = request.query_params.get('year')
        shop_floor = request.query_params.get('shop_floor')
        line = request.query_params.get('line')
        station = request.query_params.get('station')
        
        if not all([factory, month, year]):
            return Response(
                {'error': 'Factory, month, and year parameters are required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get current month's production plan
        plan_query = ProductionPlan.objects.filter(
            factory_id=factory,
            month__iexact=month,
            year=year
        )
        
        if shop_floor:
            plan_query = plan_query.filter(shop_floor_id=shop_floor)
        if line:
            plan_query = plan_query.filter(line_id=line)
        if station:
            plan_query = plan_query.filter(station_id=station)
            
        current_plan = plan_query.first()
        
        if not current_plan:
            return Response(
                {'error': f'No production plan found for {month} {year}'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Calculate 4-block data
        four_block_data = self.calculate_4_block_gap_data(current_plan)
        
        # Calculate level-wise hiring data
        level_wise_data = self.calculate_level_wise_hiring(current_plan)
        
        return Response({
            'success': True,
            'operators_availability_graph': four_block_data,
            'level_wise_hiring_graph': level_wise_data,
            'graph_title': f"Gap Analysis for {month} {year}",
            'subtitle': "This month Required - Last month Actual"
        })


    # In your views.py file, inside the ProductionPlanViewSet class

    @action(detail=False, methods=['get'], url_path='yearly-productivity')
    def yearly_productivity(self, request):
        """
        Calculates productivity for a full year, aggregated by month.
        FORMULA: total_production_plan / total_operators_required_actual
        URL: /api/production-plans/yearly-productivity/?factory=1&year=2024
        """
        factory_id = request.query_params.get('factory')
        year = request.query_params.get('year')

        if not factory_id or not year:
            return Response(
                {'error': 'Factory and year parameters are required.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        queryset = ProductionPlan.objects.filter(
            factory_id=factory_id,
            year=year
        ).order_by('created_at')

        monthly_data = {}
        month_order = [
            'January', 'February', 'March', 'April', 'May', 'June',
            'July', 'August', 'September', 'October', 'November', 'December'
        ]

        for plan in queryset:
            monthly_data[plan.month] = {
                "month": plan.month,
                # --- CHANGE 1: Using the correct fields from your formula ---
                "planned_production": plan.total_production_plan,
                "actual_manpower": plan.total_operators_required_actual
            }

        final_data = []
        for month_name in month_order:
            if month_name in monthly_data:
                final_data.append(monthly_data[month_name])
            else:
                final_data.append({
                    "month": month_name,
                    # --- CHANGE 2: Ensuring the default values match ---
                    "planned_production": 0,
                    "actual_manpower": 0
                })
        
        return Response({
            'success': True,
            'graph_title': f"Monthly Productivity for {year}",
            'subtitle': "Planned Production / Actual Manpower",
            'data': final_data
        })



# In your views.py

# --- Django & DRF Imports ---
from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.db import transaction

# --- Django DB Imports for Aggregation ---
from django.db.models import Sum, Avg
from django.db.models.functions import TruncDay, TruncMonth

# --- Python Standard Library Imports ---
from datetime import date, timedelta
from collections import OrderedDict

# --- Model & Serializer Imports ---
from .models import DailyProductionData
from .serializers import DailyProductionDataSerializer
from types import SimpleNamespace

# ===================================================================
# ### THE CORRECTED AND FINAL VIEWSET ###
class DailyProductionDataViewSet(viewsets.ModelViewSet):
    """
    ViewSet for handling daily production data, including monthly and weekly aggregations.
    """
    queryset = DailyProductionData.objects.all()
    serializer_class = DailyProductionDataSerializer

    # --- 1. THE MONTHLY ENDPOINT ---
    @action(detail=False, methods=['get'], url_path='monthly-summary')
    def monthly_summary(self, request):
        factory = request.query_params.get('factory')
        shop_floor = request.query_params.get('shop_floor')
        line = request.query_params.get('line')
        station = request.query_params.get('station')
        month_str = request.query_params.get('month')
        year_str = request.query_params.get('year')
        
        if not all([factory, month_str, year_str]):
            return Response({'error': 'Factory, month, and year are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            month = int(month_str)
            year = int(year_str)
        except (ValueError, TypeError):
            return Response({'error': 'Invalid month or year provided.'}, status=status.HTTP_400_BAD_REQUEST)
        
        filters = {'factory_id': factory, 'date__year': year, 'date__month': month}
        if shop_floor and shop_floor != 'all': filters['shop_floor_id'] = shop_floor
        if line and line != 'all': filters['line_id'] = line
        if station and station != 'all': filters['station_id'] = station
        
        aggregation = DailyProductionData.objects.filter(**filters).aggregate(
            # Standard fields
            attrition_rate=Avg('attrition_rate'), absenteeism_rate=Sum('absenteeism_rate'),
            total_production_plan=Sum('total_production_plan'), total_production_actual=Sum('total_production_actual'),
            total_operators_required_plan=Sum('total_operators_required_plan'), total_operators_required_actual=Sum('total_operators_required_actual'),
            
            # Department fields (optional, but good to keep)
            ctq_plan_l1=Sum('ctq_plan_l1'), ctq_actual_l1=Sum('ctq_actual_l1'), ctq_plan_l2=Sum('ctq_plan_l2'), ctq_actual_l2=Sum('ctq_actual_l2'), ctq_plan_l3=Sum('ctq_plan_l3'), ctq_actual_l3=Sum('ctq_actual_l3'), ctq_plan_l4=Sum('ctq_plan_l4'), ctq_actual_l4=Sum('ctq_actual_l4'),
            pdi_plan_l1=Sum('pdi_plan_l1'), pdi_actual_l1=Sum('pdi_actual_l1'), pdi_plan_l2=Sum('pdi_plan_l2'), pdi_actual_l2=Sum('pdi_actual_l2'), pdi_plan_l3=Sum('pdi_plan_l3'), pdi_actual_l3=Sum('pdi_actual_l3'), pdi_plan_l4=Sum('pdi_plan_l4'), pdi_actual_l4=Sum('pdi_actual_l4'),
            other_plan_l1=Sum('other_plan_l1'), other_actual_l1=Sum('other_actual_l1'), other_plan_l2=Sum('other_plan_l2'), other_actual_l2=Sum('other_actual_l2'), other_plan_l3=Sum('other_plan_l3'), other_actual_l3=Sum('other_actual_l3'), other_plan_l4=Sum('other_plan_l4'), other_actual_l4=Sum('other_actual_l4'),

            # =========================================================================
            # ✨ CORRECTED: Add the bifurcation fields to the aggregation
            # =========================================================================
            bifurcation_plan_l1=Sum('bifurcation_plan_l1'), bifurcation_actual_l1=Sum('bifurcation_actual_l1'),
            bifurcation_plan_l2=Sum('bifurcation_plan_l2'), bifurcation_actual_l2=Sum('bifurcation_actual_l2'),
            bifurcation_plan_l3=Sum('bifurcation_plan_l3'), bifurcation_actual_l3=Sum('bifurcation_actual_l3'),
            bifurcation_plan_l4=Sum('bifurcation_plan_l4'), bifurcation_actual_l4=Sum('bifurcation_actual_l4'),

        )
        
        # Replace any None results (from empty queries) with 0
        for key, value in aggregation.items():
            if value is None: aggregation[key] = 0
        
        # =========================================================================
        # ✨ REMOVED: The manual calculation is no longer needed
        # =========================================================================
        
        return Response(aggregation)

    # --- 2. THE WEEKLY ENDPOINT ---
    @action(detail=False, methods=['get'], url_path='aggregated-weekly-data')
    def aggregated_weekly_data(self, request):
        factory = request.query_params.get('factory')
        shop_floor = request.query_params.get('shop_floor')
        line = request.query_params.get('line')
        station = request.query_params.get('station')
        start_date_str = request.query_params.get('start_date')
        
        if not all([factory, start_date_str]):
            return Response({'error': 'Factory and start_date are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            start_date = date.fromisoformat(start_date_str)
            end_date = start_date + timedelta(days=6)
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)
        
        filters = {'factory_id': factory, 'date__range': [start_date, end_date]}
        if shop_floor and shop_floor != 'all': filters['shop_floor_id'] = shop_floor
        if line and line != 'all': filters['line_id'] = line
        if station and station != 'all': filters['station_id'] = station
        
        aggregation = DailyProductionData.objects.filter(**filters).aggregate(
            # Standard fields
            attrition_rate=Avg('attrition_rate'), absenteeism_rate=Sum('absenteeism_rate'),
            total_production_plan=Sum('total_production_plan'), total_production_actual=Sum('total_production_actual'),
            total_operators_required_plan=Sum('total_operators_required_plan'), total_operators_required_actual=Sum('total_operators_required_actual'),
            
            # Department fields
            ctq_plan_l1=Sum('ctq_plan_l1'), ctq_actual_l1=Sum('ctq_actual_l1'), ctq_plan_l2=Sum('ctq_plan_l2'), ctq_actual_l2=Sum('ctq_actual_l2'), ctq_plan_l3=Sum('ctq_plan_l3'), ctq_actual_l3=Sum('ctq_actual_l3'), ctq_plan_l4=Sum('ctq_plan_l4'), ctq_actual_l4=Sum('ctq_actual_l4'),
            pdi_plan_l1=Sum('pdi_plan_l1'), pdi_actual_l1=Sum('pdi_actual_l1'), pdi_plan_l2=Sum('pdi_plan_l2'), pdi_actual_l2=Sum('pdi_actual_l2'), pdi_plan_l3=Sum('pdi_plan_l3'), pdi_actual_l3=Sum('pdi_actual_l3'), pdi_plan_l4=Sum('pdi_plan_l4'), pdi_actual_l4=Sum('pdi_actual_l4'),
            other_plan_l1=Sum('other_plan_l1'), other_actual_l1=Sum('other_actual_l1'), other_plan_l2=Sum('other_plan_l2'), other_actual_l2=Sum('other_actual_l2'), other_plan_l3=Sum('other_plan_l3'), other_actual_l3=Sum('other_actual_l3'), other_plan_l4=Sum('other_plan_l4'), other_actual_l4=Sum('other_actual_l4'),

            # =========================================================================
            # ✨ CORRECTED: Add the bifurcation fields to the aggregation
            # =========================================================================
            bifurcation_plan_l1=Sum('bifurcation_plan_l1'), bifurcation_actual_l1=Sum('bifurcation_actual_l1'),
            bifurcation_plan_l2=Sum('bifurcation_plan_l2'), bifurcation_actual_l2=Sum('bifurcation_actual_l2'),
            bifurcation_plan_l3=Sum('bifurcation_plan_l3'), bifurcation_actual_l3=Sum('bifurcation_actual_l3'),
            bifurcation_plan_l4=Sum('bifurcation_plan_l4'), bifurcation_actual_l4=Sum('bifurcation_actual_l4'),

        )
        
        for key, value in aggregation.items():
            if value is None: aggregation[key] = 0
        
        return Response(aggregation)

    # --- 3. THE SAVE ENDPOINT (This was already correct) ---
    @action(detail=False, methods=['post'], url_path='save-batch')
    @transaction.atomic
    def save_batch(self, request):
        """
        Handles creating/updating a batch of daily records for a month or week.
        """
        data = request.data
        if not isinstance(data, list) or not data:
            return Response({'error': 'Expected a non-empty list of daily data objects.'}, status=status.HTTP_400_BAD_REQUEST)

        dates_to_clear = {item['date'] for item in data if 'date' in item}
        first_item = data[0]
        context_filters = {
            'factory_id': first_item.get('factory'),
            'shop_floor_id': first_item.get('shop_floor'),
            'line_id': first_item.get('line'),
            'station_id': first_item.get('station'),
        }
        delete_filters = {k: v for k, v in context_filters.items() if v is not None}
        delete_filters['date__in'] = list(dates_to_clear)
        
        DailyProductionData.objects.filter(**delete_filters).delete()

        try:
            serializer = self.get_serializer(data=data, many=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    # --- 4. THE TREND ENDPOINT (Unchanged) ---
    @action(detail=False, methods=['get'], url_path='productivity-trend')
    def productivity_trend(self, request):
        factory_id = request.query_params.get('factory')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        group_by = request.query_params.get('group_by', 'month').lower()

        if not all([factory_id, start_date_str, end_date_str]):
            return Response({'error': 'Factory, start_date, and end_date are required.'}, status=400)

        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)
        
        if group_by == 'monthly':
            trunc_func = TruncMonth('date')
            date_format = "%b %Y"
        else:
            trunc_func = TruncDay('date')
            date_format = "%b %d"

        db_data = DailyProductionData.objects.filter(
            factory_id=factory_id,
            date__range=[start_date, end_date]
        ).annotate(period=trunc_func).values('period').annotate(
            planned_production=Sum('total_production_plan'),
            actual_manpower=Sum('total_operators_required_actual')
        ).order_by('period')

        data_map = {item['period'].strftime('%Y-%m-%d'): item for item in db_data}

        full_range_data = OrderedDict()
        current_date = start_date

        while current_date <= end_date:
            if group_by == 'monthly':
                period_start = current_date.replace(day=1)
            else:
                period_start = current_date
            
            period_key = period_start.strftime('%Y-%m-%d')
            if period_key not in full_range_data:
                full_range_data[period_key] = {
                    'period_label': period_start.strftime(date_format),
                    'planned_production': 0,
                    'actual_manpower': 0
                }
            
            current_date += timedelta(days=1)
        
        for key, values in data_map.items():
            if key in full_range_data:
                full_range_data[key]['planned_production'] = values['planned_production'] or 0
                full_range_data[key]['actual_manpower'] = values['actual_manpower'] or 0

        return Response({
            'success': True,
            'graph_title': 'Productivity Trend',
            'data': list(full_range_data.values())
        })
    

    @action(detail=False, methods=['get'], url_path='trend-data')
    def trend_data(self, request):
        """
        A single, powerful endpoint to generate time-series data for various trend graphs.
        This version ensures a continuous timeline by filling in any missing data points with zeros.
        """
        # --- 1. Get and Validate Parameters ---
        factory_id = request.query_params.get('factory')
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        data_key = request.query_params.get('data_key')

        if not all([factory_id, start_date_str, end_date_str, data_key]):
            return Response({'error': 'factory, start_date, end_date, and data_key are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Use the corrected date parsing method
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

        # --- 2. Define the Metrics We Can Plot ---
        METRIC_MAP = {
            'operators': {
                'planned': 'total_operators_required_plan',
                'actual': 'total_operators_required_actual'
            },
            'production': {
                'planned': 'total_production_plan',
                'actual': 'total_production_actual'
            },
        }
        if data_key not in METRIC_MAP:
            return Response({'error': f"Invalid data_key. Choose from: {list(METRIC_MAP.keys())}"}, status=status.HTTP_400_BAD_REQUEST)

        metric_fields = METRIC_MAP[data_key]
        
        # --- 3. Build Filters and Grouping ---
        filters = {'factory_id': factory_id, 'date__range': [start_date, end_date]}
        for param in ['shop_floor', 'line', 'station']:
            param_id = request.query_params.get(param)
            if param_id:
                filters[f'{param}_id'] = param_id
        
        group_by = request.query_params.get('group_by', 'monthly').lower()
        
        if group_by == 'daily':
            trunc_func = TruncDay('date')
            date_format = "%b %d"
        else: # monthly
            trunc_func = TruncMonth('date')
            date_format = "%b '%y"
            
        # --- 4. The Database Query ---
        sum_annotations = {key: Sum(field) for key, field in metric_fields.items()}
        aggregated_data = DailyProductionData.objects.filter(**filters).annotate(
            period=trunc_func
        ).values('period', 'entry_mode').annotate(**sum_annotations).order_by('period')

        # --- 5. Process with Priority Logic (MONTHLY > WEEKLY > DAILY) ---
        final_period_totals = {}
        priority_map = {'MONTHLY': 3, 'WEEKLY': 2, 'DAILY': 1}
        for item in aggregated_data:
            period_key = item['period']
            current_priority = priority_map.get(item['entry_mode'], 0)
            if period_key not in final_period_totals or current_priority > final_period_totals[period_key].get('priority', 0):
                data_point = {"priority": current_priority}
                for key in metric_fields.keys():
                    data_point[key] = item[key] or 0
                final_period_totals[period_key] = data_point
        
        # --- 6. Fill in the gaps to ensure a complete timeline (FIXED LOGIC) ---
        full_range_data = OrderedDict()

        if group_by == 'monthly':
            m, y = start_date.month, start_date.year
            end_m, end_y = end_date.month, end_date.year
            
            while (y, m) <= (end_y, end_m):
                period_key = date(y, m, 1)
                template = {"name": period_key.strftime(date_format)}
                for key in metric_fields.keys():
                    template[key] = 0
                full_range_data[period_key] = template
                m += 1
                if m > 12:
                    m = 1
                    y += 1
        else: # daily
            current_d = start_date
            delta = timedelta(days=1)
            while current_d <= end_date:
                period_key = current_d
                template = {"name": period_key.strftime(date_format)}
                for key in metric_fields.keys():
                    template[key] = 0
                full_range_data[period_key] = template
                current_d += delta
        
        # --- 7. Merge real data into the template ---
        for period_key, values in final_period_totals.items():
            if period_key in full_range_data:
                for key in metric_fields.keys():
                    full_range_data[period_key][key] = values.get(key, 0)

        # --- 8. Return the final, complete list ---
        return Response(list(full_range_data.values()))



    @action(detail=False, methods=['get'], url_path='date-range-summary')
    def date_range_summary(self, request):
            """
            Provides a single aggregated summary object for any given date range.
            This is used for the top summary cards and the Operator Stats component.
            """
            # --- 1. Get Parameters ---
            factory = request.query_params.get('factory')
            start_date_str = request.query_params.get('start_date')
            end_date_str = request.query_params.get('end_date')
            shop_floor = request.query_params.get('shop_floor')
            line = request.query_params.get('line')
            station = request.query_params.get('station')

            if not all([factory, start_date_str, end_date_str]):
                return Response({'error': 'Factory, start_date, and end_date are required.'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response({'error': 'Invalid date format. Use YYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

            # --- 2. Build Filters ---
            filters = {'factory_id': factory, 'date__range': [start_date, end_date]}
            if shop_floor and shop_floor != 'all': filters['shop_floor_id'] = shop_floor
            if line and line != 'all': filters['line_id'] = line
            if station and station != 'all': filters['station_id'] = station
            
            # --- 3. Aggregate All Fields ---
            aggregation = DailyProductionData.objects.filter(**filters).aggregate(
                attrition_rate=Avg('attrition_rate'), absenteeism_rate=Avg('absenteeism_rate'),
                total_production_plan=Sum('total_production_plan'), total_production_actual=Sum('total_production_actual'),
                total_operators_required_plan=Sum('total_operators_required_plan'), total_operators_required_actual=Sum('total_operators_required_actual'),
                
                ctq_plan_l1=Sum('ctq_plan_l1'), ctq_actual_l1=Sum('ctq_actual_l1'), ctq_plan_l2=Sum('ctq_plan_l2'), ctq_actual_l2=Sum('ctq_actual_l2'), ctq_plan_l3=Sum('ctq_plan_l3'), ctq_actual_l3=Sum('ctq_actual_l3'), ctq_plan_l4=Sum('ctq_plan_l4'), ctq_actual_l4=Sum('ctq_actual_l4'),
                pdi_plan_l1=Sum('pdi_plan_l1'), pdi_actual_l1=Sum('pdi_actual_l1'), pdi_plan_l2=Sum('pdi_plan_l2'), pdi_actual_l2=Sum('pdi_actual_l2'), pdi_plan_l3=Sum('pdi_plan_l3'), pdi_actual_l3=Sum('pdi_actual_l3'), pdi_plan_l4=Sum('pdi_plan_l4'), pdi_actual_l4=Sum('pdi_actual_l4'),
                other_plan_l1=Sum('other_plan_l1'), other_actual_l1=Sum('other_actual_l1'), other_plan_l2=Sum('other_plan_l2'), other_actual_l2=Sum('other_actual_l2'), other_plan_l3=Sum('other_plan_l3'), other_actual_l3=Sum('other_actual_l3'), other_plan_l4=Sum('other_plan_l4'), other_actual_l4=Sum('other_actual_l4'),
                bifurcation_plan_l1=Sum('bifurcation_plan_l1'), bifurcation_actual_l1=Sum('bifurcation_actual_l1'),
                bifurcation_plan_l2=Sum('bifurcation_plan_l2'), bifurcation_actual_l2=Sum('bifurcation_actual_l2'),
                bifurcation_plan_l3=Sum('bifurcation_plan_l3'), bifurcation_actual_l3=Sum('bifurcation_actual_l3'),
                bifurcation_plan_l4=Sum('bifurcation_plan_l4'), bifurcation_actual_l4=Sum('bifurcation_actual_l4'),
            )
            
            # Replace any None results (from empty queries) with 0
            for key, value in aggregation.items():
                if value is None: aggregation[key] = 0
            
            return Response(aggregation)
    
    @action(detail=False, methods=['get'], url_path='gap-analysis')
    def gap_analysis(self, request):
        factory = request.query_params.get('factory')
        month_str = request.query_params.get('month')
        year_str = request.query_params.get('year')
        shop_floor = request.query_params.get('shop_floor')
        line = request.query_params.get('line')
        station = request.query_params.get('station')

        if not all([factory, month_str, year_str]):
            return Response({'error': 'Factory, month, and year are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            year = int(year_str)
            month_map = {name.lower(): num for num, name in enumerate([date(2000, i, 1).strftime('%B') for i in range(1, 13)], 1)}
            month = month_map[month_str.lower()]
        except (ValueError, KeyError):
            return Response({'error': 'Invalid month name or year format.'}, status=status.HTTP_400_BAD_REQUEST)

        filters = {'factory_id': factory}
        if shop_floor: filters['shop_floor_id'] = shop_floor
        if line: filters['line_id'] = line
        if station: filters['station_id'] = station

        # --- Use the NEW, more detailed helper function ---
        current_month_summary = self._get_detailed_summary_for_month(year, month, filters)
        if not current_month_summary or current_month_summary.get('total_operators_required_plan') is None:
            return Response({'error': f'No data found for {month_str} {year}'}, status=status.HTTP_404_NOT_FOUND)

        prev_month_date = (date(year, month, 1) - timedelta(days=1))
        prev_month_summary = self._get_detailed_summary_for_month(prev_month_date.year, prev_month_date.month, filters)

        # --- Convert dicts to objects to make calculations cleaner ---
        current_plan = SimpleNamespace(**{k: v or 0 for k, v in current_month_summary.items()})
        previous_plan = SimpleNamespace(**{k: v or 0 for k, v in prev_month_summary.items()}) if prev_month_summary else None

        # --- Perform the Business Logic Calculations ---
        # 1. Main Gap Calculation
        previous_actual_operators = getattr(previous_plan, 'total_operators_required_actual', 0)
        attrition_rate = Decimal(getattr(current_plan, 'attrition_rate', 0))
        attrition_decimal = attrition_rate / Decimal('100')
        operators_available = (Decimal(previous_actual_operators) * (Decimal('1') - attrition_decimal)).to_integral_value(rounding=ROUND_HALF_UP)
        gap = current_plan.total_operators_required_plan - operators_available

        # 2. Level-wise Hiring Calculation (FIXED)
        level_wise_data = self._calculate_level_wise_hiring(current_plan, previous_plan)

        # --- Format Response ---
        response_data = {
            "success": True,
            "operators_availability_graph": {
                "month": month_str, "year": year,
                "production_plan": {"value": current_plan.total_production_plan},
                "operators_required": {"value": current_plan.total_operators_required_plan},
                "operators_available": {"value": int(operators_available)},
                "gap": {"value": int(gap)},
                "attrition_rate": float(attrition_rate),
                "previous_month_data": { "month": prev_month_date.strftime('%B'), "actual_operators": previous_actual_operators } if previous_plan else None
            },
            "level_wise_hiring_graph": level_wise_data # This now contains correct data
        }
        return Response(response_data)

    # --- NEW HELPER METHOD ---
    def _get_detailed_summary_for_month(self, year, month, filters):
        """Helper to get a detailed monthly summary including all L1-L4 fields."""
        month_filters = filters.copy()
        month_filters['date__year'] = year
        month_filters['date__month'] = month
        
        aggregation = DailyProductionData.objects.filter(**month_filters).aggregate(
            attrition_rate=Avg('attrition_rate'),
            total_production_plan=Sum('total_production_plan'),
            total_operators_required_plan=Sum('total_operators_required_plan'),
            total_operators_required_actual=Sum('total_operators_required_actual'),
            ctq_plan_l1=Sum('ctq_plan_l1'), ctq_plan_l2=Sum('ctq_plan_l2'), ctq_plan_l3=Sum('ctq_plan_l3'), ctq_plan_l4=Sum('ctq_plan_l4'),
            pdi_plan_l1=Sum('pdi_plan_l1'), pdi_plan_l2=Sum('pdi_plan_l2'), pdi_plan_l3=Sum('pdi_plan_l3'), pdi_plan_l4=Sum('pdi_plan_l4'),
            other_plan_l1=Sum('other_plan_l1'), other_plan_l2=Sum('other_plan_l2'), other_plan_l3=Sum('other_plan_l3'), other_plan_l4=Sum('other_plan_l4'),
            ctq_actual_l1=Sum('ctq_actual_l1'), ctq_actual_l2=Sum('ctq_actual_l2'), ctq_actual_l3=Sum('ctq_actual_l3'), ctq_actual_l4=Sum('ctq_actual_l4'),
            pdi_actual_l1=Sum('pdi_actual_l1'), pdi_actual_l2=Sum('pdi_actual_l2'), pdi_actual_l3=Sum('pdi_actual_l3'), pdi_actual_l4=Sum('pdi_actual_l4'),
            other_actual_l1=Sum('other_actual_l1'), other_actual_l2=Sum('other_actual_l2'), other_actual_l3=Sum('other_actual_l3'), other_actual_l4=Sum('other_actual_l4'),
        )
        return aggregation

    # --- NEW HELPER METHOD ---
    def _calculate_level_wise_hiring(self, current_plan, previous_plan):
        """Calculates the hiring gap for each level (L1-L4)."""
        levels = ['l1', 'l2', 'l3', 'l4']
        level_data = []
        attrition_rate = Decimal(getattr(current_plan, 'attrition_rate', 0))
        attrition_decimal = attrition_rate / Decimal('100')

        for level in levels:
            current_required = (
                getattr(current_plan, f'ctq_plan_{level}', 0) +
                getattr(current_plan, f'pdi_plan_{level}', 0) +
                getattr(current_plan, f'other_plan_{level}', 0)
            )
            
            previous_actual = 0
            if previous_plan:
                previous_actual = (
                    getattr(previous_plan, f'ctq_actual_{level}', 0) +
                    getattr(previous_plan, f'pdi_actual_{level}', 0) +
                    getattr(previous_plan, f'other_actual_{level}', 0)
                )

            available = (Decimal(previous_actual) * (Decimal('1') - attrition_decimal)).to_integral_value(rounding=ROUND_HALF_UP)
            gap = current_required - available
            
            level_data.append({
                'level': level.upper(),
                'hiring_required': max(0, int(gap))
            })
        return level_data
    
    def _get_summary_for_range(self, start_date, end_date, filters):
        """Helper function to get aggregated data for a date range."""
        range_filters = filters.copy()
        range_filters['date__range'] = [start_date, end_date]
        
        aggregation = DailyProductionData.objects.filter(**range_filters).aggregate(
            attrition_rate=Avg('attrition_rate'),
            total_production_plan=Sum('total_production_plan'),
            total_operators_required_plan=Sum('total_operators_required_plan'),
            total_operators_required_actual=Sum('total_operators_required_actual'),
        )
        # Ensure we return a dictionary with 0s if no data is found, not None
        for key, value in aggregation.items():
            if value is None:
                aggregation[key] = 0
        return aggregation



    @action(detail=False, methods=['get'], url_path='weekly-gap-analysis')
    def weekly_gap_analysis(self, request):
        """
        Calculates the level-wise hiring gap for a selected week
        based on the previous week's actuals.
        Formula: (This Week's Plan) - (Last Week's Actual * Attrition)
        """
        # --- 1. Get Parameters ---
        factory = request.query_params.get('factory')
        start_date_str = request.query_params.get('start_date')
        shop_floor = request.query_params.get('shop_floor')
        line = request.query_params.get('line')
        station = request.query_params.get('station')

        if not all([factory, start_date_str]):
            return Response({'error': 'Factory and start_date are required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Date range for the CURRENT week
            current_week_start = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            
            # Date range for the PREVIOUS week
            previous_week_start = current_week_start - timedelta(days=7)

        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

        # --- 2. Get Data for CURRENT and PREVIOUS Weeks ---
        filters = {'factory_id': factory}
        if shop_floor: filters['shop_floor_id'] = shop_floor
        if line: filters['line_id'] = line
        if station: filters['station_id'] = station

        current_week_summary = self._get_detailed_summary_for_week(current_week_start, filters)
        previous_week_summary = self._get_detailed_summary_for_week(previous_week_start, filters)

        # --- 3. Perform the Calculation ---
        current_plan = SimpleNamespace(**{k: v or 0 for k, v in current_week_summary.items()}) if current_week_summary else SimpleNamespace()
        previous_plan = SimpleNamespace(**{k: v or 0 for k, v in previous_week_summary.items()}) if previous_week_summary else None

        level_wise_data = self._calculate_level_wise_hiring(current_plan, previous_plan)

        # --- 4. Format and Return Response ---
        response_data = {
            "success": True,
            "level_wise_hiring_graph": level_wise_data
        }
        return Response(response_data)

    def _get_detailed_summary_for_week(self, start_date, filters):
        """Helper to get a detailed weekly summary including all L1-L4 fields."""
        week_filters = filters.copy()
        end_date = start_date + timedelta(days=6)
        week_filters['date__range'] = [start_date, end_date]
        
        # This reuses the same aggregation logic as our detailed month summary
        aggregation = DailyProductionData.objects.filter(**week_filters).aggregate(
            attrition_rate=Avg('attrition_rate'),
            total_production_plan=Sum('total_production_plan'),
            total_operators_required_plan=Sum('total_operators_required_plan'),
            total_operators_required_actual=Sum('total_operators_required_actual'),
            ctq_plan_l1=Sum('ctq_plan_l1'), ctq_plan_l2=Sum('ctq_plan_l2'), ctq_plan_l3=Sum('ctq_plan_l3'), ctq_plan_l4=Sum('ctq_plan_l4'),
            pdi_plan_l1=Sum('pdi_plan_l1'), pdi_plan_l2=Sum('pdi_plan_l2'), pdi_plan_l3=Sum('pdi_plan_l3'), pdi_plan_l4=Sum('pdi_plan_l4'),
            other_plan_l1=Sum('other_plan_l1'), other_plan_l2=Sum('other_plan_l2'), other_plan_l3=Sum('other_plan_l3'), other_plan_l4=Sum('other_plan_l4'),
            ctq_actual_l1=Sum('ctq_actual_l1'), ctq_actual_l2=Sum('ctq_actual_l2'), ctq_actual_l3=Sum('ctq_actual_l3'), ctq_actual_l4=Sum('ctq_actual_l4'),
            pdi_actual_l1=Sum('pdi_actual_l1'), pdi_actual_l2=Sum('pdi_actual_l2'), pdi_actual_l3=Sum('pdi_actual_l3'), pdi_actual_l4=Sum('pdi_actual_l4'),
            other_actual_l1=Sum('other_actual_l1'), other_actual_l2=Sum('other_actual_l2'), other_actual_l3=Sum('other_actual_l3'), other_actual_l4=Sum('other_actual_l4'),
        )
        if aggregation.get('total_production_plan') is None:
            return None
        return aggregation



from django.db.models import Sum  # Sum is already imported earlier; safe if duplicate in file scope
from django.http import JsonResponse

def _filter_plans_by_scope(request):
    month = request.GET.get('month')
    year = request.GET.get('year')
    station_id = request.GET.get('station_id')
    line_id = request.GET.get('line_id')
    shop_floor_id = request.GET.get('shop_floor_id')
    factory_id = request.GET.get('factory_id')

    qs = ProductionPlan.objects.all()
    if month:
        qs = qs.filter(month=month)
    if year:
        qs = qs.filter(year=year)

    if station_id:
        qs = qs.filter(station_id=station_id)
    elif line_id:
        qs = qs.filter(line_id=line_id)
    elif shop_floor_id:
        qs = qs.filter(shop_floor_id=shop_floor_id)
    elif factory_id:
        qs = qs.filter(factory_id=factory_id)

    return qs

def _aggregate_plan(qs):
    agg = qs.aggregate(
        # CTQ
        ctq_plan_l1=Sum('ctq_plan_l1'), ctq_plan_l2=Sum('ctq_plan_l2'), ctq_plan_l3=Sum('ctq_plan_l3'), ctq_plan_l4=Sum('ctq_plan_l4'),
        ctq_actual_l1=Sum('ctq_actual_l1'), ctq_actual_l2=Sum('ctq_actual_l2'), ctq_actual_l3=Sum('ctq_actual_l3'), ctq_actual_l4=Sum('ctq_actual_l4'),
        ctq_plan_total=Sum('ctq_plan_total'), ctq_actual_total=Sum('ctq_actual_total'),
        # PDI
        pdi_plan_l1=Sum('pdi_plan_l1'), pdi_plan_l2=Sum('pdi_plan_l2'), pdi_plan_l3=Sum('pdi_plan_l3'), pdi_plan_l4=Sum('pdi_plan_l4'),
        pdi_actual_l1=Sum('pdi_actual_l1'), pdi_actual_l2=Sum('pdi_actual_l2'), pdi_actual_l3=Sum('pdi_actual_l3'), pdi_actual_l4=Sum('pdi_actual_l4'),
        pdi_plan_total=Sum('pdi_plan_total'), pdi_actual_total=Sum('pdi_actual_total'),
        # OTHER
        other_plan_l1=Sum('other_plan_l1'), other_plan_l2=Sum('other_plan_l2'), other_plan_l3=Sum('other_plan_l3'), other_plan_l4=Sum('other_plan_l4'),
        other_actual_l1=Sum('other_actual_l1'), other_actual_l2=Sum('other_actual_l2'), other_actual_l3=Sum('other_actual_l3'), other_actual_l4=Sum('other_actual_l4'),
        other_plan_total=Sum('other_plan_total'), other_actual_total=Sum('other_actual_total'),
    )
    return {k: (v or 0) for k, v in agg.items()}

def get_plan_data(request):
    data_type = request.GET.get('type')  # 'ctq' | 'pdi' | 'other' | 'all'

    qs = _filter_plans_by_scope(request)
    if not qs.exists():
        return JsonResponse({'error': 'No data found'}, status=404)

    if request.GET.get('station_id'):
        plan = qs.order_by('-year', '-created_at').first()
        agg = {
            # CTQ
            'ctq_plan_l1': plan.ctq_plan_l1, 'ctq_plan_l2': plan.ctq_plan_l2, 'ctq_plan_l3': plan.ctq_plan_l3, 'ctq_plan_l4': plan.ctq_plan_l4,
            'ctq_actual_l1': plan.ctq_actual_l1, 'ctq_actual_l2': plan.ctq_actual_l2, 'ctq_actual_l3': plan.ctq_actual_l3, 'ctq_actual_l4': plan.ctq_actual_l4,
            'ctq_plan_total': plan.ctq_plan_total, 'ctq_actual_total': plan.ctq_actual_total,
            # PDI
            'pdi_plan_l1': plan.pdi_plan_l1, 'pdi_plan_l2': plan.pdi_plan_l2, 'pdi_plan_l3': plan.pdi_plan_l3, 'pdi_plan_l4': plan.pdi_plan_l4,
            'pdi_actual_l1': plan.pdi_actual_l1, 'pdi_actual_l2': plan.pdi_actual_l2, 'pdi_actual_l3': plan.pdi_actual_l3, 'pdi_actual_l4': plan.pdi_actual_l4,
            'pdi_plan_total': plan.pdi_plan_total, 'pdi_actual_total': plan.pdi_actual_total,
            # OTHER
            'other_plan_l1': plan.other_plan_l1, 'other_plan_l2': plan.other_plan_l2, 'other_plan_l3': plan.other_plan_l3, 'other_plan_l4': plan.other_plan_l4,
            'other_actual_l1': plan.other_actual_l1, 'other_actual_l2': plan.other_actual_l2, 'other_actual_l3': plan.other_actual_l3, 'other_actual_l4': plan.other_actual_l4,
            'other_plan_total': plan.other_plan_total, 'other_actual_total': plan.other_actual_total,
        }
    else:
        agg = _aggregate_plan(qs)

    if data_type == 'ctq':
        data = {
            'ctq_plan_l1': agg['ctq_plan_l1'],
            'ctq_plan_l2': agg['ctq_plan_l2'],
            'ctq_plan_l3': agg['ctq_plan_l3'],
            'ctq_plan_l4': agg['ctq_plan_l4'],
            'ctq_actual_l1': agg['ctq_actual_l1'],
            'ctq_actual_l2': agg['ctq_actual_l2'],
            'ctq_actual_l3': agg['ctq_actual_l3'],
            'ctq_actual_l4': agg['ctq_actual_l4'],
            'ctq_plan_total': agg['ctq_plan_total'],
            'ctq_actual_total': agg['ctq_actual_total'],
        }
    elif data_type == 'pdi':
        data = {
            'pdi_plan_l1': agg['pdi_plan_l1'],
            'pdi_plan_l2': agg['pdi_plan_l2'],
            'pdi_plan_l3': agg['pdi_plan_l3'],
            'pdi_plan_l4': agg['pdi_plan_l4'],
            'pdi_actual_l1': agg['pdi_actual_l1'],
            'pdi_actual_l2': agg['pdi_actual_l2'],
            'pdi_actual_l3': agg['pdi_actual_l3'],
            'pdi_actual_l4': agg['pdi_actual_l4'],
            'pdi_plan_total': agg['pdi_plan_total'],
            'pdi_actual_total': agg['pdi_actual_total'],
        }
    elif data_type == 'other':
        data = {
            'other_plan_l1': agg['other_plan_l1'],
            'other_plan_l2': agg['other_plan_l2'],
            'other_plan_l3': agg['other_plan_l3'],
            'other_plan_l4': agg['other_plan_l4'],
            'other_actual_l1': agg['other_actual_l1'],
            'other_actual_l2': agg['other_actual_l2'],
            'other_actual_l3': agg['other_actual_l3'],
            'other_actual_l4': agg['other_actual_l4'],
            'other_plan_total': agg['other_plan_total'],
            'other_actual_total': agg['other_actual_total'],
        }
    elif data_type == 'all':
        data = {
            'total_plan_l1': (agg['ctq_plan_l1'] or 0) + (agg['pdi_plan_l1'] or 0) + (agg['other_plan_l1'] or 0),
            'total_plan_l2': (agg['ctq_plan_l2'] or 0) + (agg['pdi_plan_l2'] or 0) + (agg['other_plan_l2'] or 0),
            'total_plan_l3': (agg['ctq_plan_l3'] or 0) + (agg['pdi_plan_l3'] or 0) + (agg['other_plan_l3'] or 0),
            'total_plan_l4': (agg['ctq_plan_l4'] or 0) + (agg['pdi_plan_l4'] or 0) + (agg['other_plan_l4'] or 0),
            'total_actual_l1': (agg['ctq_actual_l1'] or 0) + (agg['pdi_actual_l1'] or 0) + (agg['other_actual_l1'] or 0),
            'total_actual_l2': (agg['ctq_actual_l2'] or 0) + (agg['pdi_actual_l2'] or 0) + (agg['other_actual_l2'] or 0),
            'total_actual_l3': (agg['ctq_actual_l3'] or 0) + (agg['pdi_actual_l3'] or 0) + (agg['other_actual_l3'] or 0),
            'total_actual_l4': (agg['ctq_actual_l4'] or 0) + (agg['pdi_actual_l4'] or 0) + (agg['other_actual_l4'] or 0),
            'total_plan': (agg['ctq_plan_total'] or 0) + (agg['pdi_plan_total'] or 0) + (agg['other_plan_total'] or 0),
            'total_actual': (agg['ctq_actual_total'] or 0) + (agg['pdi_actual_total'] or 0) + (agg['other_actual_total'] or 0),
        }
    else:
        return JsonResponse({'error': 'Invalid type'}, status=400)

    return JsonResponse(data)

def get_plan_totals(request):
    qs = _filter_plans_by_scope(request)
    if not qs.exists():
        return JsonResponse({'error': 'No data found'}, status=404)

    if request.GET.get('station_id'):
        plan = qs.order_by('-year', '-created_at').first()
        data = {
            'total_l1': ((plan.ctq_plan_l1 or 0) + (plan.pdi_plan_l1 or 0) + (plan.other_plan_l1 or 0)) + ((plan.ctq_actual_l1 or 0) + (plan.pdi_actual_l1 or 0) + (plan.other_actual_l1 or 0)),
            'total_l2': ((plan.ctq_plan_l2 or 0) + (plan.pdi_plan_l2 or 0) + (plan.other_plan_l2 or 0)) + ((plan.ctq_actual_l2 or 0) + (plan.pdi_actual_l2 or 0) + (plan.other_actual_l2 or 0)),
            'total_l3': ((plan.ctq_plan_l3 or 0) + (plan.pdi_plan_l3 or 0) + (plan.other_plan_l3 or 0)) + ((plan.ctq_actual_l3 or 0) + (plan.pdi_actual_l3 or 0) + (plan.other_actual_l3 or 0)),
            'total_l4': ((plan.ctq_plan_l4 or 0) + (plan.pdi_plan_l4 or 0) + (plan.other_plan_l4 or 0)) + ((plan.ctq_actual_l4 or 0) + (plan.pdi_actual_l4 or 0) + (plan.other_actual_l4 or 0)),
            'total_plan': (plan.ctq_plan_total or 0) + (plan.pdi_plan_total or 0) + (plan.other_plan_total or 0),
            'total_actual': (plan.ctq_actual_total or 0) + (plan.pdi_actual_total or 0) + (plan.other_actual_total or 0),
        }
    else:
        agg = _aggregate_plan(qs)
        data = {
            'total_l1': ((agg['ctq_plan_l1'] or 0) + (agg['pdi_plan_l1'] or 0) + (agg['other_plan_l1'] or 0)) + ((agg['ctq_actual_l1'] or 0) + (agg['pdi_actual_l1'] or 0) + (agg['other_actual_l1'] or 0)),
            'total_l2': ((agg['ctq_plan_l2'] or 0) + (agg['pdi_plan_l2'] or 0) + (agg['other_plan_l2'] or 0)) + ((agg['ctq_actual_l2'] or 0) + (agg['pdi_actual_l2'] or 0) + (agg['other_actual_l2'] or 0)),
            'total_l3': ((agg['ctq_plan_l3'] or 0) + (agg['pdi_plan_l3'] or 0) + (agg['other_plan_l3'] or 0)) + ((agg['ctq_actual_l3'] or 0) + (agg['pdi_actual_l3'] or 0) + (agg['other_actual_l3'] or 0)),
            'total_l4': ((agg['ctq_plan_l4'] or 0) + (agg['pdi_plan_l4'] or 0) + (agg['other_plan_l4'] or 0)) + ((agg['ctq_actual_l4'] or 0) + (agg['pdi_actual_l4'] or 0) + (agg['other_actual_l4'] or 0)),
            'total_plan': (agg['ctq_plan_total'] or 0) + (agg['pdi_plan_total'] or 0) + (agg['other_plan_total'] or 0),
            'total_actual': (agg['ctq_actual_total'] or 0) + (agg['pdi_actual_total'] or 0) + (agg['other_actual_total'] or 0),
        }

    return JsonResponse(data)

def get_plan_other(request):
    qs = _filter_plans_by_scope(request)
    if not qs.exists():
        return JsonResponse({'error': 'No data found'}, status=404)

    if request.GET.get('station_id'):
        plan = qs.order_by('-year', '-created_at').first()
        data = {
            'other_plan_l1': plan.other_plan_l1,
            'other_plan_l2': plan.other_plan_l2,
            'other_plan_l3': plan.other_plan_l3,
            'other_plan_l4': plan.other_plan_l4,
            'other_actual_l1': plan.other_actual_l1,
            'other_actual_l2': plan.other_actual_l2,
            'other_actual_l3': plan.other_actual_l3,
            'other_actual_l4': plan.other_actual_l4,
            'other_plan_total': plan.other_plan_total,
            'other_actual_total': plan.other_actual_total,
        }
    else:
        agg = _aggregate_plan(qs)
        data = {
            'other_plan_l1': agg['other_plan_l1'],
            'other_plan_l2': agg['other_plan_l2'],
            'other_plan_l3': agg['other_plan_l3'],
            'other_plan_l4': agg['other_plan_l4'],
            'other_actual_l1': agg['other_actual_l1'],
            'other_actual_l2': agg['other_actual_l2'],
            'other_actual_l3': agg['other_actual_l3'],
            'other_actual_l4': agg['other_actual_l4'],
            'other_plan_total': agg['other_plan_total'],
            'other_actual_total': agg['other_actual_total'],
        }

    return JsonResponse(data)





from rest_framework import viewsets
from .models import NewFactory, NewDepartment, NewLine, NewWorkstation
from .serializers import (
    NewFactorySerializer,
    NewDepartmentSerializer,
    NewLineSerializer,
    NewWorkstationSerializer
)

class NewFactoryViewSet(viewsets.ModelViewSet):
    queryset = NewFactory.objects.all()
    serializer_class = NewFactorySerializer

class NewDepartmentViewSet(viewsets.ModelViewSet):
    queryset = NewDepartment.objects.all()
    serializer_class = NewDepartmentSerializer

class NewLineViewSet(viewsets.ModelViewSet):
    queryset = NewLine.objects.all()
    serializer_class = NewLineSerializer

class NewWorkstationViewSet(viewsets.ModelViewSet):
    queryset = NewWorkstation.objects.all()
    serializer_class = NewWorkstationSerializer










from django.http import HttpResponse, JsonResponse
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch
from io import BytesIO
import json
import traceback
from .models import (
    LevelThreeQATraineeInfo, LevelThreeQATrainingTopic, LevelThreeQAOJTDay, 
    LevelThreeQAOJTScore, LevelThreeQualityLine
)

@method_decorator(csrf_exempt, name='dispatch')
class LevelThreeQAReportPDFView(View):
    def post(self, request, *args, **kwargs):
        """
        Handle PDF generation requests for Level Three QA reports
        """
        try:
            print("\n=== Received Level 3 QA PDF generation request ===")
            
            # 1. Parse input data
            trainee_id = self._get_trainee_id(request)
            line_id = self._get_line_id(request)
            
            if not trainee_id:
                return JsonResponse({'error': 'trainee_id is required'}, status=400)
            if not line_id:
                return JsonResponse({'error': 'line_id is required'}, status=400)
            
            print(f"Processing trainee_id: {trainee_id}, line_id: {line_id}")

            # 2. Get trainee record
            trainee = self._get_trainee_record(trainee_id, line_id)
            if isinstance(trainee, JsonResponse):
                return trainee
            if not trainee:
                return JsonResponse({'error': 'Trainee not found for the specified trainee_id and line_id'}, status=404)

            print(f"Found trainee: {trainee.trainee_name}")

            # 3. Generate PDF content
            print("Generating Level 3 QA PDF content...")
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), 
                                  leftMargin=0.5*inch, rightMargin=0.5*inch,
                                  topMargin=0.5*inch, bottomMargin=0.5*inch)
            story = self.create_qa_pdf_content(trainee)
            
            # 4. Build PDF document
            print("Building Level 3 QA PDF document...")
            doc.build(story)
            buffer.seek(0)
            print("Level 3 QA PDF generation completed successfully")

            # 5. Return PDF response
            response = HttpResponse(
                buffer.getvalue(), 
                content_type='application/pdf'
            )
            response['Content-Disposition'] = (
                f'attachment; filename="level3_qa_report_{trainee_id}_quality.pdf"'
            )
            return response
            
        except Exception as e:
            print("\n!!! Level 3 QA PDF generation failed !!!")
            traceback.print_exc()
            return JsonResponse(
                {
                    'error': 'Internal server error',
                    'detail': str(e),
                    'traceback': traceback.format_exc()
                }, 
                status=500
            )

    def _get_trainee_id(self, request):
        """Helper method to extract trainee_id from request"""
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                return data.get('trainee_id')
            except json.JSONDecodeError:
                return None
        return request.POST.get('trainee_id')

    def _get_line_id(self, request):
        """Helper method to extract line_id from request"""
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                return data.get('line_id')
            except json.JSONDecodeError:
                return None
        return request.POST.get('line_id')

    def _get_trainee_record(self, trainee_id, line_id):
        """Get trainee record for Level 3 quality"""
        try:
            return LevelThreeQATraineeInfo.objects.get(traineeId=trainee_id, line_id=line_id)
        except LevelThreeQATraineeInfo.DoesNotExist:
            return None
        except LevelThreeQATraineeInfo.MultipleObjectsReturned:
            try:
                return LevelThreeQATraineeInfo.objects.filter(traineeId=trainee_id, line_id=line_id).order_by('-id').first()
            except Exception as e:
                return JsonResponse(
                    {'error': f'Multiple trainee records found for trainee_id {trainee_id} and line_id {line_id}, and fallback selection failed: {str(e)}'},
                    status=400
                )

    def create_qa_pdf_content(self, trainee):
        """Generate the Level 3 QA PDF content structure for quality"""
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = "Level-3 QA MONITORING SHEET - QUALITY"
        story.append(Paragraph(title, styles['Title']))
        story.append(Spacer(1, 12))
        
        # Add all sections
        self._add_trainee_info(story, styles, trainee)
        self._add_qa_monitoring_table(story, styles, trainee)
        self._add_judgement_criteria(story, styles)
        
        return story

    def _add_trainee_info(self, story, styles, trainee):
        """Add trainee basic information for Level 3 quality"""
        # Get line_name from the ForeignKey
        line_name = trainee.line.name if trainee.line else "Not Assigned"

        info_data = [
            ["Trainee Name:", trainee.trainee_name, "Trainee Id:", trainee.traineeId],
            ["Line Name:", line_name, "", ""],
            ["Trainer Name:", trainee.trainer_name, "Training Status:", trainee.training_status]
        ]
        
        info_table = Table(info_data, colWidths=[100, 200, 100, 200])
        info_style = TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('ALIGN', (0,0), (0,-1), 'LEFT'),
            ('ALIGN', (2,0), (2,-1), 'LEFT'),
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ])
        info_table.setStyle(info_style)
        story.append(info_table)
        story.append(Spacer(1, 16))

    def _add_qa_monitoring_table(self, story, styles, trainee):
        """Add the main QA monitoring table for Level 3 quality"""
        topics = LevelThreeQATrainingTopic.objects.all().order_by('sl_no')
        days = LevelThreeQAOJTDay.objects.all().order_by('id')
        scores = LevelThreeQAOJTScore.objects.filter(trainee=trainee)

        header = ["S.NO.", "TRAINING TOPIC", "Date"]
        for day in days:
            header.append(day.name)
        
        table_data = [header]
        
        for topic in topics:
            row = [str(topic.sl_no), topic.topic, topic.date]
            for day in days:
                score_obj = scores.filter(topic=topic, day=day).first()
                score_display = str(score_obj.score) if score_obj else ""
                row.append(score_display)
            table_data.append(row)

        base_widths = [40, 300, 80]
        day_width = (720 - sum(base_widths)) / len(days) if days else 60
        col_widths = base_widths + [day_width] * len(days)

        monitoring_table = Table(table_data, colWidths=col_widths)
        monitoring_style = self._get_monitoring_table_style(len(topics), len(days))
        monitoring_table.setStyle(monitoring_style)
        
        story.append(monitoring_table)
        story.append(Spacer(1, 16))

    def _get_monitoring_table_style(self, topic_count, day_count):
        """Get styling for the monitoring table"""
        style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4682B4')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('TOPPADDING', (0,0), (-1,0), 8),
            ('ALIGN', (0,1), (0,-1), 'CENTER'),
            ('ALIGN', (1,1), (1,-1), 'LEFT'),
            ('ALIGN', (2,1), (2,-1), 'CENTER'),
            ('ALIGN', (3,1), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F8F8')]),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ])
        
        return style

    def _add_judgement_criteria(self, story, styles):
        """Add judgement criteria section for Level 3"""
        story.append(Paragraph("Judgement Criteria", styles['Heading3']))
        story.append(Spacer(1, 8))
        
        criteria_data = [
            ["Criteria", "Description"],
            ["0-10 Marks", "Score based on trainee performance"],
            ["Pass", "Achieves 100% on final assessment"],
            ["Fail", "Below 100% on final assessment"]
        ]
        
        criteria_table = Table(criteria_data, colWidths=[100, 400])
        criteria_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4682B4')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 9),
            ('ALIGN', (0,1), (0,-1), 'CENTER'),
            ('ALIGN', (1,1), (1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('LEFTPADDING', (0,0), (-1,-1), 6),
            ('RIGHTPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ])
        criteria_table.setStyle(criteria_style)
        story.append(criteria_table)
        story.append(Spacer(1, 16))

        assessment_text = """
        <b>Final Assessment should be 100%</b><br/>
        <b>Passing Criteria: 100%</b><br/>
        <i>If failed in evaluation re-training is required</i>
        """
        story.append(Paragraph(assessment_text, styles['Normal']))









from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny, BasePermission
from django.db.models import Q, Count
from django.utils import timezone
from datetime import timedelta

from .models import Notification, EmployeeMaster
from .serializers import (
    NotificationSerializer, NotificationCreateSerializer,
    NotificationUpdateSerializer, NotificationStatsSerializer
)


class NotificationPermission(BasePermission):
    """
    Custom permission for notifications:
    - Allow read access without authentication (for testing)
    - Require authentication for write operations
    """
    def has_permission(self, request, view):
        if request.method in ['GET', 'HEAD', 'OPTIONS']:
            return True
        return request.user and request.user.is_authenticated


class NotificationViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing notifications - no authentication required
    """
    serializer_class = NotificationSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        return Notification.objects.all().select_related('recipient', 'operator', 'level', 'training_schedule')

    def get_serializer_class(self):
        if self.action == 'create':
            return NotificationCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return NotificationUpdateSerializer
        return NotificationSerializer

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        # Filters
        is_read = request.query_params.get('is_read')
        if is_read is not None:
            queryset = queryset.filter(is_read=is_read.lower() == 'true')

        notification_types = request.query_params.getlist('notification_type')
        if notification_types:
            queryset = queryset.filter(notification_type__in=notification_types)

        notification_type = request.query_params.get('type')
        if notification_type:
            queryset = queryset.filter(notification_type=notification_type)

        priority = request.query_params.get('priority')
        if priority:
            queryset = queryset.filter(priority=priority)

        days = request.query_params.get('days')
        if days:
            try:
                days_int = int(days)
                since_date = timezone.now() - timedelta(days=days_int)
                queryset = queryset.filter(created_at__gte=since_date)
            except ValueError:
                pass

        queryset = queryset.order_by('-created_at')

        # Pagination
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        notification = self.get_object()
        notification.mark_as_read()
        serializer = self.get_serializer(notification)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def mark_unread(self, request, pk=None):
        notification = self.get_object()
        notification.mark_as_unread()
        serializer = self.get_serializer(notification)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        queryset = self.get_queryset().filter(is_read=False)
        count = queryset.count()
        for notification in queryset:
            notification.mark_as_read()
        return Response({'message': f'Marked {count} notifications as read', 'count': count})

    @action(detail=False, methods=['get'])
    def unread(self, request):
        queryset = self.get_queryset().filter(is_read=False)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        queryset = self.get_queryset()
        total_count = queryset.count()
        unread_count = queryset.filter(is_read=False).count()
        read_count = total_count - unread_count
        recent_count = queryset.filter(
            created_at__gte=timezone.now() - timedelta(hours=24)
        ).count()

        by_type = dict(queryset.values('notification_type').annotate(
            count=Count('id')
        ).values_list('notification_type', 'count'))

        by_priority = dict(queryset.values('priority').annotate(
            count=Count('id')
        ).values_list('priority', 'count'))

        stats_data = {
            'total_count': total_count,
            'unread_count': unread_count,
            'read_count': read_count,
            'recent_count': recent_count,
            'by_type': by_type,
            'by_priority': by_priority
        }

        serializer = NotificationStatsSerializer(stats_data)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def recent(self, request):
        since_date = timezone.now() - timedelta(hours=24)
        queryset = self.get_queryset().filter(created_at__gte=since_date)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


@api_view(['POST'])
def trigger_employee_notification(request):
    """Manually trigger an employee registration notification for testing"""
    from django.contrib.auth import get_user_model
    User = get_user_model()

    recipient = request.user if request.user.is_authenticated else User.objects.first()
    if not recipient:
        return Response({'error': 'No users found'}, status=status.HTTP_400_BAD_REQUEST)

    latest_employee = EmployeeMaster.objects.last()
    if not latest_employee:
        return Response({'error': 'No employees found'}, status=status.HTTP_400_BAD_REQUEST)

    notification = Notification.objects.create(
        title="New Employee Registered",
        message=f"New employee {latest_employee.name} (Pay Code: {latest_employee.pay_code}) has been registered.",
        notification_type='employee_registration',
        recipient=recipient,
        operator=latest_employee,
        priority='medium',
        metadata={
            'pay_code': latest_employee.pay_code,
            'department': latest_employee.department,
            'designation': latest_employee.desig_category
        }
    )

    serializer = NotificationSerializer(notification)
    return Response({
        'message': 'Employee registration notification created successfully',
        'notification': serializer.data,
        'employee': latest_employee.name
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
def trigger_all_notification_types(request):
    """Create test notifications for all 7 notification types"""
    from django.contrib.auth import get_user_model
    User = get_user_model()

    recipient = request.user if request.user.is_authenticated else User.objects.first()
    if not recipient:
        return Response({'error': 'No users found'}, status=status.HTTP_400_BAD_REQUEST)

    latest_employee = EmployeeMaster.objects.last()
    employee_name = latest_employee.name if latest_employee else "Test Employee"

    notification_types = [
        {'type': 'employee_registration', 'title': 'New Employee Registered',
         'message': f'{employee_name} has been registered in the system.', 'priority': 'medium'},
        {'type': 'level_exam_completed', 'title': 'Level Exam Completed',
         'message': f'{employee_name} has completed Level 2 evaluation with status: Pass.', 'priority': 'high'},
        {'type': 'training_reschedule', 'title': 'Training Rescheduled',
         'message': f'Training session for {employee_name} has been rescheduled.', 'priority': 'medium'},
        {'type': 'refresher_training_scheduled', 'title': 'Refresher Training Scheduled',
         'message': f'Refresher training has been scheduled for {employee_name}.', 'priority': 'medium'},
        {'type': 'refresher_training_completed', 'title': 'Refresher Training Completed',
         'message': f'{employee_name} has completed refresher training successfully.', 'priority': 'medium'},
        {'type': 'bending_training_added', 'title': 'Bending Training Added',
         'message': 'New bending training module has been added.', 'priority': 'low'},
        {'type': 'system_alert', 'title': 'System Alert',
         'message': 'System maintenance scheduled for tonight at 2:00 AM.', 'priority': 'urgent'},
    ]

    created_notifications = []
    for notif_data in notification_types:
        n = Notification.objects.create(
            title=notif_data['title'],
            message=notif_data['message'],
            notification_type=notif_data['type'],
            recipient=recipient,
            operator=latest_employee,
            priority=notif_data['priority'],
            metadata={'test': True, 'notification_type': notif_data['type']}
        )
        created_notifications.append(n)

    return Response({
        'message': f'Created {len(created_notifications)} notifications',
        'types': [n.notification_type for n in created_notifications],
        'count': len(created_notifications)
    }, status=status.HTTP_201_CREATED)






















# ==================== NOTIFICATION API VIEWS ====================

from rest_framework import viewsets, status, serializers
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, BasePermission
from django.db.models import Q, Count
from django.utils import timezone
from datetime import timedelta
from .models import Notification
from .serializers import (
    NotificationSerializer, NotificationCreateSerializer,
    NotificationUpdateSerializer, NotificationStatsSerializer
)


class NotificationPermission(BasePermission):
    """
    Custom permission for notifications:
    - Allow read access without authentication (for testing)
    - Require authentication for write operations
    """
    def has_permission(self, request, view):
        if request.method in ['GET', 'HEAD', 'OPTIONS']:
            return True
        return request.user and request.user.is_authenticated


class NotificationViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing notifications - no authentication required
    """
    serializer_class = NotificationSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        return Notification.objects.all().select_related('recipient', 'operator', 'level', 'training_schedule')

    def get_serializer_class(self):
        if self.action == 'create':
            return NotificationCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return NotificationUpdateSerializer
        return NotificationSerializer

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        is_read = request.query_params.get('is_read')
        if is_read is not None:
            queryset = queryset.filter(is_read=is_read.lower() == 'true')

        notification_types = request.query_params.getlist('notification_type')
        if notification_types:
            queryset = queryset.filter(notification_type__in=notification_types)

        notification_type = request.query_params.get('type')
        if notification_type:
            queryset = queryset.filter(notification_type=notification_type)

        priority = request.query_params.get('priority')
        if priority:
            queryset = queryset.filter(priority=priority)

        days = request.query_params.get('days')
        if days:
            try:
                days_int = int(days)
                since_date = timezone.now() - timedelta(days=days_int)
                queryset = queryset.filter(created_at__gte=since_date)
            except ValueError:
                pass

        queryset = queryset.order_by('-created_at')

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        notification = self.get_object()
        notification.mark_as_read()
        serializer = self.get_serializer(notification)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def mark_unread(self, request, pk=None):
        notification = self.get_object()
        notification.mark_as_unread()
        serializer = self.get_serializer(notification)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        queryset = self.get_queryset().filter(is_read=False)
        count = queryset.count()
        for notification in queryset:
            notification.mark_as_read()
        return Response({'message': f'Marked {count} notifications as read', 'count': count})

    @action(detail=False, methods=['get'])
    def unread(self, request):
        queryset = self.get_queryset().filter(is_read=False)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        queryset = self.get_queryset()
        total_count = queryset.count()
        unread_count = queryset.filter(is_read=False).count()
        read_count = total_count - unread_count
        recent_count = queryset.filter(created_at__gte=timezone.now() - timedelta(hours=24)).count()
        by_type = dict(queryset.values('notification_type').annotate(count=Count('id')).values_list('notification_type', 'count'))
        by_priority = dict(queryset.values('priority').annotate(count=Count('id')).values_list('priority', 'count'))

        stats_data = {
            'total_count': total_count,
            'unread_count': unread_count,
            'read_count': read_count,
            'recent_count': recent_count,
            'by_type': by_type,
            'by_priority': by_priority
        }
        serializer = NotificationStatsSerializer(stats_data)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def recent(self, request):
        since_date = timezone.now() - timedelta(hours=24)
        queryset = self.get_queryset().filter(created_at__gte=since_date)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


# ==================== API FUNCTION VIEWS ====================

@api_view(['GET'])
def notification_count(request):
    if not request.user.is_authenticated:
        return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)

    count = Notification.objects.filter(
        Q(recipient=request.user) | Q(recipient_email=request.user.email),
        is_read=False
    ).count()
    return Response({'unread_count': count})


@api_view(['GET'])
def test_notifications(request):
    notifications = Notification.objects.all()[:10]
    serializer = NotificationSerializer(notifications, many=True)
    return Response({'count': notifications.count(), 'notifications': serializer.data, 'debug': 'This is a test endpoint'})


@api_view(['POST'])
def create_system_notification(request):
    if not request.user.is_authenticated:
        return Response({'error': 'Authentication required'}, status=status.HTTP_403_FORBIDDEN)

    serializer = NotificationCreateSerializer(data=request.data)
    if serializer.is_valid():
        notification = serializer.save()
        response_serializer = NotificationSerializer(notification)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def create_test_notification(request):
    try:
        recipient = request.user if request.user.is_authenticated else None
        notification = Notification.objects.create(
            title="Test Notification",
            message="This is a test notification to verify the system is working.",
            notification_type="system_alert",
            recipient=recipient,
            priority="medium",
            metadata={"test": True}
        )
        serializer = NotificationSerializer(notification)
        return Response({'message': 'Test notification created successfully', 'notification': serializer.data}, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def trigger_employee_notification(request):
    try:
        recipient = request.user if request.user.is_authenticated else None
        latest_employee = EmployeeMaster.objects.last()

        if not latest_employee:
            return Response({'error': 'No employees found in the system'}, status=status.HTTP_400_BAD_REQUEST)

        notification = Notification.objects.create(
            title="New Employee Registered",
            message=f"New employee {latest_employee.name} (Pay Code: {latest_employee.pay_code}) has been registered.",
            notification_type='employee_registration',
            recipient=recipient,
            priority='medium',
            metadata={
                'pay_code': latest_employee.pay_code,
                'department': latest_employee.department,
                'section': latest_employee.section
            }
        )
        serializer = NotificationSerializer(notification)
        return Response({'message': 'Employee notification created', 'notification': serializer.data}, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def trigger_all_notification_types(request):
    try:
        recipient = request.user if request.user.is_authenticated else None
        latest_employee = EmployeeMaster.objects.last()
        employee_name = latest_employee.name if latest_employee else "Test Employee"

        notification_types = [
            {'type': 'employee_registration', 'title': 'New Employee Registered', 'message': f'{employee_name} has been registered.', 'priority': 'medium'},
            {'type': 'level_exam_completed', 'title': 'Level Exam Completed', 'message': f'{employee_name} completed Level 2 evaluation.', 'priority': 'high'},
            {'type': 'training_reschedule', 'title': 'Training Rescheduled', 'message': f'Training for {employee_name} rescheduled.', 'priority': 'medium'},
            {'type': 'refresher_training_scheduled', 'title': 'Refresher Training Scheduled', 'message': f'Refresher training scheduled for {employee_name}.', 'priority': 'medium'},
            {'type': 'refresher_training_completed', 'title': 'Refresher Training Completed', 'message': f'{employee_name} completed refresher training.', 'priority': 'medium'},
            {'type': 'bending_training_added', 'title': 'Bending Training Added', 'message': 'New bending training module added.', 'priority': 'low'},
            {'type': 'system_alert', 'title': 'System Alert', 'message': 'System maintenance at 2:00 AM.', 'priority': 'urgent'}
        ]

        created_notifications = []
        for notif_data in notification_types:
            n = Notification.objects.create(
                title=notif_data['title'],
                message=notif_data['message'],
                notification_type=notif_data['type'],
                recipient=recipient,
                priority=notif_data['priority'],
                metadata={'test': True, 'employee': employee_name}
            )
            created_notifications.append(n)

        return Response({'message': f'Created {len(created_notifications)} notifications', 'count': len(created_notifications)}, status=status.HTTP_201_CREATED)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
